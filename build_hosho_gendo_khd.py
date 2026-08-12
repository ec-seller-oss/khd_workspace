# 信用保証協会 翌期借入余力 試算EXCEL(KHD版)を生成する
# 元ネタ: 伊東さん共有「特典：信用保証協会から翌期いくら借りられるか試算するEXCEL」Version.2.1
# 計算式は元Excelと同一:
#   ① 自己資本比率 = B/A×100
#   ② 債務償還年数 = C/(D+E)
#   ③ 自己資本比率から見た借入上限 = B/0.25
#   ④ 債務償還年数から見た借入上限 = (D+E)×20
#   ⑤ 翌期の借入上限 = (③+④)/2
#   ⑥ 翌期の借りれ枠 = ⑤-C (マイナスはNG)
# 実行: python3 build_hosho_gendo_khd.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = "保証協会_翌期借入余力_KHD版.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "試算(KHD)"

YELLOW = PatternFill("solid", fgColor="FFF2CC")   # 入力セル
BEIGE = PatternFill("solid", fgColor="EFE6D5")    # 見出し
GRAY = PatternFill("solid", fgColor="F2F2F2")
thin = Side(style="thin", color="BBBBBB")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
BOLD = Font(bold=True)
RED = Font(bold=True, color="C00000")

ws["B2"] = "信用保証協会から翌期いくら借りられるか 試算EXCEL — KHD版"
ws["B2"].font = Font(bold=True, size=14)
ws["H2"] = "元: Version.2.1(伊東さん共有) / 単位:万円"

ROWS = [  # (区分, 記号, 項目, 補足)
    ("BS", "A", "総資産(資産は合計いくらあるか?)", ""),
    ("BS", "B", "純資産(借入を除いた資産はいくらあるか?)", "⇒役員借入金を含む"),
    ("BS", "C", "長期借入金", "⇒役員借入金を除く"),
    ("PL", "D", "減価償却", ""),
    ("PL", "E", "税引前当期純利益", ""),
]
CALC = [  # (記号, 項目, 式template, 単位, 式説明)
    ("①", "自己資本比率(自己資金は何%か?)", "={B}/{A}*100", "%", "式⇒B÷A"),
    ("②", "債務償還年数(借入は何年で返せるか?)", "={C}/({D}+{E})", "年", "式⇒C÷(D+E)"),
    ("③", "自己資本比率から見た借入上限", "={B}/0.25", "万円", "式⇒B÷0.25 (=B×4)"),
    ("④", "債務償還年数から見た借入上限", "=({D}+{E})*20", "万円", "式⇒(D+E)×20"),
    ("⑤", "翌期の借入上限", "=({r3}+{r4})/2", "万円", "式⇒(③+④)÷2"),
    ("⑥", "翌期の借りれ枠(マイナスはNG)", "={r5}-{C}", "万円", "式⇒⑤-C"),
]

def block(start, label, note):
    r = start
    ws.cell(r, 2, label).font = BOLD
    ws.cell(r, 2).fill = BEIGE
    ws.cell(r, 4, "項目").fill = BEIGE
    ws.cell(r, 5, "単位:万円").fill = BEIGE
    ws.cell(r, 7, note).font = Font(italic=True, size=9)
    r += 1
    refs = {}
    for kind, sym, item, memo in ROWS:
        ws.cell(r, 2, kind)
        ws.cell(r, 3, sym)
        ws.cell(r, 4, item)
        cell = ws.cell(r, 5)
        cell.fill = YELLOW
        cell.border = BOX
        ws.cell(r, 6, "万円")
        ws.cell(r, 7, memo).font = Font(size=9)
        refs[sym] = cell.coordinate
        r += 1
    calc_rows = []
    for i, (sym, item, tpl, unit, desc) in enumerate(CALC):
        ws.cell(r, 2, "計算式").fill = GRAY
        ws.cell(r, 3, sym)
        ws.cell(r, 4, item)
        f = tpl.format(A=refs["A"], B=refs["B"], C=refs["C"], D=refs["D"], E=refs["E"],
                       r3=f"E{calc_rows[2]}" if i >= 4 else "", r4=f"E{calc_rows[3]}" if i >= 4 else "",
                       r5=f"E{calc_rows[4]}" if i >= 5 else "")
        c = ws.cell(r, 5, f)
        c.number_format = "#,##0.0"
        c.border = BOX
        ws.cell(r, 6, unit)
        ws.cell(r, 7, desc).font = Font(size=9)
        if sym in ("⑤", "⑥"):
            ws.cell(r, 4).font = BOLD
            c.font = BOLD
        calc_rows.append(r)
        r += 1
    # 判定行(KHD追加): ⑥と E の状態からひとことで判定
    waku = f"E{calc_rows[5]}"
    e_in = refs["E"]
    ws.cell(r, 2, "判定").fill = GRAY
    ws.cell(r, 4, "この決算での保証協会の見え方")
    j = ws.cell(r, 5,
        f'=IF({e_in}="","(E未記入)",IF({e_in}<0,"⚠赤字:以降3年 設備資金NG(運転も厳しい)",'
        f'IF({waku}<=0,"⚠枠なし:既存借入が上限超過","○ 約 "&TEXT({waku},"#,##0")&"万円まで")))')
    j.font = RED
    ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)
    return r + 2

r = block(4, "＜ア＞ KIKUCHIホールディングス(株) 第4期実績(R8/1期・申告済)",
          "第4期決算書(2026-03-30電子申告)から記入")
r = block(r, "＜イ＞ 同・翌期(第5期)シミュレーション",
          "高松二丁目取得(8/31決済・岩手銀行)/公庫FC融資(9月末実行)織込後の見込みで")
r = block(r, "＜ウ＞ 個人事業(菊池研太・EC) ※参考",
          "確定申告書(青色決算書)から。保証協会は個人事業も対象")

ws.cell(r, 2, ("【注意】E(税引前当期純利益)がマイナスになったら、保証協会は以降3年設備資金は借りられない。"
               "(公庫は試算表で評価してもらえるから異なる)\n"
               "⑥の枠(金額)が翌期新たに借りられる上限。保証協会は年1回の決算書が全て。"))
ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
ws.cell(r, 2).font = Font(size=9)
ws.merge_cells(start_row=r, start_column=2, end_row=r + 2, end_column=8)

for col, w in zip("BCDEFGH", [8, 5, 44, 12, 6, 30, 14]):
    ws.column_dimensions[col].width = w

# ---- 記入ガイド ----
g = wb.create_sheet("記入ガイド(KHD)")
guide = [
    ["信用保証協会 借入余力試算 — KHDの記入ガイド", ""],
    ["", ""],
    ["■ 5つの入力を決算書のどこから取るか", ""],
    ["A 総資産", "BS「資産の部 合計」"],
    ["B 純資産", "BS「純資産の部 合計」+ 役員借入金(菊池→法人の貸付)。役員借入は資本性とみなして足す"],
    ["C 長期借入金", "BS固定負債の「長期借入金」から役員借入金を引いた額(銀行・公庫等の外部借入のみ)"],
    ["D 減価償却", "PL販管費の「減価償却費」(製造原価にあれば合算)"],
    ["E 税引前当期純利益", "PL「税引前当期純利益」"],
    ["", ""],
    ["■ KHDで分かっていること(2026-08-13時点)", ""],
    ["法人第4期(R8/1期)", "2026-03-30電子申告済み。決算書PDFは税理士(橋本先生)/岩手銀行提出分あり"],
    ["家計BSメモ(2026-06-04)", "法人は債務超過▲42万円との記録 → 帳簿上の純資産はマイナスの可能性"],
    ["★重要", "ただしBは役員借入金を含めてよい。菊池さんから法人への貸付があれば、その分Bはプラス側に戻る。ここが結果を左右する最大の変数"],
    ["既存借入(C側)", "朝日信用金庫(返済予定明細あり)/岩手銀行(高松二丁目・8/31決済で新規)/公庫 創業融資(FC・9月末実行予定)"],
    ["", ""],
    ["■ 読み方", ""],
    ["Eが赤字の場合", "保証協会は以降3年設備資金NG。その間の設備は公庫(試算表評価)か銀行プロパーが本命"],
    ["⑥がマイナスの場合", "既存借入が理論上限を超過=保証協会の新規枠なし。純資産を厚くする(役員借入の資本振替・利益計上)が先"],
    ["ア(実績)とイ(翌期)の使い分け", "アで現在地を確定→イに翌期見込み(高松の賃料収入・FC事業・新規借入)を入れて、来期の枠がどう動くかを見る"],
    ["", ""],
    ["■ 出典", ""],
    ["元Excel", "伊東さん共有「特典：信用保証協会から翌期いくら借りられるか試算するEXCEL」Version.2.1(計算式は同一)"],
]
for i, (a, b) in enumerate(guide, start=2):
    g.cell(i, 2, a).font = BOLD if a.startswith(("■", "信用")) or a == "★重要" else Font()
    g.cell(i, 3, b).alignment = Alignment(wrap_text=True, vertical="top")
g.column_dimensions["B"].width = 26
g.column_dimensions["C"].width = 90

wb.save(OUT)
print("saved:", OUT)

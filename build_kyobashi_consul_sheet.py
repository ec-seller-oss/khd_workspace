# -*- coding: utf-8 -*-
"""
京橋クリニック｜AI医療コンサル 受注設計シート（外資コンサル風・多タブ）
04コンサル本部(李牧) / 2026-05-30
配色: クリーム白#F9F6EF × レンガ赤#AA2E26（KHDスライド配色に統一）
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- design tokens ----
CREAM = "F9F6EF"; RED = "AA2E26"; INK = "1A1A1A"; RULE = "DAD6CF"
BEIGE = "F1ECE1"; INPUT_BG = "FFF2CC"; INPUT_INK = "1F4E79"; WHITE = "FFFFFF"
GREEN = "2E7D32"; GREY = "6B6B6B"
F = "Hiragino Sans"

def font(sz=11, b=False, color=INK, name=F): return Font(name=name, size=sz, bold=b, color=color)
def fill(c): return PatternFill("solid", fgColor=c)
def side(c=RULE): return Side(style="thin", color=c)
BORDER = Border(left=side(), right=side(), top=side(), bottom=side())
def align(h="left", v="center", wrap=True): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

wb = openpyxl.Workbook()

def newsheet(title, tab=RED):
    ws = wb.create_sheet(title)
    ws.sheet_properties.tabColor = tab
    ws.sheet_view.showGridLines = False
    return ws

def title_block(ws, kicker, title, sub=None, span=8):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(1, 1, kicker); c.font = font(9, True, RED); c.alignment = align()
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    c = ws.cell(2, 1, title); c.font = font(17, True, INK); c.alignment = align()
    r = 3
    if sub:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=span)
        c = ws.cell(3, 1, sub); c.font = font(10, False, GREY); c.alignment = align()
        r = 4
    # red underline row
    for col in range(1, span+1):
        ws.cell(r, col).fill = fill(RED)
    ws.row_dimensions[r].height = 3
    return r + 1

def header_row(ws, row, headers, widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i, h); c.fill = fill(RED); c.font = font(10, True, WHITE)
        c.alignment = align("center"); c.border = BORDER
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[row].height = 26

def datarow(ws, row, vals, bg=WHITE, bold_first=False, height=None, inkmap=None):
    for i, v in enumerate(vals, 1):
        c = ws.cell(row, i, v); c.border = BORDER
        c.fill = fill(bg)
        col = INK
        if inkmap and i in inkmap: col = inkmap[i]
        c.font = font(10, bold_first and i == 1, col)
        c.alignment = align("left", "top")
    if height: ws.row_dimensions[row].height = height

def section(ws, row, text, span=8):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text); c.fill = fill(BEIGE); c.font = font(11, True, RED)
    c.alignment = align(); c.border = Border(top=side(RED))
    ws.row_dimensions[row].height = 22
    return row + 1

# ============== 00 表紙・進め方 ==============
ws = newsheet("00_表紙・進め方")
r = title_block(ws, "MEDICAL CONSULTING ENGAGEMENT DESIGN",
    "京橋クリニック ｜ AI医療コンサル 受注設計シート",
    "McKinsey式イシュー分解 → ニーズ抽出 → ROI試算 → WBS → 料金体系 → リスク → KPI/出口")
for w, c in zip([22,26,22,22,16,14,14,14], range(1,9)):
    ws.column_dimensions[get_column_letter(c)].width = w
r += 1
info = [
    ("対象", "医療法人社団順医会 京橋クリニック ／ 山崎先生（内科）"),
    ("本部長ルート", "福井（TAW）OK取得済 → 6/15週 会食 → 翌週 医療事務アポ"),
    ("作成", "2026-05-30 ／ 04コンサル本部（AI本部長：李牧）"),
    ("商材", "クリニックDX「My AI」 v8（初期0円／月額0円／成果報酬30%）"),
    ("中核信条", "売り込まない／GIVE先行／信頼の対価として収益（こなくてもいい順序）"),
]
for k, v in info:
    ws.cell(r,1,k).font = font(10, True, RED)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.cell(r,2,v).font = font(10); ws.cell(r,2).alignment = align()
    r += 1
r = section(ws, r+1, "■ タブ構成（左から順に使う）")
tabs = [
    ("01 イシューツリー", "空(問)→論点分解→★菊池が抜けがちな9点（最重要）"),
    ("02 ニーズ抽出", "京橋の実データ（改善提案/問診/レセ）→真の課題→当てるAI→v8対応"),
    ("03 ROI試算", "黄色セルにベースライン入力→削減効果・成果報酬・院に残る額が自動計算"),
    ("04 WBS", "いつ・誰に・何を・どう動かすか（20ステップ）"),
    ("05 料金体系", "WBS連動・1枚印刷用。成果報酬30%モデルの提示版"),
    ("06 リスク・反論処理", "個人情報/院内抵抗/測定公平性/既存ベンダー/医療法規"),
    ("07 KPI・出口", "受注KPI→成功事例化→承継・テナント(110万/件)の本丸LTV"),
]
header_row(ws, r, ["タブ", "中身・使い方"], [24, 70]); r += 1
for t, d in tabs:
    datarow(ws, r, [t, d], height=20, inkmap={1: RED}); r += 1
r = section(ws, r+1, "■ 凡例（セルの色）")
legend = [("黄色背景＋青字", "入力セル（あなたが手入力）"),
          ("黒字", "数式（自動計算）"),
          ("赤太字", "重要マイルストーン・最重要論点"),
          ("ベージュ背景", "見出し・GIVE強調")]
for name, mean in legend:
    ws.cell(r,1,name).font = font(10, True, INK); ws.cell(r,1).fill = fill(INPUT_BG)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
    ws.cell(r,2,mean).font = font(10); ws.cell(r,2).alignment = align()
    r += 1

# ============== 01 イシューツリー ==============
ws = newsheet("01_イシューツリー")
r = title_block(ws, "ISSUE TREE & BLIND SPOTS",
    "イシューツリー：受注→成果→横展開を分解する",
    "空(問)：どうすれば京橋クリニックからAIコンサルを受注し、成果を出し、横展開できるか？")
for w, c in zip([20,40,40,8,8,8,8,8], range(1,9)):
    ws.column_dimensions[get_column_letter(c)].width = w
r = section(ws, r+1, "■ 論点分解（MECE）")
header_row(ws, r, ["論点", "問い", "本シートでの置き場"], [22, 50, 30]); r += 1
tree = [
    ("①状況把握 WHO", "誰の課題か／決裁者は誰か（山崎先生か理事長か）", "01・04(タスク4)"),
    ("②ニーズ WHAT", "真の課題は何か（顕在/潜在）", "02 ニーズ抽出"),
    ("③解決 HOW", "どのAIをどう当てるか（Lステップ/マイAI/IVR）", "02・提案デッキv8"),
    ("④価値の定量化", "削減効果をいくらと示すか（要ベースライン）", "03 ROI試算"),
    ("⑤価格・契約", "成果報酬30%をWBS連動でどう見せるか", "05 料金体系"),
    ("⑥実行 WBS", "いつ・誰が・何を・どう動かすか", "04 WBS"),
    ("⑦リスク/反論", "個人情報・院内抵抗・測定公平性・既存ベンダー・法規", "06 リスク"),
    ("⑧出口/横展開", "成功事例化→他院→承継・テナント(110万/件)", "07 KPI・出口"),
]
for a, b, c in tree:
    datarow(ws, r, [a, b, c], height=20, inkmap={1: RED}); r += 1
r = section(ws, r+1, "★ 李牧の抜け漏れ指摘 — 菊池が網羅できていない9点（デッキを開く前に埋めろ）")
header_row(ws, r, ["#", "抜けている論点", "なぜ致命的か／打ち手"], [5, 26, 63]); r += 1
gaps = [
    ("1", "【最重要】成果測定のベースライン", "成果報酬30%は『導入前の数値』が無いと請求不能。会食で現状値（電話件数/日・事務残業h・レセ工数・アイコール費）を必ず押さえる。取り損ねたら課金根拠が消える。"),
    ("2", "決裁構造", "『医療法人社団順医会』＝法人・複数院の可能性。予算決裁は山崎先生個人か理事長か。会食相手が決裁者かを福井に確認。"),
    ("3", "現場の抵抗", "事務職シートに『継続勤務 検討中』＝離職リスク。経営者だけ説得しても現場が潰す。導入が離職の引き金になり得る。現場を味方にする順序。"),
    ("4", "既存ベンダー", "アイコール（既存予約システム）が稼働中。Lステップと併存か置換か、既存契約の縛り・解約条件。"),
    ("5", "医療法規", "LINE/AI問診の薬機法・医療広告ガイドライン・要配慮個人情報・クラウド保存の3省2ガイドライン準拠。医療特有の抜け。"),
    ("6", "成果報酬の期間・上限", "30%を何ヶ月取るか未定。無限は不自然。例：24ヶ月→以降は院の資産として無償運用。"),
    ("7", "キャッシュフロー", "成果報酬は後ろ倒し＝菊池の持ち出し期間。無料診断・構築工数を何件まで撒けるか。"),
    ("8", "福井の座組み", "本案件でのTAW/福井の分け前（別途作成の覚書ドラフトと連動）。口頭で流さず書面化。"),
    ("9", "出口設計", "京橋は内科＝自費少→『売上押上げ』でなく『コスト削減』で正直に勝負。本丸は承継・テナントへの接続。"),
]
for n, t, d in gaps:
    datarow(ws, r, [n, t, d], bg=(INPUT_BG if n in ("1",) else WHITE), height=44,
            inkmap={1: RED, 2: RED}); r += 1

# ============== 02 ニーズ抽出 ==============
ws = newsheet("02_ニーズ抽出")
r = title_block(ws, "NEEDS EXTRACTION (REAL DATA)",
    "ニーズ抽出：京橋クリニックの実データから",
    "仮説でなく実資料（看護師改善提案・事務職シート・問診票・レセプト集計）から抽出。これが最強の武器。")
for w, c in zip([26,34,18,20,18,8,8,8], range(1,9)):
    ws.column_dimensions[get_column_letter(c)].width = w
r = section(ws, r+1, "■ 実データ → 真の課題 → 当てるAI → v8対応")
header_row(ws, r, ["出所（実データ）", "生の声・事実", "真の課題", "当てるAI", "v8対応"],
           [26, 38, 18, 22, 18]); r += 1
needs = [
    ("看護師 改善提案（緊急度 高）", "電話回線が受付人数より多い／化学物質過敏症の問合せ電話が長い", "電話過多・一次対応の浪費", "AI音声IVR 一次対応", "電話DX 月1200件→40%減"),
    ("看護師／事務", "アイコール番号を取った患者が時間外/遅れて来院→診察長引き・順番前後でクレーム", "予約・受付導線の崩壊", "Lステップ予約＋リマインド＋順番通知", "問診予約DX"),
    ("問診票（紙・手書き）", "紙問診を手入力、複数枚/日", "問診の二重手間", "LINE事前問診（Lステップ）", "受付問診70%減"),
    ("事務職シート", "勤務条件に不満・継続勤務『検討中』・業務固定化", "離職リスク・採用単価", "負担軽減で定着", "採用定着 70-150万"),
    ("改善提案", "診察室1のカルテ探しに時間／書類", "書類・カルテ業務", "マイAI 書類アシスト", "書類DX"),
    ("レセプト集計（月初）", "月初レセプト・初診6/日 ほか", "レセ点検工数", "マイAI レセ点検", "レセプト−45h"),
]
for row in needs:
    datarow(ws, r, list(row), height=40, inkmap={3: RED}); r += 1
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
ws.cell(r,1,"【集患データ】問診票『当院を何で知りましたか』＝集患経路が取れる→Lステップで再来院・口コミ導線。"
        "ただし京橋は内科（風邪/咳中心）＝自費少→ROIは『コスト削減』主軸で正直に見せる。").font = font(10, True, GREEN)
ws.cell(r,1).alignment = align(); ws.row_dimensions[r].height = 34; r += 2
r = section(ws, r, "■ ニーズ抽出の進め方（4段階）")
for t in ["Phase0（今/会食前）：実データから仮説ニーズを構造化（上表）",
          "Phase1（会食）：仮説をぶつけ検証＋現状値（ベースライン）を確定 ← 03タブへ入力",
          "Phase2（無料診断）：競合分析＋御院専用 削減シミュレーション報告書",
          "Phase3：最大ペイン1-2個（電話IVR＋LINE問診）に絞って段階導入"]:
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    ws.cell(r,1,t).font = font(10); ws.cell(r,1).alignment = align(); ws.row_dimensions[r].height = 18; r += 1

# ============== 03 ROI試算 ==============
ws = newsheet("03_ROI試算")
r = title_block(ws, "ROI SIMULATION (INPUT-LINKED)",
    "ROI試算：御院の数字を入れると残る額が出る",
    "黄色セルに会食で取った現状値を入れる→削減効果・成果報酬・院に残る額が自動計算。")
for w, c in zip([34,18,16,30,8,8,8,8], range(1,9)):
    ws.column_dimensions[get_column_letter(c)].width = w
r = section(ws, r+1, "■ ベースライン入力（黄色＝会食で取得して手入力）")
header_row(ws, r, ["項目", "入力値", "単位", "備考"], [34, 18, 16, 30]); r += 1
def inrow(rr, label, val, unit, note):
    ws.cell(rr,1,label).font = font(10); ws.cell(rr,1).border = BORDER; ws.cell(rr,1).alignment = align()
    c = ws.cell(rr,2,val); c.font = font(11, True, INPUT_INK); c.fill = fill(INPUT_BG)
    c.border = BORDER; c.alignment = align("center")
    ws.cell(rr,3,unit).font = font(10, color=GREY); ws.cell(rr,3).border = BORDER; ws.cell(rr,3).alignment = align("center")
    ws.cell(rr,4,note).font = font(9, color=GREY); ws.cell(rr,4).border = BORDER; ws.cell(rr,4).alignment = align()
base_start = r
inrow(r, "事務スタッフ人数", 3, "名", "京橋は要確認"); r += 1
inrow(r, "事務 月間人件費（総額）", 830000, "円", "標準モデル"); r += 1
inrow(r, "総労働時間（月・事務計）", 480, "時間", "標準モデル"); r += 1
inrow(r, "問診DXの削減見込", 42, "時間/月", "Lステップ事前問診"); r += 1
inrow(r, "レセプトDXの削減見込", 45, "時間/月", "マイAIレセ点検"); r += 1
inrow(r, "電話DXの削減見込", 40, "時間/月", "AI音声IVR"); r += 1
inrow(r, "書類DXの削減見込", 12, "時間/月", "マイAI書類アシスト"); r += 1
inrow(r, "成果報酬率", 0.30, "—", "30%"); r += 1
# rows: 2..9 are values in column B. Map:
ROW_STAFFCOST = base_start + 1   # 事務月間人件費
ROW_TOTALH = base_start + 2      # 総労働時間
ROW_Q = base_start + 3
ROW_RECE = base_start + 4
ROW_TEL = base_start + 5
ROW_DOC = base_start + 6
ROW_RATE = base_start + 7
r = section(ws, r+1, "■ 自動計算（黒字＝数式）")
header_row(ws, r, ["指標", "計算", "結果"], [34, 30, 22]); r += 1
def calcrow(rr, label, formula_desc, formula, numfmt="#,##0"):
    ws.cell(rr,1,label).font = font(10, True); ws.cell(rr,1).border = BORDER; ws.cell(rr,1).alignment = align()
    ws.cell(rr,2,formula_desc).font = font(9, color=GREY); ws.cell(rr,2).border = BORDER; ws.cell(rr,2).alignment = align()
    c = ws.cell(rr,3,formula); c.font = font(11, True, INK); c.border = BORDER; c.alignment = align("center")
    c.number_format = numfmt
ratecell = f"B{ROW_RATE}"
hsum = f"(B{ROW_Q}+B{ROW_RECE}+B{ROW_TEL}+B{ROW_DOC})"
unitcost = f"(B{ROW_STAFFCOST}/B{ROW_TOTALH})"
calcrow(r, "時間単価", "人件費÷総労働時間", f"=ROUND({unitcost},0)", "#,##0\"円\""); r += 1
calcrow(r, "月間削減時間", "問診+レセ+電話+書類", f"={hsum}", "#,##0\"時間\""); R_HSUM = r; r += 1
calcrow(r, "月間削減額", "削減時間×時間単価", f"=ROUND({hsum}*{unitcost},0)", "#,##0\"円\""); R_MSAVE = r; r += 1
calcrow(r, "年間削減額", "月×12", f"=ROUND({hsum}*{unitcost}*12,0)", "#,##0\"円\""); r += 1
calcrow(r, "成果報酬（月）", "月削減額×30%", f"=ROUND({hsum}*{unitcost}*{ratecell},0)", "#,##0\"円\""); r += 1
calcrow(r, "院に残る（月）", "月削減額×70%", f"=ROUND({hsum}*{unitcost}*(1-{ratecell}),0)", "#,##0\"円\""); r += 1
calcrow(r, "院に残る（年）", "院に残る月×12", f"=ROUND({hsum}*{unitcost}*(1-{ratecell})*12,0)", "#,##0\"円\""); r += 1
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
ws.cell(r,1,"※標準モデル（事務3名）で 月139h×約1,729円＝月約24万削減 → 成果報酬7.2万/月、院に残る16.8万/月（年202万）。"
        "会食で取った実数を黄色に入れれば、その場で『御院だといくら残るか』を提示できる。").font = font(10, True, RED)
ws.cell(r,1).alignment = align(); ws.row_dimensions[r].height = 34

# ============== 04 WBS ==============
ws = newsheet("04_WBS")
r = title_block(ws, "WORK BREAKDOWN STRUCTURE",
    "WBS：いつ・誰に・何を・どう動かすか",
    "『動かす相手』を主語に。会食→無料診断→導入→測定→課金→横展開まで20ステップ。", span=9)
header_row(ws, r, ["#","Phase","時期","タスク","主体","動かす相手","具体アクション","成果物","ステータス"],
           [4,9,10,20,8,12,34,16,10]); r += 1
wbs = [
 (1,"P0事前","〜6/14","ニーズ仮説構造化","菊池","—","実データから課題表作成（02タブ）","ニーズ抽出表",""),
 (2,"P0事前","〜6/14","会食ヒアリングシート","菊池","—","ベースライン取得項目を設計（電話/残業/アイコール費/レセ工数）","ヒアリング表",""),
 (3,"P0事前","〜6/14","競合簡易分析","菊池","—","京橋周辺クリニックのWeb/LINE/予約対応を調査","手土産の種",""),
 (4,"P0事前","〜6/14","決裁者確認","菊池","福井","『予算を決めるのは山崎先生か理事長か』をLINE確認","決裁構造メモ",""),
 (5,"P1会食","6/15週","課題ヒアリング","菊池","山崎先生","まず聞く。デッキを開かず仮説を検証","現状課題メモ",""),
 (6,"P1会食","6/15週","★ベースライン数値取得","菊池","山崎先生","電話件数/残業h/レセ工数/離職状況を聞き出す","ベースライン値",""),
 (7,"P1会食","6/15週","v8提示（当てはめ）","菊池","山崎先生","『先生がさっき言った○○、これで消えます』","—",""),
 (8,"P1会食","6/15週","★診断＋データ開示の許可","菊池","山崎先生","無料削減シミュ作成と既存データ閲覧の許可を取る","口頭合意",""),
 (9,"P1会食","6/15週","医療事務アポ内諾","菊池","山崎先生","翌週の事務ヒアリングの内諾を取る","アポ内諾",""),
 (10,"P2診断","〜+1週","削減シミュ報告書","菊池","—","既存データ精査→御院専用の定量報告書","診断書",""),
 (11,"P2診断","〜+1週","現場ヒアリング","菊池","事務/受付/看護","業務フロー・ボトルネックを実地で・現場を味方化","フロー表",""),
 (12,"P2診断","〜+1週","導入可否判断","菊池","山崎先生","報告書を提示→GO/NOをもらう","導入合意",""),
 (13,"P3導入","+2〜6週","段階導入（優先2課題）","菊池","ベンダー/院","電話IVR＋LINE問診からスモールスタート","稼働環境",""),
 (14,"P3導入","+2〜6週","★測定ルール合意書","菊池","山崎先生","同等患者数帯・人員変更月除外を書面化","測定合意書",""),
 (15,"P3導入","+2〜6週","スタッフ説明会","菊池","事務/看護","抵抗を取る・『楽になる』を体験させる","説明会",""),
 (16,"P3導入","+2〜6週","個人情報・法規ルール","菊池","山崎先生","3省2GL・要配慮個人情報の運用ルールを書面化","運用ルール",""),
 (17,"P4測定","導入30-90日","効果測定","菊池","—","測定は菊池が巻き取る（院の手間最小）","測定レポート",""),
 (18,"P4測定","導入30-90日","成果報酬請求","菊池","法人","削減額×30%を請求","請求",""),
 (19,"P4測定","導入30-90日","福井座組み精算","菊池","福井","覚書連動で分配精算","精算",""),
 (20,"P5出口","以降","成功事例化→横展開","菊池","福井","事例を福井ルート他院へ・承継/テナント(110万/件)へ接続","紹介",""),
]
for row in wbs:
    star = str(row[0]) in ("6","8","14")
    datarow(ws, r, list(row), bg=(INPUT_BG if star else WHITE), height=32,
            inkmap={4: (RED if star else INK), 6: RED}); r += 1

# ============== 05 料金体系 ==============
ws = newsheet("05_料金体系")
r = title_block(ws, "PRICING (WBS-LINKED, 1-PAGE)",
    "料金体系：WBS連動・1枚印刷用",
    "初期0円／月額0円／成果報酬30%。フェーズごとに『院の費用』と『課金根拠』を明示。")
for w, c in zip([20,26,34,16,18,8,8,8], range(1,9)):
    ws.column_dimensions[get_column_letter(c)].width = w
r = section(ws, r+1, "■ WBS連動 料金表")
header_row(ws, r, ["WBSフェーズ","院がやること","菊池が提供","院の費用","課金根拠"], [20,26,34,16,18]); r += 1
price = [
 ("P1 会食/ヒアリング","課題を話す","課題整理・提案","0円","無料"),
 ("P2 無料診断","既存データ開示","競合分析＋削減シミュ報告書","0円","無料(GIVE)"),
 ("P3 導入/伴走","院長が運用ルール決裁","設計・導入・3ヶ月伴走・補助金申請支援","0円(初期)","成果報酬に内包"),
 ("P4 成果測定後","測定に協力(手間最小)","効果測定","削減額×30%","実削減額ベース"),
]
for row in price:
    datarow(ws, r, list(row), height=30, inkmap={1: RED, 4: RED}); r += 1
r = section(ws, r+1, "■ 成果報酬の具体例（事務3名モデル）")
for label, val in [("月削減時間","139 時間"),("時間単価","1,729 円"),("月削減額","240,531 円"),
                   ("成果報酬(30%)","72,159 円"),("院に残る(70%)","168,372 円"),("年換算 院に残る","2,020,464 円")]:
    ws.cell(r,1,label).font = font(10, True); ws.cell(r,1).border = BORDER; ws.cell(r,1).fill = fill(BEIGE)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    ws.cell(r,2,val).font = font(11, True, RED); ws.cell(r,2).border = BORDER; ws.cell(r,2).alignment = align()
    r += 1
r = section(ws, r+1, "■ 課金条件")
for k, v in [("課金開始","効果測定で削減が確認された月から"),
             ("成果報酬率","削減額の30%"),
             ("契約期間","例：24ヶ月（以降は院の資産として無償運用／保守のみ）※菊池の決め事"),
             ("解約","最低期間以降30日前通知・効果0なら費用0円"),
             ("補助金","IT導入補助金で初期実費をさらに圧縮")]:
    ws.cell(r,1,k).font = font(10, True, INK); ws.cell(r,1).border = BORDER
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    ws.cell(r,2,v).font = font(10); ws.cell(r,2).border = BORDER; ws.cell(r,2).alignment = align()
    r += 1
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
ws.cell(r,1,"注記：この成果報酬は『入口』。信頼の対価として、その先の承継・テナント・不動産（110万/件）へ接続する。売り込まない。").font = font(10, True, RED)
ws.cell(r,1).alignment = align(); ws.row_dimensions[r].height = 30

# ============== 06 リスク・反論処理 ==============
ws = newsheet("06_リスク・反論処理")
r = title_block(ws, "RISK & OBJECTION HANDLING",
    "リスク・反論処理：先回りで潰す",
    "医療現場・経営者の不安を、提案前に答えで用意する。")
for w, c in zip([24,34,38,8,8,8,8,8], range(1,9)):
    ws.column_dimensions[get_column_letter(c)].width = w
header_row(ws, r, ["懸念・反論","起こり得ること","先回りの打ち手"], [24,34,40]); r += 1
risks = [
 ("患者情報・個人情報","要配慮個人情報の漏洩懸念で導入拒否","目的外利用なし・匿名化・運用ルールを最初に書面化（v8 SECURITY）。3省2ガイドライン準拠。"),
 ("院内スタッフの抵抗","『また仕事が増える』『離職検討中の事務が辞める』","段階導入＋3ヶ月伴走。説明会で『楽になる』を体験させる。経営者でなく現場を先に味方化。"),
 ("成果測定の公平性","削減を水増しと疑われ課金で揉める","導入前後で同等患者数帯・人員変更月除外を測定合意書に明記。測定は菊池が巻き取る。"),
 ("既存ベンダー(アイコール)","二重コスト・解約縛り","併存か置換かを設計。既存契約の解約条件を確認。Lステップへの移行ロードマップ。"),
 ("医療法規","薬機法・医療広告GL違反リスク","LINE/AI問診の表現を法規チェック。クラウド保存の準拠。院長決裁の運用ルール。"),
 ("効果が出なかったら","固定費だけ取られる不信","効果0なら費用0円。ロックインなし（30日前通知で解約可）。リスクは菊池が取る。"),
 ("キャッシュフロー(自社)","成果報酬後ろ倒しで持ち出し","無料診断・構築工数の上限を決める。撒く件数を絞る（京橋に集中）。"),
]
for row in risks:
    datarow(ws, r, list(row), height=42, inkmap={1: RED}); r += 1

# ============== 07 KPI・出口 ==============
ws = newsheet("07_KPI・出口")
r = title_block(ws, "KPI & EXIT TO CORE BUSINESS",
    "KPI・出口：受注の先の本丸へ",
    "AIコンサルは入口。信頼の対価として承継・テナント（110万/件）へ繋ぐのが本丸LTV。")
for w, c in zip([30,24,40,8,8,8,8,8], range(1,9)):
    ws.column_dimensions[get_column_letter(c)].width = w
r = section(ws, r+1, "■ 受注KPI（この案件）")
header_row(ws, r, ["KPI","目標","測り方"], [30,24,42]); r += 1
kpis = [
 ("会食→無料診断 許可率","100%（1院）","会食で診断＋データ開示の口頭合意を取れたか"),
 ("無料診断→導入 GO率","目標 GO","削減シミュ報告書提示後にGO/NOをもらう"),
 ("初成果報酬の発生","導入90日内","効果測定で削減確認→請求"),
 ("成功事例化","1件","数字で出た事例を横展開の弾にする"),
]
for row in kpis:
    datarow(ws, r, list(row), height=24, inkmap={1: RED}); r += 1
r = section(ws, r+1, "■ 出口：本丸LTVへの接続")
exits = [
 ("成功事例 → 福井ルート他院","京橋の数字を持って、福井紹介の他院へ横展開（CPAゼロ）"),
 ("AIコンサル → 医療テナント仲介","院の信頼を土台に、テナント・移転の相談へ（約60万/件）"),
 ("AIコンサル → 承継・事業計画","承継・事業計画コンサル（50〜150万）＝生命線（110万/件）"),
 ("座組み","福井(TAW)との利益分配は覚書ドラフトと連動・書面化"),
]
for a, b in exits:
    ws.cell(r,1,a).font = font(10, True, RED); ws.cell(r,1).border = BORDER; ws.cell(r,1).alignment = align()
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    ws.cell(r,2,b).font = font(10); ws.cell(r,2).border = BORDER; ws.cell(r,2).alignment = align()
    ws.row_dimensions[r].height = 24; r += 1
r += 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
ws.cell(r,1,"中核信条：売り込まない／GIVE先行／信頼の対価として収益。順序を逆にした瞬間（天野先生の轍）に全てを失う。").font = font(10, True, GREEN)
ws.cell(r,1).alignment = align(); ws.row_dimensions[r].height = 28

# ---- remove default sheet & set active ----
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]
wb.active = 0

out = "/Users/kikuchikenta/Library/CloudStorage/GoogleDrive-ec-seller@kikuchi-hd.net/マイドライブ/260526_AI医療コンサル/京橋クリニック_AI医療コンサル_受注設計シート.xlsx"
wb.save(out)
print("SAVED:", out)
print("SHEETS:", wb.sheetnames)

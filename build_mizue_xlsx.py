# -*- coding: utf-8 -*-
"""瑞江1-8-15 物件概要書（編集用xlsx）を飯山満フォーマットに倣って生成"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "物件概要書"

# 色
NAVY = "2C3E50"; NAVY2 = "34495E"; GRAY = "ECF0F1"; CREAM = "FDF2E9"; FLOW = "EAF2F8"
thin = Side(style="thin", color="95A5A6")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def style(cell, *, bg=None, fontcolor="1A1A1A", bold=False, size=10, align="left", wrap=False):
    if bg: cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(name="Hiragino Sans", color=fontcolor, bold=bold, size=size)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = border

# 列幅（A=セクション縦, B=ラベル, C-E=値, F=ラベル2, G=値2）
widths = {"A":4, "B":13, "C":20, "D":13, "E":18, "F":13, "G":22}
for c,w in widths.items(): ws.column_dimensions[c].width = w

r = 1
# タイトル
ws.merge_cells(f"A{r}:G{r}")
style(ws[f"A{r}"], bg=NAVY, fontcolor="FFFFFF", bold=True, size=14)
ws[f"A{r}"] = "【物件概要書】江戸川区瑞江一丁目　木造3階建戸建（再建築不可）"
ws.row_dimensions[r].height = 26
r += 1
ws.merge_cells(f"A{r}:G{r}")
style(ws[f"A{r}"], bg=NAVY2, fontcolor="FFFFFF", size=9)
ws[f"A{r}"] = "再建築不可／中古戸建　— 買取・出口提携のご依頼 —　作成日 2026年6月10日"
r += 1

def row_label_val(label, val, *, label2=None, val2=None, tbd=False, merge_val_to="G"):
    global r
    style(ws[f"B{r}"], bg=GRAY, bold=True, size=9)
    ws[f"B{r}"] = label
    fc = "C0392B" if tbd else "1A1A1A"
    if label2 is not None:
        ws.merge_cells(f"C{r}:D{r}")
        style(ws[f"C{r}"], fontcolor=fc, size=9, wrap=True)
        ws[f"C{r}"] = val
        for cc in ["C","D"]: style(ws[f"{cc}{r}"], fontcolor=fc, size=9)
        ws[f"C{r}"] = val
        style(ws[f"F{r}"], bg=GRAY, bold=True, size=9)
        ws[f"F{r}"] = label2
        style(ws[f"G{r}"], fontcolor=fc, size=9, wrap=True)
        ws[f"G{r}"] = val2
    else:
        ws.merge_cells(f"C{r}:{merge_val_to}{r}")
        for col in ["C","D","E","F","G"]:
            if col <= merge_val_to: style(ws[f"{col}{r}"], fontcolor=fc, size=9, wrap=True)
        ws[f"C{r}"] = val
    r += 1

# 物件名
row_label_val("物件名", "東京都江戸川区瑞江一丁目　木造3階建　居宅（再建築不可）")

# 土地ブロック
land_start = r
rows_land = [
    ("住居表示", "東京都江戸川区瑞江一丁目8-15", None, None, False),
    ("地番", "江戸川区瑞江一丁目 8番28・8番37", None, None, False),
    ("交通", "都営新宿線「瑞江」駅 徒歩13分（約1.1km）", None, None, False),
    ("地目", "宅地", "現況", "中古戸建（建物現存）", False),
    ("土地面積", "53.00㎡（16.03坪）　登記：8番28=42.32㎡＋8番37=10.68㎡／実測未了", None, None, False),
    ("権利", "所有権（共有2名）", "用途地域", "第一種中高層住居専用地域（確認中）", False),
    ("接道", "再建築不可（建築基準法上の接道要件 未充足の可能性／詳細確認中）", None, None, True),
    ("路線価", "相続税・固定資産税路線価とも 確認中（瑞江一丁目8番付近）", None, None, False),
]
for lbl,v,l2,v2,tbd in rows_land:
    row_label_val(lbl, v, label2=l2, val2=v2, tbd=tbd)
ws.merge_cells(f"A{land_start}:A{r-1}")
style(ws[f"A{land_start}"], bg=NAVY, fontcolor="FFFFFF", bold=True, size=10, align="center")
ws[f"A{land_start}"] = "土地"
ws[f"A{land_start}"].alignment = Alignment(horizontal="center", vertical="center", text_rotation=255)

# 建物ブロック
b_start = r
rows_b = [
    ("構造", "木造スレート葺 3階建", "種類", "居宅", False),
    ("築年", "昭和63年(1988年)新築", "築年数", "約38年", False),
    ("延床面積", "約90㎡（推定）※各階面積は登記原本／現調で要確認", None, None, False),
    ("現況", "空家想定（居住中か要確認）", None, None, True),
    ("現地写真", "別途データあり（外観・内装）", None, None, False),
]
for lbl,v,l2,v2,tbd in rows_b:
    row_label_val(lbl, v, label2=l2, val2=v2, tbd=tbd)
ws.merge_cells(f"A{b_start}:A{r-1}")
style(ws[f"A{b_start}"], bg=NAVY, fontcolor="FFFFFF", bold=True, size=10, align="center")
ws[f"A{b_start}"] = "建物現況"
ws[f"A{b_start}"].alignment = Alignment(horizontal="center", vertical="center", text_rotation=255)

# 価格
style(ws[f"B{r}"], bg=GRAY, bold=True, size=10); ws[f"B{r}"] = "価格"
ws.merge_cells(f"C{r}:D{r}")
style(ws[f"C{r}"], bg=CREAM, fontcolor="C0392B", bold=True, size=12, align="center")
ws[f"C{r}"] = "売主希望 2,000万円（応相談）"
style(ws[f"D{r}"], bg=CREAM)
style(ws[f"E{r}"], bg=GRAY, bold=True, size=9); ws[f"E{r}"] = "取引態様"
ws.merge_cells(f"F{r}:G{r}")
style(ws[f"F{r}"], size=9); ws[f"F{r}"] = "媒介（商流下記）"; style(ws[f"G{r}"])
ws.row_dimensions[r].height = 22
r += 1

# 商流
style(ws[f"B{r}"], bg=GRAY, bold=True, size=9); ws[f"B{r}"] = "商流"
ws.merge_cells(f"C{r}:G{r}")
for col in ["C","D","E","F","G"]: style(ws[f"{col}{r}"], bg=FLOW, size=9, wrap=True)
ws[f"C{r}"] = "元付業者 → 株式会社プロパティア（相部様）→ 弊社（バイセル不動産）→ 御社 cloudmil（買取・最終出口）"
r += 1
style(ws[f"B{r}"], bg=GRAY, bold=True, size=9); ws[f"B{r}"] = "手数料条件"
ws.merge_cells(f"C{r}:G{r}")
for col in ["C","D","E","F","G"]: style(ws[f"{col}{r}"], size=9, wrap=True)
ws[f"C{r}"] = "紹介業者（プロパティア）＝仲介手数料＋3%（税別）。弊社手数料・按分は別途ご相談。"
r += 1

# 注記
notes = [
 "※本物件は再建築不可のため、御社の再建築不可可の出口ルートでの買取・転売を前提に査定をお願いするものです。",
 "※接道状況・用途地域・確定面積（実測）・建物現況は元付／相部様へ確認中です。判明次第すぐ共有いたします。",
 "※掲載数値は登記・公図および周辺相場からの暫定値を含みます。秘密厳守でお願いいたします。",
]
for n in notes:
    ws.merge_cells(f"A{r}:G{r}")
    style(ws[f"A{r}"], fontcolor="555555", size=8, wrap=True)
    ws[f"A{r}"] = n
    r += 1

# 会社情報
r += 1
comp = [("会社名","株式会社B&Sエステート（バイセル不動産）"),("住所","埼玉県さいたま市大宮区宮町4-38-4"),
        ("担当","菊池（きくち）"),("TEL","080-6047-2797"),("Mail","buysell.tochi@gmail.com")]
for lbl,v in comp:
    style(ws[f"E{r}"], bg=GRAY, bold=True, size=9); ws[f"E{r}"] = lbl
    ws.merge_cells(f"F{r}:G{r}")
    style(ws[f"F{r}"], size=9); ws[f"F{r}"] = v; style(ws[f"G{r}"])
    r += 1

# 印刷設定 A4 1ページ
ws.print_area = f"A1:G{r-1}"
ws.page_setup.orientation = "portrait"
ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 1
ws.page_margins.left = ws.page_margins.right = 0.3
ws.page_margins.top = ws.page_margins.bottom = 0.3

out = "/Users/kikuchikenta/01_honbu_docs_automation/297_物件概要書_江戸川区瑞江1-8-15.xlsx"
wb.save(out)
print("saved:", out)

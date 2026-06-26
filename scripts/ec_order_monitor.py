# -*- coding: utf-8 -*-
"""
Coupang Wing 新規注文 ポーリングモニター — 03_事業運営 / EC韓国輸出
────────────────────────────────────────────────────────────────────
【何をするか】
  Coupang Wing (wing.coupang.com) の注文一覧ページを Playwright で定期巡回し、
  新規注文を検出してDBと当月販売管理表Excelに自動記録する。
  検出した注文数をLINEで通知し、エラー時は緊急通知する。

【他スクリプトとの関係】
  - ec_automation_db.py  … SQLite DB操作ヘルパー（注文のupsert/取得）
  - ec_notify.py         … LINE通知ヘルパー
  - wing_auto_download.py … 同じ認証設定・セッション保存方式を共有
  - run_monthly_ec_pipeline.sh … 月次パイプラインから呼ばれることもある

【認証設定（初回のみ）】
  cp ~/.config/khd/coupang_template.json ~/.config/khd/coupang.json
  # エディタで email/password/LINE token を記入
  chmod 600 ~/.config/khd/coupang.json

【実行コマンド】
  # 30分毎にポーリング（常駐起動）
  python3 scripts/ec_order_monitor.py

  # 1回だけ実行（cronから呼ぶ時）
  python3 scripts/ec_order_monitor.py --check-once

  # DBに書かず標準出力のみ確認
  python3 scripts/ec_order_monitor.py --dry-run --check-once

【自動実行タイミング】
  crontab 例（30分毎）:
    */30 * * * * /usr/local/bin/python3 /Users/kikuchikenta/01_honbu_docs_automation/scripts/ec_order_monitor.py --check-once >> /tmp/ec_order_monitor.log 2>&1

【セッション管理】
  ~/.config/khd/wing_sessions/ にCookieをJSON保存。
  セッション切れ時は再ログイン → 2FA/CAPTCHAが必要な場合はLINE緊急通知。
"""

import json
import os
import re
import sys
import time
import traceback
import argparse
import glob
from pathlib import Path
from datetime import date, datetime, timedelta
from typing import Optional, List, Dict, Any

# playwright のインポート（未インストール時は案内して終了）
try:
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
except ImportError:
    sys.exit(
        "❌ Playwright が未インストールです。\n"
        "   pip install playwright && playwright install chromium"
    )

# openpyxl
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("❌ openpyxl が未インストールです。\n   pip install openpyxl")

# 自作モジュール（同じ scripts/ ディレクトリから）
_SCRIPTS_DIR = Path(__file__).parent
sys.path.insert(0, str(_SCRIPTS_DIR))
from ec_automation_db import upsert_order, update_order_status, get_orders_by_status
from ec_notify import notify

# ────── 定数 ──────────────────────────────────────────────────────────────
CFG_PATH    = Path.home() / ".config/khd/coupang.json"
SESS_DIR    = Path.home() / ".config/khd/wing_sessions"
WING_BASE   = "https://wing.coupang.com"
LOGIN_URL   = f"{WING_BASE}/login"
ORDER_URL   = f"{WING_BASE}/vendor/partner/orders"

# 販売管理表の保存ディレクトリ
SALES_DIR   = Path(
    "/Users/kikuchikenta/Library/CloudStorage/"
    "GoogleDrive-ec-seller@kikuchi-hd.net/"
    "共有ドライブ/01_個人/2025_帳票、明細/韓国輸出売上"
)

# 各列のインデックス（0始まり）。ヘッダー行(6行目)で確認済み。
# A=0, B=1, C=2, F=5, G=6, H=7, I=8,
# U=20, V=21, W=22, X=23,
# AP=41, AQ=42, AR=43, AS=44, AT=45
COL_STATUS      = 0   # A: 進捗
COL_ORDER_NO    = 1   # B: 注文番号
COL_PLATFORM    = 2   # C: プラットフォーム
COL_BUYER_NAME  = 5   # F: 注文者名
COL_PRODUCT_KR  = 6   # G: 商品名(韓国語)
COL_PRODUCT_JP  = 7   # H: 日本語商品名
COL_ASIN        = 8   # I: ASIN
COL_SALE_DATE   = 20  # U: Sale Date
COL_SHIP_DL     = 21  # V: 発送予定日
COL_QTY         = 22  # W: Quantity
COL_SALE_PRICE  = 23  # X: Sale Price (KRW)
COL_POSTAL      = 41  # AP: 郵便番号
COL_ADDRESS     = 42  # AQ: 住所
COL_PHONE       = 43  # AR: 携帯電話番号
COL_CUSTOMS_NO  = 44  # AS: 個人通関番号
COL_DELIVERY_MSG= 45  # AT: 配送メッセージ

# Wing で「新規注文」とみなす韓国語ステータス
NEW_ORDER_STATUSES = {"신규 주문", "결제완료", "결제 완료"}

POLL_INTERVAL_SEC = 30 * 60  # 30分
NAV_WAIT_MS       = 15000
LOG_PREFIX        = "[ec_order_monitor]"


def log(msg: str) -> None:
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"{LOG_PREFIX} {ts} {msg}", flush=True)


# ────── 設定読み込み ──────────────────────────────────────────────────────
def load_config() -> Dict[str, Any]:
    if not CFG_PATH.exists():
        sys.exit(
            f"❌ 設定ファイルが見つかりません: {CFG_PATH}\n"
            f"   cp ~/.config/khd/coupang_template.json ~/.config/khd/coupang.json\n"
            f"   してから email/password/LINE token を記入してください。"
        )
    with open(CFG_PATH) as f:
        return json.load(f)


# ────── セッション保存・ロード ────────────────────────────────────────────
def session_path(account_name: str) -> Path:
    SESS_DIR.mkdir(parents=True, exist_ok=True)
    safe = re.sub(r"[^\w]", "_", account_name)
    return SESS_DIR / f"{safe}_session.json"


def save_session(context: Any, account_name: str) -> None:
    path = session_path(account_name)
    cookies = context.cookies()
    path.write_text(json.dumps(cookies, ensure_ascii=False, indent=2))
    log(f"セッション保存: {path}")


def load_session(context: Any, account_name: str) -> bool:
    path = session_path(account_name)
    if path.exists():
        try:
            cookies = json.loads(path.read_text())
            context.add_cookies(cookies)
            log(f"セッション復元: {path}")
            return True
        except Exception as e:
            log(f"セッション読み込みエラー（無視）: {e}")
    return False


# ────── ログイン ─────────────────────────────────────────────────────────
def do_login(page: Any, email: str, password: str, account_name: str) -> None:
    log(f"ログイン開始: {account_name} ({email})")
    page.goto(LOGIN_URL, timeout=NAV_WAIT_MS)
    page.wait_for_load_state("domcontentloaded")

    try:
        page.fill(
            'input[type="email"], input[name="email"], input[id*="email"], input[id*="Email"]',
            email, timeout=8000
        )
    except PWTimeout:
        log("⚠️ emailフィールド未検出")
        page.screenshot(path=f"/tmp/wing_order_login_err_{account_name}.png")
        raise

    page.fill('input[type="password"]', password, timeout=5000)

    for sel in [
        'button[type="submit"]',
        'button:has-text("로그인")',
        'button:has-text("Login")',
        'input[type="submit"]',
    ]:
        try:
            page.click(sel, timeout=3000)
            break
        except PWTimeout:
            continue

    try:
        page.wait_for_url(lambda u: "login" not in u, timeout=15000)
        log(f"ログイン成功: {account_name}")
    except PWTimeout:
        page.screenshot(path=f"/tmp/wing_order_2fa_{account_name}.png")
        raise RuntimeError(
            f"Wingログイン失敗 ({account_name}): 2FA/CAPTCHAが必要な可能性があります。\n"
            f"一度ブラウザで手動ログインし ~/.config/khd/wing_sessions/ を再生成してください。\n"
            f"スクリーンショット: /tmp/wing_order_2fa_{account_name}.png"
        )


def ensure_logged_in(
    page: Any, context: Any,
    email: str, password: str, account_name: str
) -> None:
    page.goto(ORDER_URL, timeout=NAV_WAIT_MS)
    page.wait_for_load_state("domcontentloaded")
    if "login" in page.url.lower():
        log(f"セッション切れ: {account_name} → 再ログイン")
        do_login(page, email, password, account_name)
        save_session(context, account_name)
    else:
        log(f"セッション有効: {account_name}")


# ────── 注文一覧を取得（全ページ走査） ──────────────────────────────────
def extract_orders_from_page(page: Any) -> List[Dict[str, Any]]:
    """
    現在開いているWing注文一覧ページから注文データを抽出する。
    複数のセレクタパターンを試行する（Wingのレイアウト変更に備え）。
    戻り値: 注文dictのリスト
    """
    orders: List[Dict[str, Any]] = []

    # テーブルが描画されるまで待機
    try:
        page.wait_for_selector(
            'table tbody tr, [class*="order-row"], [class*="orderRow"]',
            timeout=10000
        )
    except PWTimeout:
        log("注文テーブルが見つかりません（ページが空か、レイアウト変更の可能性）")
        page.screenshot(path="/tmp/wing_order_empty.png")
        return orders

    # --- パターン1: 標準テーブル ---
    rows = page.locator("table tbody tr").all()
    if not rows:
        # --- パターン2: class指定 ---
        rows = page.locator('[class*="order-row"], [class*="orderRow"], [class*="OrderRow"]').all()

    for row in rows:
        try:
            order = _parse_order_row(row)
            if order:
                orders.append(order)
        except Exception as e:
            log(f"行パースエラー（スキップ）: {e}")

    return orders


def _get_cell_text(row: Any, selectors: List[str]) -> str:
    """複数セレクタを順に試してテキストを取得。失敗は空文字を返す。"""
    for sel in selectors:
        try:
            el = row.locator(sel).first
            txt = el.inner_text(timeout=2000).strip()
            if txt:
                return txt
        except Exception:
            continue
    return ""


def _parse_order_row(row: Any) -> Optional[Dict[str, Any]]:
    """
    テーブル行から注文情報を辞書に変換する。
    注文番号が取れなければNoneを返す。
    """
    # 注文番号
    order_no = _get_cell_text(row, [
        '[class*="order-number"]',
        '[class*="orderNumber"]',
        'td:nth-child(1)',
        'td:first-child',
    ])
    # 数字のみの注文番号か確認（Wingの注文番号は通常14桁以上の数字列）
    order_no_clean = re.sub(r"[^\d]", "", order_no)
    if not order_no_clean or len(order_no_clean) < 9:
        return None

    # ステータス（韓国語）
    status_kr = _get_cell_text(row, [
        '[class*="status"]',
        '[class*="order-status"]',
        'td:nth-child(2)',
    ])

    # 商品名（韓国語）
    product_name_kr = _get_cell_text(row, [
        '[class*="product-name"]',
        '[class*="productName"]',
        'td:nth-child(3)',
    ])

    # 注文者名
    buyer_name = _get_cell_text(row, [
        '[class*="buyer-name"]',
        '[class*="buyerName"]',
        'td:nth-child(5)',
    ])

    # 金額（KRW）
    price_raw = _get_cell_text(row, [
        '[class*="price"]',
        '[class*="amount"]',
        'td:nth-child(7)',
    ])
    price_krw: Optional[float] = None
    try:
        price_krw = float(re.sub(r"[^\d.]", "", price_raw)) if price_raw else None
    except ValueError:
        pass

    # 注文日時
    sale_date_raw = _get_cell_text(row, [
        '[class*="order-date"]',
        '[class*="orderDate"]',
        'td:nth-child(4)',
    ])

    # 数量
    qty_raw = _get_cell_text(row, [
        '[class*="quantity"]',
        'td:nth-child(6)',
    ])
    qty: int = 1
    try:
        qty = int(re.sub(r"[^\d]", "", qty_raw)) if qty_raw else 1
    except ValueError:
        pass

    return {
        "coupang_order_no": order_no_clean,
        "status_kr":        status_kr,
        "product_name_kr":  product_name_kr,
        "customer_name":    buyer_name,
        "sale_price_krw":   price_krw,
        "sale_date":        sale_date_raw,
        "quantity":         qty,
    }


def fetch_new_orders(page: Any, account_name: str) -> List[Dict[str, Any]]:
    """
    Wing注文ページを全ページ走査して新規注文を収集する。
    ステータスフィルタ: 신규 주문 / 결제완료 等。
    """
    all_orders: List[Dict[str, Any]] = []

    page.goto(ORDER_URL, timeout=NAV_WAIT_MS)
    page.wait_for_load_state("networkidle", timeout=20000)

    # ステータスフィルタ: 新規注文のみ表示に絞る
    for filter_label in ["신규 주문", "결제완료", "결제 완료", "新規注文"]:
        try:
            page.select_option(
                'select[name*="status"], select[id*="status"], [class*="status-filter"] select',
                label=filter_label,
                timeout=3000
            )
            log(f"注文ステータスフィルタ: {filter_label}")
            # フィルタ適用後の再読み込み待ち
            try:
                page.click(
                    'button:has-text("검색"), button:has-text("조회"), button[type="submit"]:visible',
                    timeout=3000
                )
            except PWTimeout:
                pass
            page.wait_for_load_state("networkidle", timeout=15000)
            time.sleep(1)
            break
        except Exception:
            continue

    # 全ページを走査
    page_num = 1
    while True:
        log(f"注文一覧 ページ{page_num} 取得中 ({account_name})")
        orders_on_page = extract_orders_from_page(page)
        log(f"  → {len(orders_on_page)} 件取得")
        all_orders.extend(orders_on_page)

        # 「다음」(次へ)ボタンを探してページ送り
        next_clicked = False
        for next_sel in [
            'button:has-text("다음")',        # 次へ（韓国語）
            'a:has-text("다음")',
            '[class*="next"]:visible',
            '[aria-label*="next"]:visible',
        ]:
            try:
                next_btn = page.locator(next_sel).first
                if next_btn.is_enabled(timeout=2000):
                    next_btn.click(timeout=3000)
                    page.wait_for_load_state("networkidle", timeout=15000)
                    time.sleep(1)
                    page_num += 1
                    next_clicked = True
                    break
            except Exception:
                continue

        if not next_clicked:
            log(f"最終ページ到達（合計{page_num}ページ）")
            break

        # 無限ループ保護（最大50ページ）
        if page_num > 50:
            log("⚠️ ページ数上限(50)に達しました。ループ終了。")
            break

    return all_orders


# ────── 既存DB注文番号の取得 ──────────────────────────────────────────────
def get_existing_order_nos() -> set:
    """DBに存在する全注文番号のセットを返す（重複防止用）"""
    existing: set = set()
    for status in ["new", "purchased", "haniro_registered", "shipped", "delivered", "settled"]:
        for row in get_orders_by_status(status):
            existing.add(row.get("coupang_order_no", ""))
    return existing


# ────── 販売管理表Excel への追記 ──────────────────────────────────────────
def find_sales_excel(account_num: str) -> Optional[Path]:
    """
    当月の販売管理表Excelを探して返す。
    ファイル名パターン: YYMM_販売管理表{acct}.xlsx（例: 2505_販売管理表1.xlsx）
    account_num: "1" or "2"
    """
    today = date.today()
    ym = today.strftime("%y%m")  # 例: "2505"
    pattern = str(SALES_DIR / f"{ym}_販売管理表{account_num}.xlsx")
    matches = glob.glob(pattern)
    if matches:
        return Path(matches[0])
    # 見つからない場合は前月ファイルを探す（月初の場合）
    prev = today.replace(day=1) - timedelta(days=1)
    ym_prev = prev.strftime("%y%m")
    pattern_prev = str(SALES_DIR / f"{ym_prev}_販売管理表{account_num}.xlsx")
    matches_prev = glob.glob(pattern_prev)
    if matches_prev:
        log(f"⚠️ 当月ファイルなし。前月ファイルを使用: {matches_prev[0]}")
        return Path(matches_prev[0])
    log(f"❌ 販売管理表が見つかりません: {pattern}")
    return None


def append_to_sales_excel(
    order: Dict[str, Any],
    account_num: str,
    dry_run: bool = False
) -> bool:
    """
    販売管理表ExcelのデータシートにAppendで1行追記する。
    列マッピング（0始まり / ヘッダー行で確認済み）:
      A(0)=進捗, B(1)=注文番号, C(2)=プラットフォーム,
      F(5)=注文者名, G(6)=商品名KR, H(7)=日本語商品名, I(8)=ASIN,
      U(20)=Sale Date, V(21)=発送予定日, W(22)=数量, X(23)=Sale Price KRW,
      AP(41)=郵便番号, AQ(42)=住所, AR(43)=携帯電話番号,
      AS(44)=個人通関番号, AT(45)=配送メッセージ
    """
    excel_path = find_sales_excel(account_num)
    if not excel_path:
        return False

    if dry_run:
        log(f"[DRY RUN] Excel追記スキップ: {excel_path.name} / 注文{order['coupang_order_no']}")
        return True

    try:
        wb = openpyxl.load_workbook(str(excel_path))
        ws = wb[wb.sheetnames[0]]

        # 発送予定日 = Sale Date + 4日
        sale_date_str = order.get("sale_date", "")
        ship_deadline = ""
        try:
            if sale_date_str:
                # "2025-05-01 00:39:11" 形式に対応
                sale_dt = datetime.strptime(sale_date_str[:10], "%Y-%m-%d")
                ship_deadline = (sale_dt + timedelta(days=4)).strftime("%Y-%m-%d")
        except ValueError:
            pass

        # 追記行を構築（最大 AT列=45番まで、46セル）
        # ws.append() は row のタプル/リストで追記する
        # 列数は AS(44)=個人通関番号まで AT(45)=配送メッセージまで = 46要素
        row_data = [""] * 46

        row_data[COL_STATUS]       = "注文済"
        row_data[COL_ORDER_NO]     = order.get("coupang_order_no", "")
        row_data[COL_PLATFORM]     = "クーパン"
        row_data[COL_BUYER_NAME]   = order.get("customer_name", "")
        row_data[COL_PRODUCT_KR]   = order.get("product_name_kr", "")
        row_data[COL_PRODUCT_JP]   = order.get("product_name_jp", "")
        row_data[COL_ASIN]         = order.get("asin", "")
        row_data[COL_SALE_DATE]    = sale_date_str
        row_data[COL_SHIP_DL]      = ship_deadline
        row_data[COL_QTY]          = order.get("quantity", 1)
        row_data[COL_SALE_PRICE]   = order.get("sale_price_krw", "")
        row_data[COL_POSTAL]       = order.get("postal_code", "")
        row_data[COL_ADDRESS]      = order.get("address", "")
        row_data[COL_PHONE]        = order.get("phone", "")
        row_data[COL_CUSTOMS_NO]   = order.get("customs_no", "")
        row_data[COL_DELIVERY_MSG] = order.get("delivery_msg", "")

        ws.append(row_data)
        wb.save(str(excel_path))
        log(f"✅ Excel追記完了: {excel_path.name} / 注文{order['coupang_order_no']}")
        return True

    except Exception as e:
        log(f"❌ Excel追記エラー ({excel_path.name}): {e}")
        traceback.print_exc()
        return False


# ────── 1アカウント分の処理 ──────────────────────────────────────────────
def process_account(
    pw: Any,
    account_cfg: Dict[str, Any],
    existing_order_nos: set,
    dry_run: bool = False
) -> List[Dict[str, Any]]:
    """
    1アカウントの新規注文を検出してDBとExcelに記録する。
    戻り値: 新規検出した注文のリスト
    """
    name    = account_cfg["name"]
    email   = account_cfg["login_email"]
    pwd     = account_cfg["login_password"]

    # アカウント番号を "クーパン1" → "1" のように抽出
    acct_num_match = re.search(r"\d", name)
    account_num = acct_num_match.group(0) if acct_num_match else "1"

    log(f"=== {name} 処理開始 ===")
    SESS_DIR.mkdir(parents=True, exist_ok=True)

    browser = pw.chromium.launch(headless=True)
    context = browser.new_context()
    load_session(context, name)
    page = context.new_page()

    new_orders: List[Dict[str, Any]] = []

    try:
        ensure_logged_in(page, context, email, pwd, name)
        raw_orders = fetch_new_orders(page, name)

        for raw in raw_orders:
            order_no = raw.get("coupang_order_no", "")
            if not order_no:
                continue

            # 重複チェック
            if order_no in existing_order_nos:
                log(f"  スキップ（DB既存）: {order_no}")
                continue

            # DB用データに整形
            db_data: Dict[str, Any] = {
                "coupang_order_no": order_no,
                "account":         name,
                "status":          "new",
                "product_name_kr": raw.get("product_name_kr", ""),
                "customer_name":   raw.get("customer_name", ""),
                "sale_price_krw":  raw.get("sale_price_krw"),
                "sale_date":       raw.get("sale_date", ""),
                "quantity":        raw.get("quantity", 1),
            }

            if not dry_run:
                upsert_order(db_data)
                existing_order_nos.add(order_no)  # 同一実行内の二重挿入防止
                log(f"  DB記録: {order_no}")
            else:
                log(f"  [DRY RUN] DB記録スキップ: {order_no} / {raw.get('product_name_kr','')}")

            # Excel追記
            append_to_sales_excel(db_data, account_num, dry_run=dry_run)

            new_orders.append(db_data)

    finally:
        context.close()
        browser.close()

    log(f"=== {name} 処理完了 / 新規: {len(new_orders)} 件 ===")
    return new_orders


# ────── メインポーリングループ ────────────────────────────────────────────
def run_once(dry_run: bool = False) -> None:
    """1回分のポーリング処理を実行する"""
    log("ポーリング開始")
    cfg = load_config()
    existing_order_nos = get_existing_order_nos()
    log(f"DB既存注文数: {len(existing_order_nos)} 件")

    all_new: List[Dict[str, Any]] = []
    errors: List[str] = []
    account_summary: Dict[str, int] = {}

    try:
        with sync_playwright() as pw:
            for acct in cfg.get("accounts", []):
                try:
                    new_orders = process_account(pw, acct, existing_order_nos, dry_run=dry_run)
                    all_new.extend(new_orders)
                    account_summary[acct["name"]] = len(new_orders)
                except Exception as e:
                    err_msg = f"{acct['name']}: {e}"
                    log(f"❌ {err_msg}")
                    errors.append(err_msg)
    except Exception as e:
        err_msg = f"Playwright起動エラー: {e}"
        log(f"❌ {err_msg}")
        errors.append(err_msg)

    # 通知
    if errors:
        err_txt = "\n".join(f"・{e[:100]}" for e in errors)
        notify(
            f"注文モニター エラー発生\n"
            f"新規注文: {len(all_new)} 件\n"
            f"エラー:\n{err_txt}\n"
            f"→ /tmp/wing_order_*.png を確認",
            urgent=True
        )
    elif all_new:
        summary_txt = "\n".join(f"・{name}: {cnt} 件" for name, cnt in account_summary.items())
        notify(
            f"新規注文 {len(all_new)} 件を検出\n"
            f"{summary_txt}\n"
            f"→ 販売管理表に記録済み"
        )
        log(f"✅ 新規注文 {len(all_new)} 件を検出・記録")
    else:
        log("新規注文なし")

    log("ポーリング完了")


def main() -> None:
    parser = argparse.ArgumentParser(description="Coupang Wing 新規注文ポーリングモニター")
    parser.add_argument(
        "--dry-run", action="store_true",
        help="DBに書かず標準出力のみ（テスト確認用）"
    )
    parser.add_argument(
        "--check-once", action="store_true",
        help="1回だけ実行して終了（cronから呼ぶ時用）"
    )
    args = parser.parse_args()

    if args.check_once:
        run_once(dry_run=args.dry_run)
        return

    # 常駐ポーリングモード（30分間隔）
    log(f"常駐モード起動（{POLL_INTERVAL_SEC // 60}分間隔）。Ctrl+C で停止。")
    while True:
        try:
            run_once(dry_run=args.dry_run)
        except KeyboardInterrupt:
            log("停止シグナル受信。終了します。")
            break
        except Exception as e:
            log(f"❌ 予期しないエラー: {e}")
            traceback.print_exc()
            notify(
                f"注文モニター 予期しないエラー\n{str(e)[:200]}",
                urgent=True
            )

        log(f"次回ポーリングまで {POLL_INTERVAL_SEC // 60} 分待機...")
        try:
            time.sleep(POLL_INTERVAL_SEC)
        except KeyboardInterrupt:
            log("停止シグナル受信。終了します。")
            break


if __name__ == "__main__":
    main()

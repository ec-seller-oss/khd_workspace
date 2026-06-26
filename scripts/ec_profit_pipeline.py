# -*- coding: utf-8 -*-
"""
EC粗利パイプライン v1 (01_経営管理 CFO)
韓国クーパン輸出: 月次販売管理表24本(クーパン1/2 × 2501-2512)を集計し
売上-原価=粗利を確定。売れ筋商品分析・出品戦略の数値指針まで自動抽出。
データ源: ローカル同期Drive (.xlsx) / 行=1注文。
計算: 純売上(AB=純上高×為替, 手数料控除後の入金近似) - 総原価(T) = 粗利
"""
import openpyxl, csv, glob, re, os
from datetime import datetime
from collections import defaultdict

DIR = "/Users/kikuchikenta/Library/CloudStorage/GoogleDrive-ec-seller@kikuchi-hd.net/共有ドライブ/01_個人/2025_帳票、明細/韓国輸出売上"
OUT_CSV = "/Users/kikuchikenta/01_honbu_docs_automation/ec_orders_consolidated.csv"

# 列インデックス(0始まり)
C_ORDER, C_PLAT, C_NAME, C_ASIN, C_BUYDATE = 1, 2, 7, 8, 11
C_COST, C_SALEDATE, C_QTY, C_GROSSW, C_FEERATE, C_NETW, C_FX, C_JPY = 19, 20, 22, 23, 24, 25, 26, 27
C_SHIP, C_REFUND = 29, 35   # AD配送料, AJ還付
C_STATUS = 0                # A進捗
# 売上計上する進捗（履行済み）／除外する進捗（売上でない）
VALID_STATUS = {"発送済", "代行登録済み", "注文済"}
EXCLUDE_STATUS = {"キャンセル", "在庫切れ", "返品", "発送登録不可"}

def num(v):
    if v is None or v == "": return None
    if isinstance(v,(int,float)): return float(v)
    try: return float(str(v).replace(",","").replace("¥","").strip())
    except: return None

def is_order(b, u):
    # 主キー=注文番号(9桁以上の数字) + 有効なSale Date(datetime or 文字列日付)
    s = str(b).strip().replace(".0","")
    if not (s.isdigit() and len(s) >= 9): return False
    if isinstance(u, datetime): return True
    return bool(re.match(r"\d{4}-\d{2}-\d{2}", str(u).strip()))

orders = []
cancelled = []
for path in sorted(glob.glob(f"{DIR}/25*_販売管理表*.xlsx")):
    fn = os.path.basename(path)
    m = re.match(r"(\d{2})(\d{2})_販売管理表([12])", fn)
    if not m: continue
    yy, mm, acct = m.group(1), m.group(2), m.group(3)
    ym = f"20{yy}-{mm}"
    account = f"クーパン{acct}"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    cnt = 0
    for row in ws.iter_rows(min_row=7, values_only=True):
        if len(row) <= C_JPY: continue
        b, u = row[C_ORDER], row[C_SALEDATE]
        if not is_order(b, u): continue
        status = str(row[C_STATUS] or "").strip()
        cost0 = num(row[C_COST]) or 0
        # キャンセル等は売上から除外。仕入済みなら原価リスクとして別計上
        if status in EXCLUDE_STATUS or status not in VALID_STATUS:
            cancelled.append({"年月": ym, "アカウント": account, "進捗": status,
                              "注文番号": str(b).replace(".0",""),
                              "商品名": str(row[C_NAME] or "")[:40],
                              "原価円": round(cost0)})
            continue
        cost = num(row[C_COST]) or 0
        grossw = num(row[C_GROSSW]) or 0
        fx = num(row[C_FX])
        netw = num(row[C_NETW])
        jpy = num(row[C_JPY])
        # 純売上(円)=AB。無ければ 純上高W×為替、それも無ければ 商品総額×0.89×為替
        if jpy is None and netw is not None and fx is not None: jpy = netw*fx
        if jpy is None and fx is not None: jpy = grossw*0.89*fx
        if jpy is None: continue
        gross_jpy = grossw*fx if fx else jpy/0.89
        fee_jpy = gross_jpy - jpy
        ship = num(row[C_SHIP]) or 0
        qty = num(row[C_QTY]) or 1
        profit = jpy - cost                # 粗利(売上総利益)
        op_profit = jpy - cost - ship      # 営業利益(配送料控除後)
        orders.append({
            "年月": ym, "アカウント": account,
            "プラットフォーム": str(row[C_PLAT] or ""),
            "注文番号": str(b).replace(".0",""),
            "商品名": str(row[C_NAME] or "")[:40],
            "ASIN": str(row[C_ASIN] or ""),
            "数量": int(qty),
            "総売上円": round(gross_jpy), "手数料円": round(fee_jpy),
            "純売上円": round(jpy), "原価円": round(cost),
            "配送料円": round(ship), "粗利円": round(profit),
            "営業利益円": round(op_profit),
            "粗利率": round(profit/jpy,4) if jpy else 0,
        })
        cnt += 1
    wb.close()
    print(f"  {fn:30s} {account} {ym}: {cnt}件")

print(f"\n総注文件数: {len(orders)}")

# CSV出力
if orders:
    with open(OUT_CSV,"w",newline="",encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=list(orders[0].keys()))
        w.writeheader(); w.writerows(orders)
    print(f"明細CSV出力: {OUT_CSV}")

# ===== 月次×アカウント集計 =====
def agg(rows, *keys):
    d = defaultdict(lambda: dict(件数=0,数量=0,総売上=0,手数料=0,純売上=0,原価=0,配送料=0,粗利=0,営業利益=0))
    for r in rows:
        k = tuple(r[x] for x in keys)
        a = d[k]
        a["件数"]+=1; a["数量"]+=r["数量"]; a["総売上"]+=r["総売上円"]; a["手数料"]+=r["手数料円"]
        a["純売上"]+=r["純売上円"]; a["原価"]+=r["原価円"]; a["配送料"]+=r["配送料円"]
        a["粗利"]+=r["粗利円"]; a["営業利益"]+=r["営業利益円"]
    return d

print("\n"+"="*100)
print("【月次×アカウント】純売上=手数料控除後入金 / 粗利=純売上-原価")
print("="*100)
print(f"{'年月':9s}{'アカ':9s}{'件数':>5s}{'純売上':>11s}{'原価':>11s}{'粗利':>11s}{'粗利率':>8s}{'営業利益':>11s}")
ma = agg(orders,"年月","アカウント")
for k in sorted(ma):
    a=ma[k]; gm=a['粗利']/a['純売上'] if a['純売上'] else 0
    print(f"{k[0]:9s}{k[1]:9s}{a['件数']:>5d}{a['純売上']:>11,d}{a['原価']:>11,d}{a['粗利']:>11,d}{gm:>7.1%}{a['営業利益']:>11,d}")

print("\n"+"="*100)
print("【年間合計：アカウント別】")
print("="*100)
ya = agg(orders,"アカウント")
tot=dict(件数=0,純売上=0,原価=0,粗利=0,営業利益=0,手数料=0)
for k in sorted(ya):
    a=ya[k]; gm=a['粗利']/a['純売上'] if a['純売上'] else 0
    print(f"{k[0]:10s} 件数{a['件数']:>4d} 純売上{a['純売上']:>10,d} 原価{a['原価']:>10,d} 手数料{a['手数料']:>9,d} 粗利{a['粗利']:>10,d}({gm:.1%}) 営業利益{a['営業利益']:>10,d}")
    for x in tot: tot[x]+=a[x]
gm=tot['粗利']/tot['純売上'] if tot['純売上'] else 0
print(f"{'合計(1+2)':10s} 件数{tot['件数']:>4d} 純売上{tot['純売上']:>10,d} 原価{tot['原価']:>10,d} 手数料{tot['手数料']:>9,d} 粗利{tot['粗利']:>10,d}({gm:.1%}) 営業利益{tot['営業利益']:>10,d}")

# ===== キャンセル/返品の原価リスク（売上から除外した分） =====
print("\n"+"="*100)
print("【除外行：キャンセル/返品/在庫切れ等】売上計上せず・仕入済みなら原価リスク")
print("="*100)
cstat = defaultdict(lambda: dict(件数=0, 原価=0))
for r in cancelled:
    cstat[r["進捗"]]["件数"]+=1; cstat[r["進捗"]]["原価"]+=r["原価円"]
ctot_n=ctot_c=0
for k in sorted(cstat, key=lambda x:-cstat[x]["件数"]):
    print(f"  {k:12s} {cstat[k]['件数']:>4d}件  原価計 {cstat[k]['原価']:>9,d}円")
    ctot_n+=cstat[k]['件数']; ctot_c+=cstat[k]['原価']
print(f"  {'除外計':12s} {ctot_n:>4d}件  原価計 {ctot_c:>9,d}円  ← うちキャンセルで仕入済みは滞留/損失リスク")

# Q1照合（MF海外売上Q1=2,796,335 と比較）
q1 = sum(r["純売上円"] for r in orders if r["年月"] in ("2025-01","2025-02","2025-03"))
q1g = sum(r["総売上円"] for r in orders if r["年月"] in ("2025-01","2025-02","2025-03"))
print(f"\n【Q1照合】純売上(手数料控除後)={q1:,}円 / 総売上(手数料前)={q1g:,}円 / MF海外売上Q1=2,796,335円")

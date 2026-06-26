# -*- coding: utf-8 -*-
"""
EC精算実額 記帳照合エンジン (01_経営管理 CFO) — 記帳の正確性の追求
販売管理表(見込み: Z=清算予定額KRW, 11%固定手数料) ↔ 売掛金確定(実精算: 精算金額KRW) を
注文番号キーでネット相殺(返金リバーサル±を相殺)して突合し、記帳に使える「実額売上」を出す。

入力:
  - 販売管理表 25*_販売管理表{1,2}.xlsx (見込み・通年)
  - 最新の *_売掛金.xlsx (確定=実精算スナップショット) ← フォルダ内で日付プレフィクス最大のものを自動採用
出力:
  - ec_settlement_recon.csv  … 注文単位 (見込みZ / 実精算 / 差 / 突合区分)
  - コンソール要約           … カバレッジ・KRW一致・実額売上・要確認フラグ

※確定ファイルは月次DLの追記でリバーサル(+X/−X)や重複が混じる → 注文番号で合算(ネット)するのが正。
※円記帳には 精算KRW×実入金レート が必要。本v1は管理表の行別為替(AA)で近似JPYを併記。
"""
import openpyxl, glob, re, os, csv
from datetime import datetime
from collections import defaultdict

DIR = "/Users/kikuchikenta/Library/CloudStorage/GoogleDrive-ec-seller@kikuchi-hd.net/共有ドライブ/01_個人/2025_帳票、明細/韓国輸出売上"
OUT = "/Users/kikuchikenta/01_honbu_docs_automation/ec_settlement_recon.csv"
VALID = {"発送済", "代行登録済み", "注文済"}

def norm(x):
    s = str(x).strip(); s = re.sub(r"\.0$", "", s); return s.rstrip(".")
def num(v):
    if v is None or v == "": return None
    if isinstance(v, (int, float)): return float(v)
    try: return float(str(v).replace(",", "").replace("¥", "").strip())
    except: return None
def is_ord(s):
    s = norm(s); return s.isdigit() and len(s) >= 9

# ===== 1. 販売管理表(見込み) =====
def load_mgmt():
    mgmt = {}
    for path in sorted(glob.glob(f"{DIR}/25*_販売管理表*.xlsx")):
        fn = os.path.basename(path)
        m = re.match(r"(\d{2})(\d{2})_販売管理表([12])", fn)
        if not m: continue
        ym = f"20{m.group(1)}-{m.group(2)}"; acct = f"クーパン{m.group(3)}"
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=7, values_only=True):
            if len(row) <= 27: continue
            b = row[1]
            if not is_ord(b): continue
            u = row[20]
            if not (isinstance(u, datetime) or re.match(r"\d{4}-\d{2}-\d{2}", str(u).strip())): continue
            on = norm(b)
            mgmt[on] = dict(status=str(row[0] or "").strip(), ym=ym, account=acct,
                            name=str(row[7] or "")[:30],
                            Zkrw=num(row[25]) or 0, Xkrw=num(row[23]) or 0,
                            fx=num(row[26]), AB=num(row[27]) or 0, cost=num(row[19]) or 0)
        wb.close()
    return mgmt

# ===== 2. 売掛金確定(実精算) — 最新ファイルを自動採用・注文番号でネット =====
def latest_urikake():
    files = glob.glob(f"{DIR}/*_売掛金.xlsx")
    if not files: return None
    def keyf(p):
        m = re.match(r"(\d{6})_売掛金", os.path.basename(p)); return m.group(1) if m else "000000"
    return sorted(files, key=keyf)[-1]

def colmap(header):
    idx = {}
    for i, h in enumerate(header):
        hs = str(h or "").strip()
        if hs in ("注文番号", "A1") and "order" not in idx: idx["order"] = i
        if hs == "精算金額": idx["seika"] = i
        if hs == "販売額": idx["hanbai"] = i
        if hs == "販売手数料": idx["fee"] = i
        if hs == "キャンセル完了日": idx["cancel"] = i
    # フォールバック(既知レイアウト)
    idx.setdefault("order", 0); idx.setdefault("seika", 17)
    idx.setdefault("hanbai", 9); idx.setdefault("fee", 13)
    return idx

def load_settle(path):
    settle = defaultdict(lambda: dict(精算=0.0, 販売額=0.0, 手数料=0.0, 行数=0))
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    target = [s for s in wb.sheetnames if s.endswith("確定") and not s.endswith("未確定")]
    for sn in target:
        ws = wb[sn]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None: continue
        ix = colmap(header)
        for row in rows:
            if len(row) <= ix["seika"]: continue
            b = row[ix["order"]]
            if not is_ord(b): continue
            on = norm(b); s = settle[on]
            s["精算"] += num(row[ix["seika"]]) or 0
            s["販売額"] += num(row[ix["hanbai"]]) or 0
            s["手数料"] += num(row[ix["fee"]]) or 0
            s["行数"] += 1
    wb.close()
    return dict(settle), target

def main():
    mgmt = load_mgmt()
    mgmt_valid = {k: v for k, v in mgmt.items() if v["status"] in VALID}
    upath = latest_urikake()
    print(f"売掛金(確定)ファイル: {os.path.basename(upath)}")
    settle, sheets = load_settle(upath)
    print(f"  対象シート: {sheets}")
    print(f"販売管理表: 全{len(mgmt)}件 / 売上計上valid {len(mgmt_valid)}件")
    print(f"売掛金確定: ネット後ユニーク注文 {len(settle)}件")

    mset, sset = set(mgmt_valid), set(settle)
    both, only_m, only_s = mset & sset, mset - sset, sset - mset

    sum_z = sum_s = exact = near = 0; rows = []
    for k in both:
        z = mgmt_valid[k]["Zkrw"]; s = settle[k]["精算"]; fx = mgmt_valid[k]["fx"]
        sum_z += z; sum_s += s
        d = z - s
        if abs(d) < 1: exact += 1
        elif abs(d) <= max(50, abs(s) * 0.01): near += 1
        seika_jpy = round(s * fx) if fx else ""
        rows.append(dict(注文番号=k, 区分="照合OK", アカウント=mgmt_valid[k]["account"], 年月=mgmt_valid[k]["ym"],
                         商品名=mgmt_valid[k]["name"], 見込みZ_KRW=round(z), 実精算_KRW=round(s),
                         差KRW=round(d), 行別為替=fx, 実精算_円=seika_jpy, 見込みAB_円=round(mgmt_valid[k]["AB"])))
    diff = sum_z - sum_s
    print("\n=== 実精算 vs 見込み (照合できた {0}件) ===".format(len(both)))
    print(f"  KRW: 見込みZ {sum_z:,.0f} / 実精算 {sum_s:,.0f} / 差 {diff:,.0f}W ({diff/sum_s*100 if sum_s else 0:+.2f}%)")
    print(f"  一致: 完全{exact} / 近似{near} / その他{len(both)-exact-near}")
    print(f"  → 実精算が見込みより {(-diff)/sum_s*100 if sum_s else 0:+.2f}% (手数料11%固定の過大計上ぶん)")

    # 要確認: キャンセルstatusだが実精算あり(ネット後>0)
    cancelled_paid = [(k, settle[k]["精算"]) for k in only_s if k in mgmt and settle[k]["精算"] > 1]
    print(f"\n=== 要確認フラグ ===")
    print(f"  管理表のみ(8月以降の精算待ち等): {len(only_m)}件 / 清算予定 {sum(mgmt_valid[k]['Zkrw'] for k in only_m):,.0f}W")
    print(f"  確定のみ(管理表validに無い): {len(only_s)}件 / 実精算 {sum(settle[k]['精算'] for k in only_s):,.0f}W")
    print(f"   └ うちキャンセル等statusなのに実入金あり: {len(cancelled_paid)}件 / {sum(v for _,v in cancelled_paid):,.0f}W (集計漏れ候補)")
    for k, v in sorted(cancelled_paid, key=lambda x: -x[1])[:8]:
        print(f"      注文{k}: status='{mgmt[k]['status']}' 実精算={v:,.0f}W")
    for k in only_m:
        m = mgmt_valid[k]
        rows.append(dict(注文番号=k, 区分="精算待ち(確定なし)", アカウント=m["account"], 年月=m["ym"],
                         商品名=m["name"], 見込みZ_KRW=round(m["Zkrw"]), 実精算_KRW="", 差KRW="",
                         行別為替=m["fx"], 実精算_円="", 見込みAB_円=round(m["AB"])))
    for k, v in cancelled_paid:
        rows.append(dict(注文番号=k, 区分="要確認(除外statusだが入金)", アカウント=mgmt.get(k,{}).get("account",""),
                         年月=mgmt.get(k,{}).get("ym",""), 商品名=mgmt.get(k,{}).get("name",""),
                         見込みZ_KRW="", 実精算_KRW=round(v), 差KRW="", 行別為替="", 実精算_円="", 見込みAB_円=""))

    with open(OUT, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["注文番号","区分","アカウント","年月","商品名","見込みZ_KRW","実精算_KRW","差KRW","行別為替","実精算_円","見込みAB_円"])
        w.writeheader(); w.writerows(rows)
    print(f"\n照合明細CSV: {OUT}")

if __name__ == "__main__":
    main()

# -*- coding: utf-8 -*-
"""
EC精算実額 記帳照合エンジン v2 (01_経営管理 CFO) — 通年カバレッジ版
v1からの進化点:
  - 精算源を「MSF生エクスポート18本(Wing決済確定DL) + 前半タレ(売掛金スナップショット)」の合算に変更。
    オンライン売掛金シートは1〜6月だけ完備・8月以降が空だった(=v1が53%しか照合できなかった原因)。
    Wingから2ヶ月窓で落としたMSF18本が6月〜2026年をカバー → 合算で通年95%まで到達。
  - 全ファイル横断の「全列重複ガード」で再DL行を完全排除(±リバーサル/調整行は注文番号でネット)。
  - MSF(生エクスポート)を真実源として優先。MSF未収載の前半(1〜5月)のみスナップショットで補完。

精算源の構造(2026-05-30実測):
  MSF18本 정산예정일 2025-08-22〜2026-05-26 / 결제완료일 2025-06〜2026-04 (口座1:10本 口座2:8本)
  スナップショット 250811_売掛金.xlsx = 2025-08-11時点 → 1〜7月の実精算を保持

列レイアウト(MSF韓国語・売掛金日本語とも同一25列順):
  0=注文番号(주문번호) 9=販売額(판매액) 13=販売手数料(판매수수료) 17=精算金額(정산금액)
  19=決済完了日 22=キャンセル完了日 24=精算予定日

出力:
  - ec_settlement_recon_v2.csv … 注文単位(見込みZ / 実精算 / 差 / 区分 / JPY併記)
  - コンソール要約 … 月別カバレッジ・通年KRW一致・実額売上(JPY)・要確認フラグ
"""
import openpyxl, glob, os, re, csv, warnings
from datetime import datetime
from collections import defaultdict
warnings.filterwarnings("ignore")

DIR = "/Users/kikuchikenta/Library/CloudStorage/GoogleDrive-ec-seller@kikuchi-hd.net/共有ドライブ/01_個人/2025_帳票、明細/韓国輸出売上"
MSFDIR = "/Users/kikuchikenta/Library/CloudStorage/GoogleDrive-ec-seller@kikuchi-hd.net/共有ドライブ/01_個人/2025_帳票、明細/韓国輸出売上/_精算実額MSF"
OUT = "/Users/kikuchikenta/01_honbu_docs_automation/ec_settlement_recon_v2.csv"
VALID = {"発送済", "代行登録済み", "注文済"}
# 精算源の列インデックス(MSF/売掛金 共通)
ORDER_I, HANBAI_I, FEE_I, SEIKA_I = 0, 9, 13, 17

def norm(x):
    s = str(x).strip(); s = re.sub(r"\.0$", "", s); return s.rstrip(".")
def num(v):
    if v is None or v == "": return None
    if isinstance(v, (int, float)): return float(v)
    try: return float(str(v).replace(",", "").replace("¥", "").strip())
    except: return None
def is_ord(s):
    s = norm(s); return s.isdigit() and len(s) >= 9

# ===== 1. 販売管理表(見込み) =====
def load_mgmt():
    mgmt = {}
    for path in sorted(glob.glob(f"{DIR}/25*_販売管理表*.xlsx")):
        m = re.match(r"(\d{2})(\d{2})_販売管理表([12])", os.path.basename(path))
        if not m: continue
        ym = f"20{m.group(1)}-{m.group(2)}"; acct = f"クーパン{m.group(3)}"
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=7, values_only=True):
            if len(row) <= 27: continue
            b = row[1]
            if not is_ord(b): continue
            u = row[20]
            if not (isinstance(u, datetime) or re.match(r"\d{4}-\d{2}-\d{2}", str(u).strip())): continue
            mgmt[norm(b)] = dict(status=str(row[0] or "").strip(), ym=ym, account=acct,
                                 name=str(row[7] or "")[:30],
                                 Zkrw=num(row[25]) or 0, Xkrw=num(row[23]) or 0,
                                 fx=num(row[26]), AB=num(row[27]) or 0, cost=num(row[19]) or 0)
        wb.close()
    return mgmt

# ===== 2. 精算実額 — MSF18本+前半タレを横断ネット(全列重複ガード) =====
def add_rows(ws, settle, gseen):
    rows = ws.iter_rows(values_only=True); next(rows, None)
    for r in rows:
        if len(r) <= SEIKA_I or not is_ord(r[ORDER_I]): continue
        key = tuple(str(c) for c in r[:25])
        if key in gseen: continue            # 全列一致=再DL/重複ペースト→1回だけ
        gseen.add(key)
        on = norm(r[ORDER_I]); s = settle[on]
        s["精算"] += num(r[SEIKA_I]) or 0
        s["販売額"] += num(r[HANBAI_I]) or 0
        s["手数料"] += num(r[FEE_I]) or 0
        s["行数"] += 1

def latest_snapshot():
    files = glob.glob(f"{DIR}/*_売掛金.xlsx")
    if not files: return None
    keyf = lambda p: (re.match(r"(\d{6})_売掛金", os.path.basename(p)) or [None,"000000"])[1] \
        if re.match(r"(\d{6})_売掛金", os.path.basename(p)) else "000000"
    return sorted(files, key=lambda p: re.sub(r"\D","",os.path.basename(p))[:6] or "000000")[-1]

def load_settle():
    # (A) MSF生エクスポート18本 = 真実源(6月〜2026)。全横断ガード。
    msf = defaultdict(lambda: dict(精算=0.0, 販売額=0.0, 手数料=0.0, 行数=0)); gseen = set()
    msf_files = sorted(glob.glob(f"{MSFDIR}/MSF_PAYMENT_REVENUE_DETAIL*.xlsx"))
    for p in msf_files:
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
        add_rows(wb[wb.sheetnames[0]], msf, gseen); wb.close()
    # (B) 前半タレ = 売掛金スナップショット(1〜7月)。独立ガードでネット。
    snap = latest_snapshot(); snap_settle = defaultdict(lambda: dict(精算=0.0, 販売額=0.0, 手数料=0.0, 行数=0))
    sseen = set()
    if snap:
        wb = openpyxl.load_workbook(snap, read_only=True, data_only=True)
        for sn in [s for s in wb.sheetnames if s.endswith("確定") and not s.endswith("未確定")]:
            add_rows(wb[sn], snap_settle, sseen)
        wb.close()
    # (C) 合算: MSF優先、未収載のみスナップショットで補完
    settle = {k: dict(v, src="MSF") for k, v in msf.items()}
    for on, v in snap_settle.items():
        if on not in settle: settle[on] = dict(v, src="snap")
    return settle, os.path.basename(snap) if snap else "-", len(msf_files), len(msf), len(snap_settle)

def main():
    mgmt = load_mgmt()
    mgmt_valid = {k: v for k, v in mgmt.items() if v["status"] in VALID}
    settle, snapname, nmsf, n_msf_ord, n_snap_ord = load_settle()
    print(f"精算源: MSF生エクスポート{nmsf}本({n_msf_ord}注文) + スナップショット{snapname}({n_snap_ord}注文)")
    print(f"  → 合算ユニーク注文 {len(settle)}件")
    print(f"販売管理表: 全{len(mgmt)}件 / 売上計上valid {len(mgmt_valid)}件")

    mset, sset = set(mgmt_valid), set(settle)
    both, only_m, only_s = mset & sset, mset - sset, sset - mset

    # 月別カバレッジ
    cov = defaultdict(lambda: [0, 0])
    for k, v in mgmt_valid.items():
        cov[v["ym"]][0] += 1
        if k in settle: cov[v["ym"]][1] += 1
    print(f"\n=== 月別カバレッジ(valid→実精算照合) 全体 {len(both)}/{len(mgmt_valid)} ({len(both)/len(mgmt_valid)*100:.0f}%) ===")
    for ym in sorted(cov):
        v, m = cov[ym]; bar = "✅" if m == v else ("🟡" if v and m/v >= 0.8 else "❌")
        print(f"  {ym}: {m:3}/{v:3} ({m/v*100 if v else 0:3.0f}%) {bar}")

    sum_z = sum_s = exact = near = 0; sum_seika_jpy = sum_ab_jpy = 0.0; rows = []
    for k in both:
        z = mgmt_valid[k]["Zkrw"]; s = settle[k]["精算"]; fx = mgmt_valid[k]["fx"]
        sum_z += z; sum_s += s; d = z - s
        if abs(d) < 1: exact += 1
        elif abs(d) <= max(50, abs(s) * 0.01): near += 1
        seika_jpy = round(s * fx) if fx else ""
        if fx: sum_seika_jpy += s * fx
        sum_ab_jpy += mgmt_valid[k]["AB"]
        rows.append(dict(注文番号=k, 区分="照合OK", アカウント=mgmt_valid[k]["account"], 年月=mgmt_valid[k]["ym"],
                         商品名=mgmt_valid[k]["name"], 見込みZ_KRW=round(z), 実精算_KRW=round(s),
                         差KRW=round(d), 精算源=settle[k]["src"], 行別為替=fx,
                         実精算_円=seika_jpy, 見込みAB_円=round(mgmt_valid[k]["AB"])))
    diff = sum_s - sum_z
    print(f"\n=== 通年 実精算 vs 見込み (照合 {len(both)}件) ===")
    print(f"  KRW: 見込みZ {sum_z:,.0f} / 実精算 {sum_s:,.0f} / 差 {diff:+,.0f}W ({diff/sum_z*100 if sum_z else 0:+.2f}%)")
    print(f"  一致: 完全{exact} / 近似{near} / その他{len(both)-exact-near}")
    print(f"  JPY: 実精算(実額×行別為替) ¥{sum_seika_jpy:,.0f} / 見込みAB ¥{sum_ab_jpy:,.0f} / 差 ¥{sum_seika_jpy-sum_ab_jpy:+,.0f}")
    print(f"  → 実精算が見込みより {diff/sum_z*100 if sum_z else 0:+.2f}%（11%固定手数料の過大計上ぶん＝ダッシュボード粗利は控えめ）")

    # 要確認: 除外statusだが実精算あり → 自動2分類
    #   A=売上漏れ候補(実精算が商品売上規模=管理表statusの誤り。status修正で売上計上)
    #   B=キャンセル補償(定額補償 精算≈29,100W。商品売上ではなく雑収入。status'キャンセル'は正しい)
    HOSHO = 29100          # クーパン キャンセル定額補償の精算額
    THRESH = 40000         # これ未満は補償、以上は売上漏れ候補
    def classify(v):
        return "B_キャンセル補償(雑収入)" if v < THRESH else "A_売上漏れ候補(status要修正)"
    cancelled_paid = [(k, settle[k]["精算"]) for k in only_s if k in mgmt and settle[k]["精算"] > 1]
    leak = [(k, v) for k, v in cancelled_paid if classify(v).startswith("A")]
    hosho = [(k, v) for k, v in cancelled_paid if classify(v).startswith("B")]
    print(f"\n=== 要確認フラグ ===")
    print(f"  管理表のみ(実精算なし=精算保留/未精算): {len(only_m)}件 / 清算予定 {sum(mgmt_valid[k]['Zkrw'] for k in only_m):,.0f}W")
    print(f"  精算のみ(管理表validに無い): {len(only_s)}件")
    leak_uri = leak_genka = 0.0   # 売上漏れの粗利インパクト(JPY)
    print(f"   └ A.売上漏れ候補(status要修正): {len(leak)}件 / {sum(v for _,v in leak):,.0f}W ≈ ¥{sum(v for _,v in leak)*0.105:,.0f} ★売上計上もれ")
    for k, v in sorted(leak, key=lambda x: -x[1]):
        fx = mgmt[k].get("fx") or 0.105/1.0   # 行別為替が無ければ概算
        uri = v * (mgmt[k].get("fx") or 0.105); genka = mgmt[k].get("cost") or 0
        leak_uri += uri; leak_genka += genka
        print(f"      注文{k}: status='{mgmt[k]['status']}' 実精算={v:,.0f}W(¥{uri:,.0f}) 原価¥{genka:,.0f} 粗利¥{uri-genka:+,.0f} '{mgmt[k]['name'][:18]}'")
    print(f"      → A合計: 売上+¥{leak_uri:,.0f} − 原価¥{leak_genka:,.0f} = 粗利インパクト ¥{leak_uri-leak_genka:+,.0f}")
    print(f"   └ B.キャンセル補償(雑収入・status正): {len(hosho)}件 / {sum(v for _,v in hosho):,.0f}W ≈ ¥{sum(v for _,v in hosho)*0.105:,.0f}")

    for k in only_m:
        m = mgmt_valid[k]
        rows.append(dict(注文番号=k, 区分="精算待ち(実精算なし)", アカウント=m["account"], 年月=m["ym"],
                         商品名=m["name"], 見込みZ_KRW=round(m["Zkrw"]), 実精算_KRW="", 差KRW="",
                         精算源="", 行別為替=m["fx"], 実精算_円="", 見込みAB_円=round(m["AB"])))
    for k, v in cancelled_paid:
        rows.append(dict(注文番号=k, 区分=classify(v), アカウント=mgmt.get(k,{}).get("account",""),
                         年月=mgmt.get(k,{}).get("ym",""), 商品名=mgmt.get(k,{}).get("name",""),
                         見込みZ_KRW="", 実精算_KRW=round(v), 差KRW="", 精算源=settle[k]["src"],
                         行別為替="", 実精算_円="", 見込みAB_円=""))

    with open(OUT, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.DictWriter(f, fieldnames=["注文番号","区分","アカウント","年月","商品名","見込みZ_KRW","実精算_KRW","差KRW","精算源","行別為替","実精算_円","見込みAB_円"])
        w.writeheader(); w.writerows(rows)
    print(f"\n照合明細CSV: {OUT}  ({len(rows)}行)")

    # ===== 月次運用ランブック(毎回表示=作り込み) =====
    print(f"""
┌─ 月次運用ランブック（クーパンが精算した直後に回す）─────────────
│ ②精算待ち{len(only_m)}件の対策＝時間が解決。クーパンの精算は決済の2〜4ヶ月後。
│   毎月、精算予定日を過ぎたら Wing→決済→精算状況→決済確定 を【両口座】2ヶ月窓でDL
│   → 260530_資金繰り/ 等にMSF_PAYMENT_REVENUE_DETAIL*.xlsx を追加 → 本スクリプト再実行。
│   今回の12月未収はクーパン2の精算予定日2026-04-21/05-26を追加DLでほぼ100%に到達。
│ ①売上漏れ{len(leak)}件の対策＝毎回このAフラグを管理表へ反映。
│   A=実精算が商品売上規模なのにstatusがキャンセル/発送登録不可 → status'発送済'へ修正し売上計上。
│   B=精算≈29,100Wの定額キャンセル補償 → 売上でなく雑収入。statusは'キャンセル'のまま正。
│   根治策：管理表のstatusは「精算が出てから最終確定」する。出る前の早合点除外をやめる。
└──────────────────────────────────────────────""")

if __name__ == "__main__":
    main()

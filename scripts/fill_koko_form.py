# -*- coding: utf-8 -*-
"""
公庫公式様式 koko.xlsx の実セルに直接記入して提出用xlsxを生成（オーロラFC船橋）。
埋まる部分は記入、本部依存は［本部確認］、個人未確定は［要記入］。
"""
import openpyxl
from openpyxl.styles import Alignment, Font

SRC = "/tmp/koko.xlsx"
OUT = "/tmp/創業計画書_オーロラFC船橋_記入版_260531.xlsx"

DOUKI = ("当社（KHD）は不動産物件調達および医療クリニックの承継コンサルティングを営み、医療機関・ケアマネジャーとのネットワークと事業運営基盤を有している。"
         "事業エリアの船橋市は後期高齢者が8万人を超え、在宅・施設での訪問マッサージ（医療保険適用のマッサージ・はり灸）需要が構造的に拡大している。"
         "既存事業で培った医療・介護分野の人的ネットワークを活かせる隣接領域であり、訪問医療マッサージFC「オーロラ」に加盟することで、"
         "集客・保険請求（レセプト）代行・運営ノウハウを本部から取得し、未経験リスクを抑えて確実に立ち上げられると判断し創業する。")
JIYU = ("・船橋市の後期高齢者8万人超という需要基盤と、本部territory承認済みの独占エリアで開業する。\n"
        "・既存のKHD医療・介護ネットワーク（ケアマネ・医療機関）を集客に活用できる。\n"
        "・加盟前にIndeedで採用テストを実施し、即戦力人材（53歳・あマ指師＋鍼灸師・訪問実務／ケアマネ営業／管理者経験）の応募を確認済み＝人員確保の蓋然性を実証。\n"
        "・本部のレセプト代行により保険請求の未収リスクを抑制。慎重に採算を検証したうえで開業判断を行う。")
KONKYO = ("売上高＝施術者数 × 1日訪問件数 × 訪問単価 × 稼働日数（訪問1回の保険診療単価・1人あたり標準モデルは本部受領後に確定）。"
          "売上原価は施術用消耗品で僅少。経費＝施術者人件費／ロイヤリティ13.2%（税込）／家賃・車両・通信・広告分担金／公庫返済利息。")

wb = openpyxl.load_workbook(SRC)
ws = wb["創業計画書"]

# ── 複数行テキスト欄を結合し直す（列はそのまま・行方向に拡張）──
def remerge(unmerge_list, target):
    for rng in unmerge_list:
        try: ws.unmerge_cells(rng)
        except Exception: pass
    ws.merge_cells(target)

remerge(["B7:AM7", "B8:AM8", "B9:AM9", "B10:AM10"], "B7:AM10")
remerge(["H26:AM26", "H27:AM27"], "H26:AM27")
remerge(["H35:AM35", "H36:AM36", "H37:AM37"], "H35:AM37")
remerge(["H38:AM38", "H39:AM39", "H40:AM40"], "H38:AM40")
remerge(["H41:AM41", "H42:AM42", "H43:AM43"], "H41:AM43")

WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")

def put(coord, val, wrap=False, size=9):
    c = ws[coord]
    c.value = val
    if wrap:
        c.alignment = WRAP
    f = c.font
    c.font = Font(name=f.name or "ＭＳ Ｐゴシック", size=size, bold=f.bold)

# ── 記入 ──
put("V3", "菊池 研太", size=11)
# 1 動機
put("B7", DOUKI, wrap=True, size=9)
# 2 略歴
put("B13", "2023年〜現在", size=9)
put("H13", "KHD（kikuchi-hd.net）代表。不動産物件調達（仕入〜加工〜売却）・医療クリニック承継コンサル・EC（韓国輸出）を運営。複数事業の収支・資金繰りを管理。", wrap=True, size=9)
put("H14", "［以前の勤務先・担当業務・役職・身につけた技能を記入］", wrap=True, size=9)
put("AG20", "不動産・医療コンサル・EC", size=9)   # 過去の事業経験 事業内容
put("P23", "あマ指師の施術所開設届・出張施術業務開始届を予定（施術者側で要確認）", wrap=True, size=9)  # 許認可
# 3 取扱商品
put("H26", "訪問マッサージ／訪問はり灸（医療保険適用、医師の同意書に基づく施術）。訪問医療マッサージFC「オーロラ」加盟。提供エリア：船橋市（本部territory承認済み）。", wrap=True, size=9)
put("I28", "① 訪問マッサージ", size=9)
put("I29", "② 訪問はり灸", size=9)
put("I30", "③ ―", size=9)
put("H31", "［本部確認：訪問1回の保険診療単価］", size=9)  # 客単価
put("H32", "［ ］", size=9)  # 営業日数
put("H35", "本部ブランド＋レセプト代行で未収・査定リスクが低い。ケアマネ・施設営業網による安定集客。KHDの医療・介護ネットワークを活用。", wrap=True, size=9)
put("H38", "船橋市の在宅高齢者・施設入居者。ケアマネ・医療機関への営業、本部の集客支援、紹介ルートで獲得。", wrap=True, size=9)
put("H41", "船橋市は後期高齢者8万人超で在宅医療マッサージ需要が拡大。本部territory承認により当該エリアで独占的に展開。", wrap=True, size=9)
# 4 従業員
put("I45", "1", size=10)
# 5 取引先
put("C51", "後期高齢者医療広域連合 等（保険者）／入金まで約［本部確認］ヶ月", wrap=True, size=9)
put("C57", "施術用消耗品ほか ［ ］", size=9)
put("C63", "オーロラ本部（ロイヤリティ13.2%税込・毎月）", wrap=True, size=9)
# 6 関連企業
put("AU7", "KHD", size=9)
put("AU8", "菊池 研太", size=9)
put("AU9", "［ ］", size=9)
put("AU10", "不動産・医療コンサル・EC", size=9)
# 7 お借入
put("AO13", "［要記入：住宅・車・教育・カード等。無ければ「該当なし」］", wrap=True, size=9)
# 8 必要な資金と調達方法
put("AP20", "加盟金・研修費・保証金・備品・機材・初期システム", wrap=True, size=9)
put("AY20", "オーロラ本部", size=9)
put("BD20", "［本部確認］", size=9)
put("AP32", "立上げ人件費・採用費・広告分担金・家賃・車両・予備費", wrap=True, size=9)
put("BD32", "［算定］", size=9)
put("BV18", 600, size=10)                 # 自己資金
put("BV24", "［所要−600万］", size=9)      # 公庫からの借入
# 9 事業の見通し（数値は本部待ち、根拠のみ記入）
put("BF41", KONKYO, wrap=True, size=9)
# 10 自由記述
put("AO60", JIYU, wrap=True, size=9)

# 行高を確保（複数行欄）
for r, h in {7: 78, 26: 46, 35: 46, 38: 46, 41: 46}.items():
    ws.row_dimensions[r].height = h

wb.save(OUT)
print("SAVED:", OUT)

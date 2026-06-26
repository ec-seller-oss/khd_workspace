# -*- coding: utf-8 -*-
"""
📦 HANIRO CSV自動生成スクリプト (01_経営管理 CFO)
─────────────────────────────────────────────────
HANIROへの代行業者一括CSV を自動生成する。

【入力】
  ① 販売管理表 (25*_販売管理表*.xlsx) の「Amazon発注済・HANIRO未登録」行
  ② ec_automation.db の orders テーブル (自動化後)

【出力】
  HANIRO_batch_YYYYMMDD_HHMMSS.csv → Drive「_HANIRO登録CSV/」フォルダへ保存

【HANIROに送る項目】（実際のHANIROフォームに合わせて微調整可能）
  受取人名・郵便番号・住所・電話番号・個人通関番号・
  商品名・重量・Amazon追跡番号（出荷後）

【使い方】
  # 販売管理表から未登録行を抽出してCSV生成
  python3 scripts/haniro_csv_generator.py --from-mgmt

  # DBから生成（自動化後）
  python3 scripts/haniro_csv_generator.py --from-db

  # テスト出力（標準出力のみ）
  python3 scripts/haniro_csv_generator.py --from-mgmt --dry-run
"""

import os, csv, glob, re, argparse
from datetime import datetime, date
from pathlib import Path
import openpyxl
import sys

sys.path.insert(0, os.path.dirname(__file__))

DIR = "/Users/kikuchikenta/Library/CloudStorage/GoogleDrive-ec-seller@kikuchi-hd.net/共有ドライブ/01_個人/2025_帳票、明細/韓国輸出売上"
HANIRO_DIR = Path(DIR) / "_HANIRO登録CSV"
DB_PATH = os.path.expanduser("~/01_honbu_docs_automation/ec_automation.db")

# 販売管理表の列インデックス（0始まり）
COL = dict(
    status=0,        # A: 進捗
    order_no=1,      # B: 注文番号
    platform=2,      # C: プラットフォーム
    carrier=3,       # D: 配送業者
    customer_name=5, # F: 注文者名
    product_kr=6,    # G: 商品名（韓国語）
    product_jp=7,    # H: 日本語商品名
    asin=8,          # I: ASIN
    amazon_url=9,    # J: Amazon URL
    cost=12,         # M: 仕入
    amazon_account=13, # N: 仕入先
    amazon_order=14, # O: 仕入れ先注文番号
    total_cost=19,   # T: 総原価
    sale_date=20,    # U: Sale Date
    ship_deadline=21,# V: 発送予定日
    quantity=22,     # W: Quantity
    sale_price=23,   # X: Sale Price
    weight=28,       # ] : 商品重量
    postal_code=41,  # j: 郵便番号
    address=42,      # k: 住所
    phone=43,        # l: 携帯電話番号
    customs_no=44,   # m: 個人通関番号
    delivery_msg=45, # n: 配送メッセージ
    customs1=46,     # o: 通関確認①
    customs2=47,     # p: 通関確認②
    customs3=48,     # q: 通関確認③
)

# HANIROが要求するCSVカラム（実際のフォーマットに合わせて変更）
HANIRO_COLS = [
    "受取人名",
    "郵便番号",
    "住所",
    "電話番号",
    "個人通関番号",
    "商品名（日本語）",
    "商品名（韓国語）",
    "数量",
    "重量(kg)",
    "Amazon注文番号",
    "ASIN",
    "送付メッセージ",
    "注文日",
    "発送期限",
    "Coupang注文番号",
    "通関確認①",
    "通関確認②",
    "通関確認③",
]

VALID_STATUSES = {"代行登録済み", "注文済", "発送済"}
PURCHASE_STATUSES = {"注文済", "代行登録済み"}  # Amazon発注済みの行


def norm(v):
    if v is None: return ""
    return str(v).strip()


def load_from_mgmt(only_unregistered=True):
    """
    販売管理表から代行未登録・Amazon発注済みの行を抽出
    only_unregistered=True: HANIRO未登録（配送業者D列が空か"注文済"）のみ
    """
    rows = []
    for path in sorted(glob.glob(f"{DIR}/25*_販売管理表*.xlsx")):
        fn = os.path.basename(path)
        m = re.match(r"(\d{2})(\d{2})_販売管理表([12])", fn)
        if not m: continue
        acct = f"クーパン{m.group(3)}"
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        for row in ws.iter_rows(min_row=7, values_only=True):
            if len(row) <= COL["asin"]: continue
            status = norm(row[COL["status"]])
            if status not in PURCHASE_STATUSES:
                continue
            order_no = norm(row[COL["order_no"]])
            if not order_no or not order_no.isdigit():
                continue
            # Amazon注文番号があること（発注済み確認）
            amazon_order = norm(row[COL["amazon_order"]])
            if only_unregistered and not amazon_order:
                continue
            carrier = norm(row[COL["carrier"]])
            # HANIRO登録済みかどうか: D列に追跡番号っぽい文字列があれば登録済み
            if only_unregistered and carrier and carrier not in ("", "注文済", "キャンセル"):
                continue  # すでに配送業者が登録されている→登録済み

            rows.append({
                "coupang_order_no": order_no,
                "account": acct,
                "customer_name": norm(row[COL["customer_name"]]),
                "postal_code": norm(row[COL["postal_code"]]),
                "address": norm(row[COL["address"]]),
                "phone": norm(row[COL["phone"]]),
                "customs_no": norm(row[COL["customs_no"]]),
                "product_jp": norm(row[COL["product_jp"]]),
                "product_kr": norm(row[COL["product_kr"]]),
                "quantity": norm(row[COL["quantity"]]) or "1",
                "weight": norm(row[COL["weight"]]) or "",
                "amazon_order": amazon_order,
                "asin": norm(row[COL["asin"]]),
                "delivery_msg": norm(row[COL["delivery_msg"]]),
                "sale_date": norm(row[COL["sale_date"]])[:10] if row[COL["sale_date"]] else "",
                "ship_deadline": norm(row[COL["ship_deadline"]])[:10] if row[COL["ship_deadline"]] else "",
                "customs1": norm(row[COL.get("customs1", 46)]) if len(row) > 46 else "",
                "customs2": norm(row[COL.get("customs2", 47)]) if len(row) > 47 else "",
                "customs3": norm(row[COL.get("customs3", 48)]) if len(row) > 48 else "",
            })
        wb.close()
    return rows


def load_from_db():
    """DBのorders (status=purchased, haniro_registered=0) から生成"""
    try:
        from ec_automation_db import get_unregistered_haniro_orders
        return get_unregistered_haniro_orders()
    except Exception as e:
        print(f"[haniro] DB読み込みエラー: {e} → 管理表から代替")
        return load_from_mgmt()


def to_haniro_row(r: dict) -> list:
    return [
        r.get("customer_name", ""),
        r.get("postal_code", ""),
        r.get("address", ""),
        r.get("phone", ""),
        r.get("customs_no", ""),
        r.get("product_jp", "")[:100],
        r.get("product_kr", "")[:100],
        r.get("quantity", "1"),
        r.get("weight", ""),
        r.get("amazon_order", r.get("amazon_order_no", "")),
        r.get("asin", ""),
        r.get("delivery_msg", ""),
        r.get("sale_date", ""),
        r.get("ship_deadline", ""),
        r.get("coupang_order_no", ""),
        r.get("customs1", ""),
        r.get("customs2", ""),
        r.get("customs3", ""),
    ]


def generate_csv(rows: list, dry_run=False):
    if not rows:
        print("[haniro] 対象注文なし（全件登録済み or Amazon未発注）")
        return None

    batch_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"HANIRO_batch_{batch_id}.csv"

    if dry_run:
        print(f"[DRY RUN] {len(rows)}件 → {filename}")
        import io
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(HANIRO_COLS)
        for r in rows:
            w.writerow(to_haniro_row(r))
        print(buf.getvalue())
        return None

    HANIRO_DIR.mkdir(parents=True, exist_ok=True)
    out_path = HANIRO_DIR / filename
    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        w.writerow(HANIRO_COLS)
        for r in rows:
            w.writerow(to_haniro_row(r))

    print(f"✅ HANIRO CSV生成: {out_path}")
    print(f"   {len(rows)}件")

    # DB側にバッチ登録
    try:
        import sqlite3
        conn = sqlite3.connect(DB_PATH)
        conn.execute(
            "INSERT INTO haniro_batch_log (batch_id,csv_path,order_count) VALUES (?,?,?)",
            [batch_id, str(out_path), len(rows)]
        )
        for r in rows:
            on = r.get("coupang_order_no", "")
            if on:
                conn.execute(
                    "UPDATE orders SET haniro_registered=1, haniro_batch_id=?, updated_at=datetime('now','localtime') "
                    "WHERE coupang_order_no=?",
                    [batch_id, on]
                )
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[haniro] DBログ記録スキップ: {e}")

    return str(out_path)


def main():
    parser = argparse.ArgumentParser(description="HANIRO CSV自動生成")
    parser.add_argument("--from-mgmt", action="store_true", help="販売管理表から生成（デフォルト）")
    parser.add_argument("--from-db",   action="store_true", help="自動化DBから生成")
    parser.add_argument("--dry-run",   action="store_true", help="CSVを保存せず標準出力に表示")
    parser.add_argument("--all",       action="store_true", help="登録済みも含めて全件出力")
    args = parser.parse_args()

    if args.from_db:
        rows = load_from_db()
    else:
        rows = load_from_mgmt(only_unregistered=not args.all)

    print(f"対象: {len(rows)}件")
    generate_csv(rows, dry_run=args.dry_run)


if __name__ == "__main__":
    main()

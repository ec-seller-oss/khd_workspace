#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
仕入一次スクリーナー（融資資料自動化TTP・自社版 スコープA）
================================================================
マイソク/物件概要の数値を入れる → 玉川式KPIで「買い / 要検討 / 見送り」を秒判定。
PDF抽出・スクレイピングは後フェーズ。まずは判定エンジンの心臓部。

【設計原則（B&S設計書 原則②を踏襲）】
  取れない値は勝手に補完しない。missing は ❌ として正直に表示する。
  最終判断は人。本ツールは判断材料の収集・整理に徹する。

使い方:
  python3 property_screener.py                 # 同梱サンプルで実行
  python3 property_screener.py 物件.json        # JSON入力で実行
  → コンソールに判定 + Excel判定シートを固定フォルダに出力
"""
import re
import sys
import json
import datetime
from pathlib import Path

# ── 玉川式 判定基準（~/.claude/CLAUDE.md より）─────────────────
KIJUN = {
    "土地値割合": {"safe": 0.40, "warn": 0.30},   # 0.4以上=安全圏
    "粗利率":     {"safe": 0.20, "warn": 0.12},   # 20%以上=キャピタルOK
    "積算割合":   {"safe": 0.70, "warn": 0.50},   # 0.7↑融資見込み/0.5未満=融資出にくく弾く（築古は低め）
}
# インカム（保有収益）判定基準（玉川式）
KIJUN_INCOME = {
    "実質利回り": {"safe": 0.08, "warn": 0.05},   # 経費15%控除後
    "CF率":      {"safe": 0.015, "warn": 0.010},  # 税引前CF/価格 1.5〜2%
    "CCR":       {"safe": 0.15, "warn": 0.10},    # 税引前CF/自己資金 15〜20%
}
# 簡易ボリューム用の1戸あたり専有面積（仮値・後で実測で上書き）
KOSU_MODELS = {"1DK": 25.0, "1LDK": 40.0}
# 専有面積→延床への割戻し係数（共用部込み・仮 0.75）
SENYU_RATIO = 0.75
KEIHI_RATIO = 0.15  # 運営経費率（実質利回り算出・仮）

# 構造別 法定耐用年数 と 再調達単価(円/㎡・銀行積算の目安)
TAIYO = {"SRC": 47, "RC": 47, "重量鉄骨": 34, "鉄骨": 27, "軽量鉄骨": 19, "木造": 22}
SAICHOTATSU = {"SRC": 200000, "RC": 200000, "重量鉄骨": 180000, "鉄骨": 170000,
               "軽量鉄骨": 140000, "木造": 150000}


def detect_struct(s):
    """構造文字列 → 種別キー。"""
    if not s:
        return None
    for k in ("SRC", "RC", "重量鉄骨", "軽量鉄骨", "鉄骨", "木造"):
        if k in s:
            return k
    return None


def chikunen_from(s):
    """『2017年11月（築8年）』『築58年』等 → 築年数(int)。"""
    if not s:
        return None
    m = re.search(r"築\s*(\d+)\s*年", s)
    if m:
        return int(m.group(1))
    m = re.search(r"(\d{4})\s*年", s)
    if m:
        return max(0, 2026 - int(m.group(1)))
    return None


def pmt_annual(principal, rate_pct, years):
    """元利均等の年間返済額。principal/years 不足なら None。"""
    principal = f(principal); years = f(years); rate_pct = f(rate_pct)
    if not principal or not years:
        return None
    r = (rate_pct or 0) / 100 / 12
    n = int(years * 12)
    m = principal / n if r == 0 else principal * r / (1 - (1 + r) ** -n)
    return m * 12

OUT_DIR = Path.home() / "01_honbu_docs_automation" / "out_screener"


def f(v):
    """数値 or None を安全に float 化。空文字/None→None"""
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", ""))
    except ValueError:
        return None


def calc(p):
    """物件dict → 算出結果dict（取れない項目は None のまま残す）"""
    price = f(p.get("価格_円"))
    land = f(p.get("土地面積_m2"))
    rosenka = f(p.get("路線価_円per_m2"))   # 相続税路線価（円/㎡）
    yoseki = f(p.get("容積率_pct"))          # 例: 200
    uridashi = f(p.get("想定売却額_円"))
    kenchikuhi = f(p.get("建築費_円"))

    r = {"missing": []}

    # ① 土地値割合 = 路線価 × 面積 ÷ 価格
    if rosenka and land and price:
        r["土地値_円"] = rosenka * land
        r["土地値割合"] = r["土地値_円"] / price
    else:
        r["土地値割合"] = None
        for k, v in (("路線価_円per_m2", rosenka), ("土地面積_m2", land), ("価格_円", price)):
            if not v:
                r["missing"].append(k)

    # ② 簡易ボリューム（延床 = 面積 × 容積率、戸数 = 延床×専有率 ÷ 1戸面積）
    if land and yoseki:
        enyuka = land * (yoseki / 100.0)
        r["延床_m2"] = enyuka
        r["戸数"] = {m: int((enyuka * SENYU_RATIO) // a) for m, a in KOSU_MODELS.items()}
    else:
        r["戸数"] = None
        if not yoseki:
            r["missing"].append("容積率_pct")

    # ④ インカム（保有収益）：年賃料があれば収益物件として判定
    nenchin = f(p.get("年賃料_円"))
    r["is_income"] = bool(nenchin)
    if nenchin and price:
        r["表面利回り"] = nenchin / price
        r["実質利回り"] = nenchin * (1 - KEIHI_RATIO) / price
        noi = nenchin * (1 - KEIHI_RATIO)
        ann_pay = pmt_annual(p.get("融資額_円"), p.get("金利_pct"), p.get("期間_年"))
        if ann_pay is not None:
            r["年返済_円"] = ann_pay
            r["税引前CF"] = noi - ann_pay
            r["CF率"] = r["税引前CF"] / price
            jiko = f(p.get("自己資金_円"))
            if jiko:
                r["CCR"] = r["税引前CF"] / jiko

    # ③ 粗利率 =（売却額 −（仕入 + 建築費））÷ 売却額（キャピタル時のみ）
    if uridashi and price is not None and kenchikuhi is not None:
        genka = price + kenchikuhi
        r["粗利_円"] = uridashi - genka
        r["粗利率"] = r["粗利_円"] / uridashi
    else:
        r["粗利率"] = None
        if not r["is_income"]:   # 収益物件では売却額/建築費の欠損は不問
            for k, v in (("想定売却額_円", uridashi), ("建築費_円", kenchikuhi)):
                if v is None:
                    r["missing"].append(k)

    # ⑤ 積算評価（銀行目線・融資が出るか）＝ 土地積算 ＋ 建物積算
    ref = p.get("_参考情報") or {}
    bldg_area = f(ref.get("建物面積_m2"))
    struct = detect_struct(ref.get("構造"))
    chikunen = chikunen_from(ref.get("築年月"))
    if rosenka and land:
        r["土地積算_円"] = rosenka * land
        tatemono = 0
        if bldg_area and struct:
            taiyo = TAIYO[struct]
            zonzon = max(0, taiyo - (chikunen or 0))
            tatemono = SAICHOTATSU[struct] * bldg_area * (zonzon / taiyo)
            r["建物積算_円"] = tatemono
        r["積算評価_円"] = r["土地積算_円"] + tatemono
        if price:
            r["積算割合"] = r["積算評価_円"] / price   # 1以上=積算が出る=融資◎
    # B（旧バイセル式・融資カバー率）＝仕入×0.8÷(面積×路線価)。1以下で担保堅い（参考）
    if rosenka and land and price:
        r["融資カバー率B"] = price * 0.8 / (land * rosenka)

    return r


def judge_one(name, value, kijun):
    """単一KPIの信号（🟢/🟡/🔴/❌）を返す"""
    if value is None:
        return "❌", "データ未取得"
    if value >= kijun["safe"]:
        return "🟢", "安全圏"
    if value >= kijun["warn"]:
        return "🟡", "要検討"
    return "🔴", "基準未達"


def overall(signals):
    """総合判定。🔴が1つでも→見送り、❌や🟡が残る→要検討、全🟢→買い"""
    s = [x[0] for x in signals]
    if "🔴" in s:
        return "見送り", "🔴"
    if "❌" in s or "🟡" in s:
        return "要検討（データ補完 or 条件交渉）", "🟡"
    return "買い", "🟢"


def render_console(p, r):
    lines = []
    lines.append("=" * 56)
    lines.append(f"  仕入一次スクリーナー判定 — {p.get('物件名','(無名)')}")
    lines.append("=" * 56)
    lines.append(f"  価格    : {fmt_yen(p.get('価格_円'))}")
    lines.append(f"  土地面積: {p.get('土地面積_m2','—')} ㎡")
    lines.append("-" * 56)

    signals = []
    # 土地値割合（担保・両モード共通）
    sig, note = judge_one("土地値割合", r.get("土地値割合"), KIJUN["土地値割合"])
    signals.append((sig, note))
    lines.append(f"  {sig} 土地値割合 : {pct(r.get('土地値割合'))}  (基準 0.40↑)  {note}")

    if r.get("is_income"):
        # ── インカム（保有収益）モード ──
        lines.append(f"  🏠 モード: 収益物件（インカム評価）")
        lines.append(f"  ・表面利回り: {pct(r.get('表面利回り'))} （参考）")
        sig, note = judge_one("実質利回り", r.get("実質利回り"), KIJUN_INCOME["実質利回り"])
        signals.append((sig, note))
        lines.append(f"  {sig} 実質利回り : {pct(r.get('実質利回り'))}  (基準 8%↑/経費15%控除)  {note}")
        if r.get("CF率") is not None:
            sig, note = judge_one("CF率", r.get("CF率"), KIJUN_INCOME["CF率"])
            signals.append((sig, note))
            lines.append(f"  {sig} CF率       : {pct(r.get('CF率'))}  (基準 1.5%↑)  税引前CF {fmt_yen(r.get('税引前CF'))}/年")
        else:
            lines.append(f"  ❔ CF率       : 融資条件未入力（--loan/--rate/--years）")
        if r.get("CCR") is not None:
            sig, note = judge_one("CCR", r.get("CCR"), KIJUN_INCOME["CCR"])
            signals.append((sig, note))
            lines.append(f"  {sig} CCR        : {pct(r.get('CCR'))}  (基準 15%↑)  {note}")
    else:
        # ── キャピタル（再販）モード ──
        sig, note = judge_one("粗利率", r.get("粗利率"), KIJUN["粗利率"])
        signals.append((sig, note))
        lines.append(f"  {sig} 粗利率     : {pct(r.get('粗利率'))}  (基準 0.20↑)  {note}")
    # ボリューム（参考・判定対象外）
    if r.get("戸数"):
        ko = r["戸数"]
        lines.append(f"  📐 簡易ボリューム: 延床{r['延床_m2']:.0f}㎡ → "
                     f"1DK {ko['1DK']}戸 / 1LDK {ko['1LDK']}戸 (参考)")
    else:
        lines.append(f"  ❌ 簡易ボリューム: 容積率データ未取得")

    # 積算割合（銀行目線・融資が出るか）＝参考表示のみ（買い/見送りは玉川で決める）
    if r.get("積算割合") is not None:
        sig, _ = judge_one("積算割合", r.get("積算割合"), KIJUN["積算割合"])
        b = r.get("融資カバー率B")
        bstr = f"／B={b:.2f}" if b is not None else ""
        lines.append(f"  🏦 積算割合 : {sig}{pct(r['積算割合'])}  (融資目安0.7↑・参考)  積算{fmt_yen(r.get('積算評価_円'))}{bstr}")
    elif r.get("融資カバー率B") is not None:
        lines.append(f"  🏦 融資カバー率B: {r['融資カバー率B']:.2f}（1以下で担保堅い・参考）")

    lines.append("-" * 56)
    verdict, mark = overall(signals)
    lines.append(f"  {mark} 総合判定 : 【 {verdict} 】")
    if r["missing"]:
        lines.append(f"  ⚠️ 未取得（要補完）: {', '.join(sorted(set(r['missing'])))}")
    lines.append("=" * 56)
    return "\n".join(lines), verdict, mark, signals


def fmt_yen(v):
    v = f(v)
    return f"{v:,.0f} 円" if v is not None else "—"


def pct(v):
    return f"{v*100:.1f}%" if v is not None else "—（未取得）"


def write_excel(p, r, verdict, mark, signals):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return None
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "仕入判定"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 16

    brick = "AA2E26"
    cream = "F9F6EF"
    bold_w = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor=brick)
    thin = Border(*[Side(style="thin", color="DDDDDD")] * 4)

    ws["A1"] = f"仕入一次スクリーナー判定 — {p.get('物件名','(無名)')}"
    ws["A1"].font = Font(bold=True, size=14, color=brick)
    ws.merge_cells("A1:C1")
    ws["A2"] = f"作成: {datetime.date.today().isoformat()}  /  玉川式KPI"
    ws["A2"].font = Font(size=9, color="888888")

    row = 4
    ws.cell(row, 1, "総合判定").font = Font(bold=True)
    c = ws.cell(row, 2, f"{mark} {verdict}")
    c.font = Font(bold=True, size=13,
                  color="C00000" if mark == "🔴" else ("BF8F00" if mark == "🟡" else "1F7A1F"))
    row += 2

    for h in ("項目", "値", "判定"):
        cc = ws.cell(row, ("項目", "値", "判定").index(h) + 1, h)
        cc.font = bold_w
        cc.fill = head_fill
        cc.alignment = Alignment(horizontal="center")
    row += 1

    rows = [
        ("価格", fmt_yen(p.get("価格_円")), ""),
        ("土地面積", f"{p.get('土地面積_m2','—')} ㎡", ""),
        ("土地値割合", pct(r.get("土地値割合")), f"{signals[0][0]} {signals[0][1]}"),
    ]
    if r.get("is_income"):
        rows.append(("モード", "収益物件（インカム）", "🏠"))
        rows.append(("表面利回り", pct(r.get("表面利回り")), "参考"))
        rows.append(("実質利回り", pct(r.get("実質利回り")), f"{signals[1][0]} {signals[1][1]}"))
        if r.get("CF率") is not None:
            rows.append(("CF率", pct(r.get("CF率")), f"税引前CF {fmt_yen(r.get('税引前CF'))}/年"))
        if r.get("CCR") is not None:
            rows.append(("CCR", pct(r.get("CCR")), ""))
    else:
        rows.append(("粗利率", pct(r.get("粗利率")), f"{signals[1][0]} {signals[1][1]}"))
    if r.get("戸数"):
        rows.append(("簡易ボリューム",
                     f"延床{r['延床_m2']:.0f}㎡ / 1DK {r['戸数']['1DK']}戸・1LDK {r['戸数']['1LDK']}戸",
                     "📐参考"))
    if r["missing"]:
        rows.append(("未取得（要補完）", ", ".join(sorted(set(r["missing"]))), "⚠️"))

    for label, val, jud in rows:
        ws.cell(row, 1, label).border = thin
        ws.cell(row, 2, str(val)).border = thin
        ws.cell(row, 3, jud).border = thin
        row += 1

    safe_name = str(p.get("物件名", "property")).replace("/", "_").replace(" ", "_")
    out = OUT_DIR / f"判定_{safe_name}_{datetime.date.today().isoformat()}.xlsx"
    wb.save(out)
    return out


SAMPLE = {
    "物件名": "サンプル_船橋テスト地",
    "価格_円": 30000000,
    "土地面積_m2": 165.0,
    "路線価_円per_m2": 95000,
    "容積率_pct": 200,
    "想定売却額_円": 52000000,
    "建築費_円": 16000000,
}


def main():
    if len(sys.argv) > 1:
        p = json.loads(Path(sys.argv[1]).read_text(encoding="utf-8"))
    else:
        p = SAMPLE
        print("（入力ファイル未指定 → 同梱サンプルで実行）\n")
    r = calc(p)
    console, verdict, mark, signals = render_console(p, r)
    print(console)
    out = write_excel(p, r, verdict, mark, signals)
    if out:
        print(f"\n📄 判定シート出力: {out}")
    else:
        print("\n(openpyxl 未導入のためExcel出力スキップ)")


if __name__ == "__main__":
    main()

# -*- coding: utf-8 -*-
"""
京橋クリニック｜AI/DX 提案の幅 × 課題と対策 網羅マトリクス（+iPad段取り+補助金）
04コンサル(李牧) / 2026-06-17 / 配色: クリーム白#F9F6EF × レンガ赤#AA2E26
Web裏取り済み（出典URL付）。未確認は「要確認」。
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CREAM="F9F6EF"; RED="AA2E26"; INK="1A1A1A"; RULE="DAD6CF"; BEIGE="F1ECE1"
WHITE="FFFFFF"; GREEN="2E7D32"; GREY="6B6B6B"; AMBER="B5821E"; INPUT_BG="FFF2CC"
F="Hiragino Sans"

def fn(sz=10,b=False,c=INK): return Font(name=F,size=sz,bold=b,color=c)
def fl(c): return PatternFill("solid",fgColor=c)
def sd(c=RULE): return Side(style="thin",color=c)
BORDER=Border(left=sd(),right=sd(),top=sd(),bottom=sd())
def al(h="left",v="top",w=True): return Alignment(horizontal=h,vertical=v,wrap_text=w)

wb=openpyxl.Workbook()

def newsheet(t,tab=RED):
    ws=wb.create_sheet(t); ws.sheet_properties.tabColor=tab; ws.sheet_view.showGridLines=False
    return ws

def title_block(ws,kick,title,sub,span):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=span)
    c=ws.cell(1,1,kick); c.font=fn(9,True,RED); c.alignment=al()
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=span)
    c=ws.cell(2,1,title); c.font=fn(16,True,INK); c.alignment=al()
    ws.merge_cells(start_row=3,start_column=1,end_row=3,end_column=span)
    c=ws.cell(3,1,sub); c.font=fn(10,False,GREY); c.alignment=al()
    for col in range(1,span+1): ws.cell(4,col).fill=fl(RED)
    ws.row_dimensions[4].height=3
    return 5

def header(ws,row,heads,widths):
    for i,h in enumerate(heads,1):
        c=ws.cell(row,i,h); c.fill=fl(RED); c.font=fn(9.5,True,WHITE); c.alignment=al("center","center"); c.border=BORDER
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.row_dimensions[row].height=26

def row(ws,r,vals,h=46,bg=WHITE,inks=None,bolds=None):
    for i,v in enumerate(vals,1):
        c=ws.cell(r,i,v); c.border=BORDER; c.fill=fl(bg)
        col=INK; b=False
        if inks and i in inks: col=inks[i]
        if bolds and i in bolds: b=True
        if i==1: b=True
        c.font=fn(9.5,b,col); c.alignment=al("left","top")
    ws.row_dimensions[r].height=h

# ============== 00 表紙 ==============
ws=newsheet("00_表紙")
r=title_block(ws,"KYOBASHI CLINIC ｜ AI/DX PROPOSAL MATRIX",
    "京橋クリニック AI/DX 提案マトリクス（課題 × 対策の網羅）",
    "2026-06-17 面談を受けた follow-up ／ Web裏取り済（出典URL付・未確認は「要確認」）",6)
for w,c in zip([22,30,22,18,14,14],range(1,7)):
    ws.column_dimensions[get_column_letter(c)].width=w
r+=1
for k,v in [("対象","医療法人社団順医会 京橋クリニック ／ 山崎先生"),
            ("フロー","【菊池・宮崎】作業→【福井】確認→【先生】提出"),
            ("最速提出","Task1・2は 6/22(月) 目標／Stream Deckは月曜 現地デモ"),
            ("本部","04コンサル（李牧） ／ テナントアシスト・ウイン")]:
    ws.cell(r,1,k).font=fn(10,True,RED)
    ws.merge_cells(start_row=r,start_column=2,end_row=r,end_column=6)
    ws.cell(r,2,v).font=fn(10); ws.cell(r,2).alignment=al()
    r+=1
r+=1
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
ws.cell(r,1,"■ タブ：01_課題×対策マトリクス（網羅）／02_iPad導入の段取り（2トラック）／03_IT補助金まとめ").font=fn(11,True,RED)
ws.cell(r,1).alignment=al(); r+=1
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
ws.cell(r,1,"凡例：確度 ◎=導入しやすい/事例あり ○=現実的/要見積 △=要件・規制注意 ／ 黄色セル=確認待ち").font=fn(9.5,False,GREY)
ws.cell(r,1).alignment=al()

# ============== 01 課題×対策マトリクス ==============
ws=newsheet("01_課題×対策マトリクス")
r=title_block(ws,"ISSUE × SOLUTION MATRIX",
    "提案の幅 ── 課題ごとの対策・電カル連携・補助金・確度",
    "電子カルテ=中央ビジコン Medicom 前提。Medicom Cloud Connect API（2022公開）で外部取込の実例あり。",8)
heads=["領域","現場の課題（会議メモ/実データ）","対策（実在ツール・手段）","Medicom電カルとの兼ね合い","IT補助金","確度","参考事例・出典","次アクション・確認先"]
header(ws,r,heads,[12,26,26,24,12,7,24,22]); r+=1
data=[
 ("問診","紙問診を毎日手入力・受付過多","AI問診（Ubie等）をiPadで。事前問診→カルテ下書き→ワンクリック転記","Cloud Connect APIで患者IDキーに取込の実例あり","ソフト対象（iPadはインボイス枠で一部）","◎","Ubie AI問診 intro.dr-ubie.com","Ubieに京橋＝Medicom連携可否を確認"),
 ("電話","鳴り止まない電話で受付パンク（既存アイコール導入済）","AI音声IVRが一次対応・LINE誘導・営業電話ブロック","電カル非依存（電話側）","ソフト対象","○","IVRy 公開事例","既存アイコールと併存/置換・契約縛りを確認"),
 ("紹介状","作成の二度手間・診療後残業","紹介状AI（カルテから取込→下書き）","ひろつ内科＝CLINICSから取込で5-15分→3分","ソフト対象","○","ひろつ内科 hirotsu.clinic","Medicom取込の可否／3省2GL準拠を書面確認"),
 ("音声カルテ","会話を手入力でカルテ化","アンビエントAI（Ubie生成AI/MEDISMA AIクラーク等）会話→SOAP下書き","連携設定要・Medicom対応は要確認","ソフト対象","○","MEDISMA/OPTiM 公式","費用見積／Medicom連携／3省2GLを確認"),
 ("レセ点検","月初レセプトが毎月重い","レセ点検AI（算定漏れ自動チェック）","レセコン連携（Medicom-HRf）","ソフト対象","○","各レセ点検AI","Medicom対応製品を絞る"),
 ("定型文/入力","最初の文字入力・定型文の手間","Stream Deck（物理ボタンで定型文/マクロ）＋テキスト展開（ペースター等）","電カル非依存（キー操作で挿入）","ハード少額/ソフト","◎ 即効・低コスト","Stream Deck医療活用 hospital-management.net","宮崎が実機検証（日本語はSuperMacro対策）"),
 ("薬照合","薬セットの取り違え・期限","GS1バーコードで処方オーダーと照合","Medicom連携・読取データ取込は要確認","一部対象","○ 安全性","GS1 Japan gs1jp.org","Medicomのバーコード連携を販社確認"),
 ("画像所見","胸部CT・XPの最初の変換・所見","⚠️読影AIは薬機法SaMD（承認必要）＝別レベル。所見定型文はStream Deck等で時短","電カル非依存（定型文側）","読影AIは別／定型文は対象外","△ 規制注意","富士CXR-AID等は承認済","「読影AI」と「定型文時短」を分けて提案"),
 ("予約/HP","LINE・HP導入の検討","Lステップ（予約・問診・リマインド）＋HP","既存アイコールと連携整理","ソフト対象","○","Lステップ公開事例","アイコールとの役割分担を整理"),
 ("データ抽出","薬等の入力情報の自動抽出","Medicom Cloud Connect API活用で外部取込","公式にAPI実在（具体仕様は非公開）","ケースによる","△ 要確認","PHC Medicom 公式","中央ビジコン/ウィーメックスに連携仕様確認"),
 ("証明書","障害年金・化学物質過敏症の証明書（書ける医師が全国数名）","まず情報整理（専門医ネットワーク・更新時記載要否）","別件（電カル外）","対象外","△ 調査","—","みよし先生5000名等の現状を情報整理"),
]
for d in data:
    ink={6: (GREEN if d[5]=="◎" else (RED if d[5]=="△" else AMBER))}
    bold={6:True}
    bg=BEIGE if d[5]=="◎" else WHITE
    row(ws,r,list(d),h=54,bg=bg,inks=ink,bolds=bold); r+=1
r+=1
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=8)
ws.cell(r,1,"★最有力の入口＝①AI問診(iPad) ②Stream Deck定型文（即効・低コスト・月曜デモ向き） ③紹介状AI（事例が固い）。読影AIは薬機法のため切り離す。").font=fn(10,True,RED)
ws.cell(r,1).alignment=al(); ws.row_dimensions[r].height=34

# ============== 02 iPad導入の段取り ==============
ws=newsheet("02_iPad導入の段取り")
r=title_block(ws,"iPad ROLLOUT ｜ 2 TRACKS",
    "iPad導入の段取り ── 「月曜デモ」と「補助金」は別トラック",
    "李牧の注意：補助金とKHDの事業者登録は数週間かかる。月曜デモは補助金に依存させない。",6)
for w,c in zip([20,40,16,16,16,12],range(1,7)):
    ws.column_dimensions[get_column_letter(c)].width=w
r+=1
# トラックA
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
ws.cell(r,1,"【トラックA】即効・月曜デモ（補助金に依存しない）").fill=fl(BEIGE)
ws.cell(r,1).font=fn(12,True,RED); ws.cell(r,1).alignment=al(); ws.row_dimensions[r].height=22; r+=1
header(ws,r,["ステップ","内容","担当","費用感","期限","状態"],[20,40,12,14,12,10]); r+=1
trackA=[
 ("①機種選定","最新iPad（無印 or iPad Air）+ スタンド/ケース。AI問診・定型文デモ用","菊池","iPad約7-10万＋付属","6/19",""),
 ("②購入","KHD名義で購入（経費）。Apple/量販で即日〜","菊池","—","6/20",""),
 ("③ツール設定","AI問診(Ubie等トライアル)/定型文/Stream Deck接続をセットアップ","菊池/宮崎","トライアルは無償枠","6/21",""),
 ("④現地デモ","月曜 現地で先生・事務に実演（紙問診→iPad問診、定型文ワンタッチ）","菊池","—","6/22",""),
]
for d in trackA: row(ws,r,list(d),h=34,bg=WHITE); r+=1
r+=1
# トラックB
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
ws.cell(r,1,"【トラックB】IT導入補助金トラック（数週間・後追いで本導入）").fill=fl(BEIGE)
ws.cell(r,1).font=fn(12,True,RED); ws.cell(r,1).alignment=al(); ws.row_dimensions[r].height=22; r+=1
header(ws,r,["ステップ","内容","担当","費用感","期限","状態"],[20,40,12,14,12,10]); r+=1
trackB=[
 ("①KHD事業者登録","KHDをIT導入支援事業者として登録＋ITツール登録（審査あり・数週間）","菊池","—","要確認",""),
 ("②共同申請","クリニックが申請者＝KHDと共同申請。gBizID・SECURITY ACTION","菊池+先生","—","登録後",""),
 ("③交付決定後に発注","交付決定の後にツール契約・iPad購入（先に買うと対象外）","—","補助1/2等","決定後",""),
 ("④実績報告→交付","導入・実績報告→補助金交付","菊池","—","後日",""),
]
for d in trackB: row(ws,r,list(d),h=34,bg=WHITE); r+=1
r+=1
# 先生提出書類
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
ws.cell(r,1,"■ 先生（クリニック）に出してもらう書類（トラックB・要・最新公募要領で最終確認）").fill=fl(RED)
ws.cell(r,1).font=fn(11,True,WHITE); ws.cell(r,1).alignment=al(); ws.row_dimensions[r].height=22; r+=1
docs=["gBizIDプライム（発行に2-3週間→最優先）","直近の決算書／確定申告書（医療法人 or 個人）",
      "本人確認書類","SECURITY ACTION（★）自己宣言","納税証明（求められる場合）","（導入ツールの見積・申込関連）"]
for d in docs:
    ws.cell(r,1,"▢ "+d).font=fn(10); ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
    ws.cell(r,1).alignment=al(); ws.row_dimensions[r].height=18; r+=1
r+=1
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=6)
ws.cell(r,1,"⚠️ iPad(ハード)はIT補助金で原則対象外。インボイス枠で上限10万・1/2まで（要件付・要確認）。月曜デモはKHD経費で先行し、補助金は本導入時に。").font=fn(10,True,RED)
ws.cell(r,1).alignment=al(); ws.row_dimensions[r].height=34

# ============== 03 補助金まとめ ==============
ws=newsheet("03_IT補助金まとめ")
r=title_block(ws,"SUBSIDY SUMMARY",
    "IT導入補助金（2026〜「デジタル化・AI導入補助金」）まとめ",
    "クリニックは申請可。金額は提出前に公式公募要領で再確認（名称・定義が2026変更）。",4)
for w,c in zip([22,34,26,22],range(1,5)):
    ws.column_dimensions[get_column_letter(c)].width=w
r+=1
header(ws,r,["項目","内容","数字/条件","出典・確認"],[22,34,26,22]); r+=1
sub=[
 ("対象者","医療法人（従業員300人以下）・個人クリニックも対象","小規模事業者は補助率UP","it-shien.smrj.go.jp"),
 ("通常枠","業務効率化ソフト","1/2（賃上げ2/3）・5万〜450万","要・公募要領"),
 ("インボイス枠","ソフト＋ハード（PC/タブレット）","ソフト3/4上限350万・ハード1/2上限10万","要・公募要領"),
 ("セキュリティ枠","対策ソフト","1/2（小規模2/3）・5万〜150万","—"),
 ("必須条件","IT導入支援事業者と共同申請／登録ツールのみ／gBizID／SECURITY ACTION","クリニック単独申請は不可","公式"),
 ("発注タイミング","交付決定の後に契約・発注","先に契約すると対象外","公式"),
 ("別制度","医療情報化支援基金=オン資/電子処方箋/標準型カルテ","同一経費の二重取り不可","mhlw.go.jp"),
]
for d in sub: row(ws,r,list(d),h=40,bg=WHITE); r+=1
r+=1
ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
ws.cell(r,1,"要確認(6/19まで)：①導入ツールが事前登録済か（公式ITツール検索）②2026各枠の確定額③KHD事業者登録の要件・所要期間④Medicom連携可否").font=fn(10,True,RED)
ws.cell(r,1).alignment=al(); ws.row_dimensions[r].height=40

if "Sheet" in wb.sheetnames: del wb["Sheet"]
wb.active=0
out="/Users/kikuchikenta/Library/CloudStorage/GoogleDrive-ec-seller@kikuchi-hd.net/マイドライブ/260526_AI医療コンサル/京橋_AI-DX提案マトリクス_課題と対策.xlsx"
wb.save(out)
print("SAVED:",out); print("SHEETS:",wb.sheetnames)

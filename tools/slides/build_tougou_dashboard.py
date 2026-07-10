# -*- coding: utf-8 -*-
# KHD 統合ダッシュボード v1（2026-06-02）
# 設計＝菊池の「過去→現在→未来＋行動の4軸／本部番号付き全事業マトリクス／時間×金×家族／日次ループ」
# 既存 build_zaimu_cockpit.py(財務5タブ)を土台に、足りない3要素を追加して1ファイルに統合。
# タブ: ①統合司令塔(4軸1枚) ②本部マトリクス(時間×金×家族) ③資産負債SSoT ④資金繰り予測 ⑤借入SSoT ⑥日次ループ・使い方
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---- スタイル（スライド標準＝レンガ赤×クリーム） ----
H   = Font(bold=True, size=12, color="FFFFFF")
H14 = Font(bold=True, size=14, color="FFFFFF")
HF  = PatternFill("solid", fgColor="AA2E26")   # レンガ赤
SUB = Font(bold=True, size=11)
SUBF= PatternFill("solid", fgColor="F0E2DF")
HI  = PatternFill("solid", fgColor="DDF3DD")   # ハイ/緑
MID = PatternFill("solid", fgColor="FFF4D6")   # 中/黄
LOW = PatternFill("solid", fgColor="ECECEC")   # 非流動/灰
FAM = PatternFill("solid", fgColor="CDE9D6")   # 家族＝目的(濃緑)
AX_P= PatternFill("solid", fgColor="EAE0DA")   # 過去
AX_N= PatternFill("solid", fgColor="FCEFE7")   # 現在
AX_F= PatternFill("solid", fgColor="E7EEF6")   # 未来
AX_A= PatternFill("solid", fgColor="FBF3D6")   # 行動
YEN = '#,##0"円"'
HRS = '0.0"h"'
ROI = '#,##0"円/h"'
thin= Side(style="thin", color="CCCCCC")
BD  = Border(left=thin,right=thin,top=thin,bottom=thin)
def st(c, font=None, fill=None, num=None, al=None, bd=False, wrap=False):
    if font:c.font=font
    if fill:c.fill=fill
    if num:c.number_format=num
    if al or wrap:c.alignment=Alignment(horizontal=al or "left",vertical="center",wrap_text=wrap)
    if bd:c.border=BD

# ============================================================
# ③ 資産負債SSoT（cockpit v1から流用＝単一入力源）
# ============================================================
ss = wb.active; ss.title="③資産負債SSoT"
ss["A1"]="③ 資産負債 SSoT（4主格 × 流動性 × 月次）— 毎月初ここだけ更新"; st(ss["A1"],H,HF)
for col in "BCDEFG": st(ss[col+"1"],fill=HF)
ss["A2"]="基準日：2026/06/01（毎月初に「6/1残高」列を更新→他タブは自動連動）"
hdr=["区分","主格","種類","口座/名称","流動性","6/1残高","データ元/備考"]
for i,h in enumerate(hdr):
    st(ss.cell(4,i+1,h),SUB,SUBF,al="center",bd=True)
assets=[
 ("資産","①法人","銀行","城北信金","ハイ",281012,"260601_城北信金_明細.csv"),
 ("資産","①法人","銀行","法人SBI(代表)","ハイ",139320,"6/1実読・5/19役員借入で減"),
 ("資産","①法人","銀行","朝日信金(普通)","ハイ",25000,""),
 ("資産","①法人","銀行","朝日信金(積立)","ハイ",360000,"付合い定額積立"),
 ("資産","①法人","銀行","大東京信金","ハイ",769486,"※6/1未更新(5/1値)"),
 ("資産","①法人","銀行","法人TB(東京ベイ)","ハイ",1677563,"※6/1未更新(5/1値)"),
 ("資産","①法人","保険","経営セーフティ共済","非流動",1100000,"解約返戻・拘束"),
 ("資産","②③研太","銀行","個人事業SBI(2613479)","ハイ",1224610,"260601_個人事業SBI_明細.csv"),
 ("資産","②③研太","銀行","みずほ","ハイ",71,""),
 ("資産","②③研太","銀行","ゆうちょ","ハイ",243582,"※6/1未更新(5/1値)"),
 ("資産","②③研太","銀行","楽天銀行RB1","ハイ",97587,"※6/1未更新(5/1値)"),
 ("資産","②③研太","銀行","楽天銀行RB2","ハイ",1674,""),
 ("資産","②③研太","銀行","東京ベイTB(研太)","ハイ",404939,"260601_東京ベイ信金_明細.csv"),
 ("資産","②③研太","現金","現金","ハイ",40000,""),
 ("資産","②③研太","証券","楽天証券(NISA+特定/持株/DC)","中",11287071,"※6/1未更新・株式評価"),
 ("資産","②③研太","金","金(研太)","非流動",2000000,""),
 ("資産","②③研太","現物","現物","中",250000,""),
 ("資産","②③研太","不動産","自宅","非流動",42472875,"住宅ローン対応"),
 ("資産","②③研太","不動産","Kハウス北千住","中",13000000,""),
 ("資産","②③研太","その他","マリパパ譲渡","中",2000000,""),
 ("資産","②③研太","年金","年金(研太)","非流動",3300000,"将来資産"),
 ("資産","④麻梨奈","銀行","麻梨奈SBI2(普通)","ハイ",1117644,"※6/1未更新(5/1値)"),
 ("資産","④麻梨奈","証券","麻梨奈SBI2(証券)","中",2763600,""),
 ("資産","④麻梨奈","証券","麻梨奈 楽天2 NISA","非流動",2444704,""),
 ("資産","④麻梨奈","金","金(麻梨奈)","非流動",2000000,""),
 ("資産","④麻梨奈","年金","年金(麻梨奈)","非流動",3000000,"将来資産"),
]
liabs=[
 ("負債","①法人","借入","城北信金(東京保証)","-",2833330,"0返済予定表 r409"),
 ("負債","①法人","借入","朝日信金(東京保証)","-",600000,"0返済予定表 r474"),
 ("負債","①法人","借入","大東京(オリコ保証)","-",792000,"0返済予定表 r525"),
 ("負債","①法人","カード","JAL VIEWカード未払","-",1907342,""),
 ("負債","②③研太","借入","住宅ローン(自宅)","-",42349508,"0返済予定表 r287"),
 ("負債","②③研太","借入","公庫コロナ","-",1238000,"0返済予定表 r138"),
 ("負債","②③研太","借入","TB浦安創業","-",3112000,"0返済予定表 r225"),
 ("負債","②③研太","借入","TBセゾン不動産","-",4739240,"0返済予定表 r15"),
 ("負債","②③研太","カード","カード未払(MB/AGP)","-",218558,""),
 ("負債","②③研太","カード","カード未払(R1楽天)","-",180921,""),
 ("負債","②③研太","カード","PayPayカード","-",10000,""),
]
r=5; first_asset=r
for row in assets+liabs:
    for i,v in enumerate(row):
        c=ss.cell(r,i+1,v); st(c,bd=True)
        if i==5: st(c,num=YEN,al="right")
    fill={"ハイ":HI,"中":MID,"非流動":LOW}.get(row[4])
    if fill: st(ss.cell(r,5),fill=fill,al="center")
    r+=1
last=r-1
r+=1
def line(lbl, formula, bold=True, fill=None):
    global r
    st(ss.cell(r,4,lbl),SUB if bold else None,fill)
    st(ss.cell(r,6,formula),SUB if bold else None,fill,num=YEN,al="right")
    r+=1
line("流動性ハイ（即現金化）", f'=SUMIFS(F{first_asset}:F{last},E{first_asset}:E{last},"ハイ")', fill=HI)
line("中（数ヶ月）", f'=SUMIFS(F{first_asset}:F{last},E{first_asset}:E{last},"中")', fill=MID)
line("非流動", f'=SUMIFS(F{first_asset}:F{last},E{first_asset}:E{last},"非流動")', fill=LOW)
line("総資産", f'=SUMIF(A{first_asset}:A{last},"資産",F{first_asset}:F{last})')
line("総負債", f'=SUMIF(A{first_asset}:A{last},"負債",F{first_asset}:F{last})')
networth_row=r
line("★純資産（自己資本）", f'=F{networth_row-2}-F{networth_row-1}', fill=SUBF)
ss.column_dimensions["D"].width=30; ss.column_dimensions["F"].width=14; ss.column_dimensions["G"].width=34
for col in "ABCE": ss.column_dimensions[col].width=10
SSREF="'③資産負債SSoT'"
HI_CELL   =f"{SSREF}!F{networth_row-5}"
ASSET_CELL=f"{SSREF}!F{networth_row-2}"
LIAB_CELL =f"{SSREF}!F{networth_row-1}"
NET_CELL  =f"{SSREF}!F{networth_row}"

# ============================================================
# ④ 資金繰り予測（cockpit v1から流用＝未来軸の本体）
# ============================================================
cf=wb.create_sheet("④資金繰り予測")
cf["A1"]="④ 資金繰り予測（6ヶ月・月初現金はSSoT連動）"; st(cf["A1"],H,HF)
for col in "BCDEFGHI": st(cf[col+"1"],fill=HF)
months=["2026/06","2026/07","2026/08","2026/09","2026/10","2026/11","2026/12","2027/01"]
cf.append(["項目"]+months)
for i in range(9): st(cf.cell(2,i+1),SUB,SUBF,al="center",bd=True)
cols=[get_column_letter(i+2) for i in range(8)]
def crow(label, vals):
    cf.append([label]+vals); rr=cf.max_row
    for i in range(9):
        c=cf.cell(rr,i+1); st(c,bd=True)
        if i>=1: st(c,num=YEN,al="right")
    return rr
cf.append(["■入金"])
ec_in_栄町=crow("物件売却(栄町6月)",[1900000,0,0,0,0,0,0,0])
医療=crow("医療コンサル(9月想定)",[0,0,0,660000,0,0,0,0])
ecr=crow("EC粗利(クーパン1+2)",[150000]*8)
iku=crow("麻梨奈 育休給付(月割)",[156659]*8)
in_first=ec_in_栄町; in_last=iku
in_tot=crow("入金合計",[f"=SUM({c}{in_first}:{c}{in_last})" for c in cols])
cf.append(["■出金"])
fixed_rows=[
 ("法人:税理士(ZH橋本)",110000),("法人:外注(江藤)",99000),("法人:社保",22167),
 ("個人事業:税理士",22000),("住宅ローン",130668),("SMBC管理費",22130),("PayPay",10000),
 ("★世帯生活費(楽天18+麻梨奈12.1)",301000),
 ("★法人借入返済(城北+朝日+大東京)",76035),("★個人事業借入返済(TB2+公庫)",145001),
 ("投信積立(任意)",100000),("役員賞与(見送り確定=0)",0),
]
out_first=None
for lbl,val in fixed_rows:
    rr=crow(lbl,[val]*8)
    if out_first is None: out_first=rr
    out_last=rr
out_tot=crow("出金合計",[f"=SUM({c}{out_first}:{c}{out_last})" for c in cols])
net=crow("当月純増減",[f"={c}{in_tot}-{c}{out_tot}" for c in cols])
end=crow("月末現金（体力計）",[0]*8)
for i in range(8):
    if i==0: cf.cell(end,2,f"={HI_CELL}+B{net}")
    else:    cf.cell(end,i+2,f"={get_column_letter(i+1)}{end}+{cols[i]}{net}")
    st(cf.cell(end,i+2),SUB,num=YEN,al="right")
cf.append([])
rwr=crow("純月次燃焼(出金−経常入金)",[f"={c}{out_tot}-{c}{ecr}-{c}{iku}-100000" for c in cols])
sig=cf.max_row+1
st(cf.cell(sig,1,"ランウェイ(月初現金÷燃焼)"),SUB)
st(cf.cell(sig,2,f"={HI_CELL}/B{rwr}"),SUB,num='0.0"ヶ月"')
st(cf.cell(sig,3,f'=IF(B{sig}<3,"🔴守り:投資凍結",IF(B{sig}<6,"🟡注意","🟢攻めOK"))'),SUB)
cf.column_dimensions["A"].width=30
for c in cols: cf.column_dimensions[c].width=12

# ============================================================
# ② 本部マトリクス（時間 × 金 × 家族）＝予実DBの本体
# ============================================================
BURN=631342   # 世帯 純月次燃焼(memory project_household_cashflow)。家族◯ヶ月分の分母。
mx=wb.create_sheet("②本部マトリクス")
mx["A1"]="② 本部マトリクス（時間×金×家族）— パパの1時間がいくらを生み、家族の暮らし何ヶ月分になるか"; st(mx["A1"],H,HF)
for col in "BCDEFGHIJKLM": st(mx[col+"1"],fill=HF)
# 妻と見る1行サマリー（プレーン言語＋計算値）
mx["A2"]=("👨‍👩‍👦 妻と見る：高い活動に時間を寄せる＝将来の家族時間を“買う”こと。"
          "実績h=カレンダー自動／見込み粗利×確度＝期待粗利／円per h＝1時間の価値／家族◯ヶ月分＝その金で暮らせる月数(世帯燃焼63万)。")
mx.merge_cells("A2:M2")
st(mx["A2"],None,FAM,wrap=True)
mh=["本部","活動","実績h(月)","構成比","見込み粗利","確度","期待粗利","円/h(期待)","家族◯ヶ月分","営業直結","判断","投下予定h","メモ"]
for i,h in enumerate(mh): st(mx.cell(4,i+1,h),SUB,SUBF,al="center",bd=True,wrap=True)
# (本部, 活動, 見込み粗利満額, 確度, 投下予定h, 営業直結, 判断, メモ)  ※実績h=0でseed→カレンダー自動書込
biz=[
 ("05 物件調達","不動産 売却/買取(栄町6/20決済)",1900000,0.9,40,"営業","続ける(6月集中)","6/20決済・確度90%。一回性なので継続源(医療)と両睨み"),
 ("04 コンサル","医療テナント/承継コンサル",660000,0.5,60,"営業","増やす","1件66万・継続性◎の本命現金源"),
 ("03 事業運営","EC 韓国輸出(クーパン)",150000,0.9,40,"","維持(黒字回転)","月次継続・粗利は各DBから月末転記"),
 ("04 調査士","土地家屋調査士(将来ROI)",0,0,80,"仕込","続ける","将来ROI・今期見込0。2027合格で単価UP"),
 ("本命","メディア×AI(YouTube/HP/MyAI)",0,0,50,"仕込","増やす","将来の継続収益源・今は仕込み"),
 ("03/05 協働","買取再販テレアポ・採用(宮崎)",100000,0.3,25,"営業","様子見","オーロラ次第・確度低の補助線"),
 ("01-03 内務","経営/資金/運営(朝礼終礼・台帳等)",0,0,60,"","減らす","★最大の塊=コスト。圧縮し営業へ振替"),
 ("-","その他",0,0,10,"","減らす","映画・雑など。意図的に減らす"),
 ("00 家族","親子/夫婦(目的・死守)",None,None,125,"目的","死守","稼ぐ目的そのもの。ROI対象外・増やすのがゴール"),
]
mr=5; first_mx=mr
for b in biz:
    honbu,name,mikomi,kakudo,plan_h,eig,handan,memo=b
    fam=(eig=="目的")
    rev=(isinstance(mikomi,(int,float)) and mikomi>0)   # 当期に金を生む行か
    st(mx.cell(mr,1,honbu),bd=True)
    st(mx.cell(mr,2,name),bd=True)
    st(mx.cell(mr,3,0),bd=True,num=HRS,al="right")                          # C 実績h(カレンダー自動)
    st(mx.cell(mr,4,0),bd=True,num='0.0%',al="right")                       # D 構成比(下で式)
    if rev:
        st(mx.cell(mr,5,mikomi),bd=True,num=YEN,al="right")                 # E 見込み粗利
        st(mx.cell(mr,6,kakudo),bd=True,num='0%',al="center")               # F 確度
        st(mx.cell(mr,7,f"=E{mr}*F{mr}"),bd=True,num=YEN,al="right")        # G 期待粗利
        st(mx.cell(mr,8,f"=IF(L{mr}=0,\"-\",G{mr}/L{mr})"),bd=True,num=ROI,al="right")        # H 円/h(期待)
        st(mx.cell(mr,9,f"=G{mr}/{BURN}"),bd=True,num='0.0"ヶ月"',al="right")                  # I 家族◯ヶ月分
    else:
        for col in (5,6,7,8,9): st(mx.cell(mr,col,"-"),bd=True,al="center")
    st(mx.cell(mr,10,eig),bd=True,al="center")
    st(mx.cell(mr,11,handan),bd=True,al="center")
    st(mx.cell(mr,12,plan_h),bd=True,num=HRS,al="right")
    st(mx.cell(mr,13,memo),bd=True,wrap=True)
    if fam:
        for col in range(1,14): st(mx.cell(mr,col),fill=FAM)
    mr+=1
last_mx=mr-1
tot_row=mr+1
# 構成比の式
for rr in range(first_mx,last_mx+1):
    mx.cell(rr,4,f"=IF($C${tot_row}=0,0,C{rr}/$C${tot_row})")
# 合計行
st(mx.cell(tot_row,2,"合計（全活動＝総時間）"),SUB,SUBF)
st(mx.cell(tot_row,3,f"=SUM(C{first_mx}:C{last_mx})"),SUB,SUBF,num=HRS,al="right")
st(mx.cell(tot_row,4,f"=IF(C{tot_row}=0,0,1)"),SUB,SUBF,num='0%',al="right")
st(mx.cell(tot_row,7,f"=SUM(G{first_mx}:G{last_mx})"),SUB,SUBF,num=YEN,al="right")   # 期待粗利合計
st(mx.cell(tot_row,9,f"=SUM(I{first_mx}:I{last_mx})"),SUB,SUBF,num='0.0"ヶ月"',al="right")  # 家族月数合計
st(mx.cell(tot_row,12,f"=SUM(L{first_mx}:L{last_mx})"),SUB,SUBF,num=HRS,al="right")
# 妻サマリー（計算値・合計行を参照）
sp_row=tot_row+1
st(mx.cell(sp_row,2,"👨‍👩‍👦 今月の期待粗利 合計 →"),SUB,FAM)
st(mx.cell(sp_row,7,f"=G{tot_row}"),Font(bold=True,size=12),FAM,num=YEN,al="right")
st(mx.cell(sp_row,8,"＝家族"),SUB,FAM,al="right")
st(mx.cell(sp_row,9,f"=I{tot_row}"),Font(bold=True,size=12),FAM,num='0.0"ヶ月"',al="center")
st(mx.cell(sp_row,10,"分の暮らしを確保。だから今は時間を寄せる。"),None,FAM,wrap=True)
mx.merge_cells(start_row=sp_row,start_column=10,end_row=sp_row,end_column=13)
# 営業直結％行（秘書KPI 目標60%）
eig_row=sp_row+1
st(mx.cell(eig_row,2,"★営業直結比率(不動産+医療+テレアポ)／目標60%"),SUB)
st(mx.cell(eig_row,3,f'=IF(C{tot_row}=0,"-",SUMIF(J{first_mx}:J{last_mx},"営業",C{first_mx}:C{last_mx})/C{tot_row})'),Font(bold=True,size=12),HI,num='0%',al="right")
note_row=eig_row+2
st(mx.cell(note_row,1,"【日次】予定どおり時間を使えたか(カレンダー実測)。【週次】営業直結%とメディア時間を点検し配分調整。【月次】粗利確定→実績で見込みを更新し続ける/減らす/やめるを決定。"),None,wrap=True)
mx.merge_cells(start_row=note_row,start_column=1,end_row=note_row,end_column=13)
mx.column_dimensions["A"].width=13; mx.column_dimensions["B"].width=30; mx.column_dimensions["M"].width=34
for col in "CDEFGHIJKL": mx.column_dimensions[col].width=11
MXREF="'②本部マトリクス'"
MX_HOURS=f"{MXREF}!C{tot_row}"; MX_EXP=f"{MXREF}!G{tot_row}"; MX_FAMMO=f"{MXREF}!I{tot_row}"
MX_EIGYO=f"{MXREF}!C{eig_row}"; MX_FAM_H=f"{MXREF}!C{last_mx}"

# ============================================================
# ① 統合司令塔（過去 → 現在 → 未来 ＋ 行動 の4軸を1枚）
# ============================================================
db=wb.create_sheet("①統合司令塔",0)
db.sheet_view.showGridLines=False
db["A1"]="① 統合司令塔 — 過去 → 現在 → 未来 ＋ 行動"; st(db["A1"],H14,HF)
for col in "BCDE": st(db[col+"1"],fill=HF)
db["A2"]="基準日 2026/06/02（②本部マトリクス／③SSoT／④予測 と自動連動）"
def block(rr,title,fill):
    st(db.cell(rr,1,title),Font(bold=True,size=12),fill)
    for col in range(2,5): st(db.cell(rr,col),fill=fill)
def kv(rr,label,formula,num=YEN,note=""):
    st(db.cell(rr,1,label),SUB,SUBF)
    st(db.cell(rr,2,formula),Font(bold=True,size=12),num=num,al="right")
    if note: st(db.cell(rr,3,note),None,wrap=True)

# ── 過去（実績ストック）──
block(4,"◆ 過去（実績ストック・②③連動）",AX_P)
kv(5,"純資産（自己資本）",f"={NET_CELL}",note="世間中央値超だが流動性は薄い")
kv(6,"総資産",f"={ASSET_CELL}")
kv(7,"総負債",f"={LIAB_CELL}",note="法人は債務超過に注意")
# ── 現在（今の体力）──
block(9,"◆ 現在（今の体力・③④連動）",AX_N)
kv(10,"現預金（流動性ハイ）",f"={HI_CELL}",note="即現金化できる額")
st(db.cell(11,1,"ランウェイ"),SUB,SUBF)
st(db.cell(11,2,f"='④資金繰り予測'!B{sig}"),Font(bold=True,size=12),num='0.0"ヶ月"')
st(db.cell(11,3,f"='④資金繰り予測'!C{sig}"),Font(bold=True,size=12))
# ── 未来（予測の谷）──
block(13,"◆ 未来（資金繰り予測・④連動）",AX_F)
st(db.cell(14,1,"通期末(2027/1)現金"),SUB,SUBF)
st(db.cell(14,2,f"='④資金繰り予測'!I{end}"),Font(bold=True,size=12),num=YEN,al="right")
st(db.cell(15,1,"最小月末現金(谷)"),SUB,SUBF)
st(db.cell(15,2,f"=MIN('④資金繰り予測'!B{end}:I{end})"),Font(bold=True,size=12),num=YEN,al="right")
st(db.cell(15,3,"マイナスなら谷=ショート月あり。月初に必ず確認。"),None,wrap=True)
# ── 行動（本部マトリクス＝時間×金×家族・②連動）──
block(17,"◆ 行動（時間×金×家族・②連動）",AX_A)
st(db.cell(18,1,"★営業直結比率(目標60%)"),SUB,SUBF)
st(db.cell(18,2,f"={MX_EIGYO}"),Font(bold=True,size=12),HI,num='0%',al="right")
st(db.cell(18,3,"不動産+医療+テレアポ。内務に溶けていないか毎週確認。"),None,wrap=True)
st(db.cell(19,1,"今月 期待粗利→家族換算"),SUB,FAM)
st(db.cell(19,2,f"={MX_EXP}"),Font(bold=True,size=12),FAM,num=YEN,al="right")
st(db.cell(19,3,f'=CONCATENATE("＝家族 ",TEXT({MX_FAMMO},"0.0"),"ヶ月分の暮らし")'),Font(bold=True,size=11),FAM)
st(db.cell(20,1,"今月 家族時間(目的)／総実績h"),SUB,SUBF)
st(db.cell(20,2,f"={MX_FAM_H}"),Font(bold=True,size=11),FAM,num=HRS,al="right")
st(db.cell(20,3,f"={MX_HOURS}"),Font(bold=True,size=11),num=HRS,al="right")
st(db.cell(21,1,"次の一手（手記入）"),SUB,SUBF)
st(db.cell(21,2,"内務を削り営業へ／栄町6/20前倒し／医療もう1件"),None,wrap=True)
db.merge_cells("B21:E21")
# ── 今月の判断ポイント ──
st(db.cell(22,1,"■ 今月の判断ポイント（絞る）"),SUB)
for i,t in enumerate([
 "□ 法人現預金のうち運転資金(EC仕入で消える)と自由現金の切り分け＝真のランウェイ確定の鍵",
 "□ ランウェイ🟢6ヶ月以上で初めて『資産拡大の投資』に現金を回す",
 "□ 時間ROI(粗利/h)が最も高い事業に翌月の時間を寄せる。低ROIは縮小・撤退を検討",
]): st(db.cell(23+i,1,t),None,wrap=True)
db.column_dimensions["A"].width=30; db.column_dimensions["B"].width=18; db.column_dimensions["C"].width=40
db.column_dimensions["D"].width=10; db.column_dimensions["E"].width=10

# ============================================================
# ⑤ 借入SSoT（証票＝0返済予定表ベース）
# ============================================================
ln=wb.create_sheet("⑤借入SSoT")
ln["A1"]="⑤ 借入 SSoT（残高の単一真実源・証票=0返済予定表）"; st(ln["A1"],H,HF)
for col in "BCDEFG": st(ln[col+"1"],fill=HF)
lh=["主格","借入先","残高","金利","月返済","証票/リンク","備考"]
ln.append([""])
for i,h in enumerate(lh): st(ln.cell(2,i+1,h),SUB,SUBF,al="center",bd=True)
loans=[
 ("①法人","城北信金(東京保証)",2833330,"1.8%固定",21253,"0返済予定表 r409",""),
 ("①法人","朝日信金(東京保証)",600000,"2.575%変動",17732,"0返済予定表 r474","年1回11月改定"),
 ("①法人","大東京(オリコ保証)",792000,"2.8%固定",18340,"0返済予定表 r525","250929確認"),
 ("②③研太","公庫コロナ",1238000,"1.21%固定",22531,"0返済予定表 r138","利子補給終了"),
 ("②③研太","TB浦安創業",3112000,"2.1%変動",91348,"0返済予定表 r225","R8.2.2金利変更"),
 ("②③研太","TBセゾン不動産",4739240,"要確認","-","0返済予定表 r15",""),
 ("②③研太","住宅ローン(自宅)",42349508,"要確認",130668,"0返済予定表 r287","2065年"),
]
for row in loans:
    ln.append(list(row)); rr=ln.max_row
    for i in range(7):
        c=ln.cell(rr,i+1); st(c,bd=True)
        if i==2: st(c,num=YEN,al="right")
tr=ln.max_row+2
st(ln.cell(tr,2,"借入合計"),SUB)
st(ln.cell(tr,3,f"=SUM(C3:C{tr-2})"),SUB,num=YEN,al="right")
ln.column_dimensions["B"].width=22; ln.column_dimensions["C"].width=14
ln.column_dimensions["F"].width=20; ln.column_dimensions["G"].width=20

# ============================================================
# ⑥ 日次ループ・使い方（夜=抽出→タスク化／朝=ブリーフ→再配置）
# ============================================================
hp=wb.create_sheet("⑥日次ループ・使い方")
hp["A1"]="⑥ 日次ループ — 夜に抽出してタスク化 → 朝にブリーフして再配置"; st(hp["A1"],H,HF)
for col in "BCD": st(hp[col+"1"],fill=HF)
rows=[
 ("",""),
 ("🌙 夜ループ（仕込み・抽出→タスク化）","18時以降は家族最優先。家族明けの短時間で回す"),
 ("  1. その日の進捗・学び・気づきを抽出","秘書がnotes/learningsから自動抽出も可"),
 ("  2. 抽出をタスク化（誰を・いつまでに・どう動かす）","Google Tasks(SSoT)へ投入。プッシュ営業視点"),
 ("  3. 翌日の最優先1〜3件を確定","明日の『起点1つ』を決めて寝る"),
 ("",""),
 ("🌅 朝ループ（ブリーフ→再配置）","feedback_morning_brief 準拠"),
 ("  1. カレンダー今週分＋Google Tasks未完了全件を1枚に統合",""),
 ("  2. ①期限切れ ②今日 ③今週 ④将来 に仕分け","期限切れは件数明示"),
 ("  3. 配置換え：今日やる/今週へ/凍結/完了 を菊池がYES/NOで決定","利益直結=営業60%を死守"),
 ("  4. ①統合司令塔で過去→現在→未来＋行動を確認","数字を見てから動く"),
 ("",""),
 ("【月初ルーティン】","毎月初の5分"),
 ("  ・③資産負債SSoTの『6/1残高』列をMF/明細から更新","他タブは自動連動"),
 ("  ・②本部マトリクスに先月の実績粗利・実績hを記入","時間ROIで翌月の時間配分を決める"),
 ("  ・①司令塔のランウェイ信号と谷(最小月末現金)を確認","🟢6ヶ月以上で投資解禁"),
 ("",""),
 ("【沼化の歯止め】","秘書が毎回チェック"),
 ("  ・このダッシュは『判断を速くする表示層』。配線の作り込みは本業を食う","完成磨き込み禁止"),
 ("  ・カレンダー自動集計・ライブ連携はv1では手更新。必要になったら最小で足す","done_criteria準拠"),
 ("",""),
 ("【データ元】","SSoT本体=メインWB収支タブ/新スプシv2(1ofLJOFuW…)"),
 ("  ・残高=MF優先(自動読取)、MF非連携(信金3)のみ手動","月初DLは記帳/税務用"),
 ("  ・借入残高=0返済予定表が証票(0借入一覧は使わない)",""),
]
for i,(a,b) in enumerate(rows):
    st(hp.cell(i+3,1,a))
    if b: st(hp.cell(i+3,2,b))
    if a.startswith(("🌙","🌅","【")): st(hp.cell(i+3,1),SUB,SUBF)
hp.column_dimensions["A"].width=52; hp.column_dimensions["B"].width=48

# ---- 保存 ----
out="/Users/kikuchikenta/Library/CloudStorage/GoogleDrive-ec-seller@kikuchi-hd.net/マイドライブ/KHD_統合ダッシュボード_v1.xlsx"
wb.save(out)
print("保存完了:",out)
print("タブ:", wb.sheetnames)

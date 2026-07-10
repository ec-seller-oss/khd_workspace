# -*- coding: utf-8 -*-
# W杯2026観戦ハブ スプシ生成（KHD視覚ルール: タブ番号+色分け+凡例+使い方）
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

FONT = "Yu Gothic"
C_HEAD = "AA2E26"   # レンガ赤(KHDデザイン)
C_CREAM = "F9F6EF"
C_INPUT = "FFF2CC"  # 🟡入力
C_AUTO = "EFEFEF"   # 灰=自動
C_DAILY = "FCE5CD"  # 🟧毎日
C_WEEK = "CFE2F3"   # 🟦週次
thin = Side(style="thin", color="BBBBBB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

def style_header(ws, row, cols, fill=C_HEAD, color="FFFFFF"):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(name=FONT, bold=True, color=color, size=10)
        cell.fill = PatternFill("solid", start_color=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

def set_rows(ws, start, rows, widths=None, fills=None, fontsize=10):
    for i, r in enumerate(rows):
        for j, v in enumerate(r):
            cell = ws.cell(row=start + i, column=j + 1, value=v)
            cell.font = Font(name=FONT, size=fontsize)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = BORDER
            if fills and fills[i] and fills[i][j]:
                cell.fill = PatternFill("solid", start_color=fills[i][j])
    if widths:
        for k, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(k + 1)].width = w

def title(ws, text, cols):
    ws.cell(row=1, column=1, value=text)
    ws.cell(row=1, column=1).font = Font(name=FONT, bold=True, size=14, color=C_HEAD)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)

# ============ 00_README ============
ws = wb.active
ws.title = "00_README"
ws.sheet_properties.tabColor = "999999"
title(ws, "⚽ W杯2026 観戦ハブ（オーナー視点でサッカーを楽しむ常設台帳）", 6)
rows = [
    ["", "", "", "", "", ""],
    ["■ このファイルの目的", "W杯2026を入口に『お金の動き×選手の活躍』をオーナー視点で追う趣味台帳。大会後も使い続ける。", "", "", "", ""],
    ["", "", "", "", "", ""],
    ["■ タブの使い方", "", "", "", "", ""],
    ["01_勝ち点シミュ 🟡", "日本の試合結果を入れると突破判定が自動で出る。試合のたび更新", "", "", "", ""],
    ["02_観戦スケジュール 🟧", "全試合の日本時間＆観れない日の回収ルール（カレンダー連動）", "", "", "", ""],
    ["03_日本代表名鑑", "26人＋追加招集。ブルーロック例え付き＝麻梨奈さん用ガイド", "", "", "", ""],
    ["04_追っかけ×マネー 🟦", "W杯後に移籍金が動く注目銘柄。週1（日曜21:00）で更新", "", "", "", ""],
    ["05_観戦日報 🟧", "毎日21:30のハイライト枠で1行ずつ積む（趣味の日報）", "", "", "", ""],
    ["06_TIPS_師匠とリンク集", "誰の視点を真似るか／情報源／不動産⇄サッカー対訳表", "", "", "", ""],
    ["", "", "", "", "", ""],
    ["■ 凡例（色）", "🟡黄=手で入力するセル ／ 🟧橙=毎日さわるタブ ／ 🟦青=週次タブ ／ 灰=自動計算（さわらない）", "", "", "", ""],
    ["", "", "", "", "", ""],
    ["■ 毎日のルーチン", "21:30-22:00（カレンダー登録済）: FotMobで結果→DAZN/FIFA公式YouTubeでハイライト→05に1行→お金ネタはロマーノX", "", "", "", ""],
    ["■ 週次ルーチン", "日曜21:00-21:30（カレンダー登録済）: Transfermarktで04タブの市場価値を見て更新", "", "", "", ""],
    ["", "", "", "", "", ""],
    ["■ 関連リンク", "Notion観戦計画書: https://app.notion.com/p/37e7d27fe295814aa0e2cdbde0012c97", "", "", "", ""],
]
set_rows(ws, 2, rows, widths=[26, 90, 8, 8, 8, 8])
for r in (3, 5, 13, 15, 16, 18):
    ws.cell(row=r + 1, column=1).font = Font(name=FONT, bold=True, size=11, color=C_HEAD)

# ============ 01_勝ち点シミュ ============
ws = wb.create_sheet("01_勝ち点シミュ")
ws.sheet_properties.tabColor = "FFD966"
title(ws, "🟡 日本の勝ち点シミュレーション（結果を入れるだけ）", 7)

ws.cell(row=3, column=1, value="◆ A. いまの日本（🟡黄セルに 勝/分/負 を入力）").font = Font(name=FONT, bold=True, size=11, color=C_HEAD)
head = ["試合", "対戦相手", "日時(JST)", "🟡結果", "勝ち点(自動)", "", ""]
set_rows(ws, 4, [head])
style_header(ws, 4, 5)
matches = [
    ["第1戦", "オランダ", "6/15(月) 5:00"],
    ["第2戦", "チュニジア", "6/21(日) 13:00"],
    ["第3戦", "スウェーデン", "6/26(金) 8:00"],
]
for i, m in enumerate(matches):
    r = 5 + i
    set_rows(ws, r, [[m[0], m[1], m[2], "", None]])
    ws.cell(row=r, column=4).fill = PatternFill("solid", start_color=C_INPUT)
    ws.cell(row=r, column=5, value=f'=IF(D{r}="勝",3,IF(D{r}="分",1,IF(D{r}="負",0,"")))')
    ws.cell(row=r, column=5).fill = PatternFill("solid", start_color=C_AUTO)
    ws.cell(row=r, column=5).font = Font(name=FONT)
    ws.cell(row=r, column=5).border = BORDER
ws.cell(row=8, column=4, value="合計").font = Font(name=FONT, bold=True)
ws.cell(row=8, column=5, value="=SUM(E5:E7)").font = Font(name=FONT, bold=True)
ws.cell(row=8, column=5).fill = PatternFill("solid", start_color=C_AUTO)
ws.cell(row=9, column=4, value="判定").font = Font(name=FONT, bold=True)
ws.cell(row=9, column=5, value=('=IF(E8="","結果を入力してください",IF(E8>=7,"1位通過濃厚→32強はC組2位(モロッコ/スコットランド想定)",'
    'IF(E8>=6,"突破ほぼ確実(1〜2位)",IF(E8>=5,"2位有力→32強でブラジル(C組1位想定)の可能性",'
    'IF(E8>=4,"2〜3位。突破可能性高",IF(E8=3,"3位勝負＝得失点差と他組次第",IF(E8>=1,"ほぼ敗退・奇跡待ち","敗退")))))))'))
ws.merge_cells(start_row=9, start_column=5, end_row=9, end_column=7)

dv = DataValidation(type="list", formula1='"勝,分,負"', allow_blank=True)
ws.add_data_validation(dv)
dv.add("D5:D7")

ws.cell(row=11, column=1, value="◆ B. 全27シナリオ早見表（自動・さわらない）").font = Font(name=FONT, bold=True, size=11, color=C_HEAD)
head2 = ["蘭戦", "チュニジア戦", "瑞戦", "勝ち点(自動)", "判定(自動)", "", ""]
set_rows(ws, 12, [head2])
style_header(ws, 12, 5)
r = 13
for o1 in ["勝", "分", "負"]:
    for o2 in ["勝", "分", "負"]:
        for o3 in ["勝", "分", "負"]:
            set_rows(ws, r, [[o1, o2, o3]])
            ws.cell(row=r, column=4, value=(f'=IF(A{r}="勝",3,IF(A{r}="分",1,0))+IF(B{r}="勝",3,IF(B{r}="分",1,0))'
                                            f'+IF(C{r}="勝",3,IF(C{r}="分",1,0))'))
            ws.cell(row=r, column=5, value=(f'=IF(D{r}>=7,"1位通過濃厚",IF(D{r}>=6,"突破ほぼ確実",IF(D{r}>=5,"2位有力(32強ブラジル注意)",'
                                            f'IF(D{r}>=4,"突破可能性高",IF(D{r}=3,"3位勝負(得失点差)",IF(D{r}>=1,"ほぼ敗退","敗退"))))))'))
            for c in (4, 5):
                ws.cell(row=r, column=c).font = Font(name=FONT, size=10)
                ws.cell(row=r, column=c).fill = PatternFill("solid", start_color=C_AUTO)
                ws.cell(row=r, column=c).border = BORDER
            r += 1

r0 = r + 1
ws.cell(row=r0, column=1, value="◆ C. グループF順位表（🟡試合ごとに手入力）").font = Font(name=FONT, bold=True, size=11, color=C_HEAD)
head3 = ["チーム", "試合", "勝", "分", "負", "得点-失点", "勝ち点(自動)"]
set_rows(ws, r0 + 1, [head3])
style_header(ws, r0 + 1, 7)
for i, t in enumerate(["日本", "オランダ", "スウェーデン", "チュニジア"]):
    rr = r0 + 2 + i
    set_rows(ws, rr, [[t, 0, 0, 0, 0, "0-0", None]])
    for c in range(2, 7):
        ws.cell(row=rr, column=c).fill = PatternFill("solid", start_color=C_INPUT)
    ws.cell(row=rr, column=7, value=f"=C{rr}*3+D{rr}")
    ws.cell(row=rr, column=7).fill = PatternFill("solid", start_color=C_AUTO)
    ws.cell(row=rr, column=7).font = Font(name=FONT)
    ws.cell(row=rr, column=7).border = BORDER
ws.cell(row=r0 + 7, column=1, value="※突破ライン目安: 6点=ほぼ確実／4点=有力／3点=3位勝負／12組の3位のうち成績上位8チームも突破").font = Font(name=FONT, size=9, italic=True)
for k, w in enumerate([10, 14, 14, 12, 26, 12, 40]):
    ws.column_dimensions[get_column_letter(k + 1)].width = w

# ============ 02_観戦スケジュール ============
ws = wb.create_sheet("02_観戦スケジュール")
ws.sheet_properties.tabColor = "F6B26B"
title(ws, "🟧 観戦スケジュール（全部は見れない前提の回収設計・カレンダー連動）", 7)
head = ["日時(JST)", "カード", "種別", "見方", "前後の情報枠", "ハイライト回収", "メモ"]
set_rows(ws, 3, [head])
style_header(ws, 3, 7)
sched = [
    ["6/14(日) 21:00-21:30", "—", "📋準備", "プレビュー収集枠(カレンダー済)", "スタメン予想/オランダ情報", "—", "見る:Goal/Sportsnavi"],
    ["6/15(月) 5:00", "🇳🇱オランダ vs 🇯🇵日本", "🔴日本戦", "リアタイ(NHKor日テレ/DAZN無料)", "終了後そのまま採点チェック", "—", "朝活はこの日7時以降へ"],
    ["6/15(月) 11:00", "🇸🇪スウェーデン vs 🇹🇳チュニジア", "ライバル", "結果だけ即チェック", "—", "21:30枠でハイライト", "勝ち点計算に直結"],
    ["6/21(日) 2:00", "🇳🇱オランダ vs 🇸🇪スウェーデン", "ライバル", "寝る(深夜)", "—", "21:30枠でハイライト", "日曜は家族日"],
    ["6/20(土) 21:00-21:30", "—", "📋準備", "プレビュー収集枠(カレンダー済)", "チュニジア戦スタメン予想", "—", ""],
    ["6/21(日) 13:00", "🇹🇳チュニジア vs 🇯🇵日本", "🔴日本戦", "家族でリアタイ📺", "終了後そのまま", "—", "葵斗くん初W杯"],
    ["6/25(木) 12:30-13:00", "—", "📋準備", "プレビュー収集枠(カレンダー済)", "瑞戦+突破条件の整理", "—", "夜は結婚記念日のため昼枠"],
    ["6/26(金) 8:00", "🇯🇵日本 vs 🇸🇪スウェーデン", "🔴日本戦", "リアタイ", "裏のチュニジア-オランダをスマホ", "—", "⚠️麻梨奈さん誕生日: 観戦は朝だけ宣言"],
    ["毎日 21:30-22:00", "その日の全試合", "🟧回収", "ハイライト枠(カレンダー済・毎日繰返し)", "FotMob→DAZN/FIFA YouTube", "◎ここで全試合追える", "05_観戦日報に1行"],
    ["6/28-7/3", "32強(R32)", "決勝T", "日本の試合確定後リアタイ枠化", "組合せ確定後に更新", "21:30枠", "F2位ならブラジル戦の可能性"],
    ["7/4-7/7", "16強", "決勝T", "同上", "", "21:30枠", ""],
    ["7/9-7/11", "準々決勝", "決勝T", "同上", "", "21:30枠", ""],
    ["7/14-7/15", "準決勝", "決勝T", "", "", "21:30枠", ""],
    ["7/19(日)", "決勝(MetLife)", "決勝T", "リアタイ推奨(日本時間は翌朝の見込み)", "", "21:30枠", "大会総括を05に書く"],
]
set_rows(ws, 4, sched, widths=[20, 30, 10, 32, 28, 22, 30])
for i in range(len(sched)):
    if "日本戦" in str(sched[i][2]):
        for c in range(1, 8):
            ws.cell(row=4 + i, column=c).fill = PatternFill("solid", start_color="FDE9E7")
ws.cell(row=4 + len(sched) + 1, column=1, value="■ ハイライトの出る時間目安: 朝5時/8時の試合→昼頃 ／ 昼13時の試合→夕方 ／ 深夜2時の試合→朝。つまり21:30枠で当日分は全部回収できる").font = Font(name=FONT, size=9, italic=True)

# ============ 03_日本代表名鑑 ============
ws = wb.create_sheet("03_日本代表名鑑")
ws.sheet_properties.tabColor = "CCCCCC"
title(ws, "日本代表名鑑（ブルーロック例え付き・麻梨奈さんと見る用）", 8)
ws.cell(row=2, column=1, value="監督: 森保一 ＝ 絵心甚八（26人のエゴイストを選び抜いて束ねる人）／想定布陣 3-4-2-1／平均26.8歳").font = Font(name=FONT, size=10, bold=True)
head = ["背番", "名前", "Pos", "年齢", "クラブ(リーグ)", "クラブでの実績・状態", "今大会の期待", "🔥ブルーロックでいうと", "妻向けひとこと"]
set_rows(ws, 3, [head])
style_header(ws, 3, 9)
players = [
    [1, "鈴木彩艶", "GK", 23, "パルマ(伊・セリエA)", "セリエAで正GK定着", "守護神。ビッグセーブ連発に期待", "我牙丸吟(野性の反射神経)", "ゴールを守る最後の砦。スーパーセーブしたらこの人"],
    [12, "大迫敬介", "GK", 26, "サンフレッチェ広島", "J1屈指の守護神", "第2GK", "我牙丸の控え", "2番手キーパー"],
    [23, "早川友基", "GK", 27, "鹿島アントラーズ", "鹿島の正GK", "第3GK", "—", "3番手キーパー"],
    [2, "菅原由勢", "DF", 25, "ブレーメン(独)", "ブンデスで主力SB", "右サイドの上下動", "—(BLに守備職はいない)", "右端を走りまくる人"],
    [3, "谷口彰悟", "DF", 34, "シント=トロイデン(白)", "ベテランCB", "経験で守備を整える", "—", "落ち着きのお手本おじさん"],
    [4, "板倉滉", "DF", 29, "アヤックス(蘭)", "名門アヤックスの主力CB", "⭐新キャプテン。守備の大黒柱", "蟻生十兵衛(高さの壁)", "遠藤離脱で急きょキャプテンに。チームの兄貴"],
    [5, "長友佑都", "DF", 39, "FC東京", "39歳・5大会目の鉄人", "精神的支柱・ブラボー担当", "BLに居ない“魂”枠", "「ブラボー!」の人。39歳はW杯では超レジェンド"],
    [16, "渡辺剛", "DF", 29, "フェイエノールト(蘭)", "上田と同僚のCB", "対人の強さ", "—", "上田選手の同僚"],
    [20, "瀬古歩夢", "DF", 25, "ル・アーヴル(仏)", "リーグアンでCB", "3バックの一角候補", "—", "若手の守備職人"],
    [21, "伊藤洋輝", "DF", 27, "バイエルン(独)", "世界的名門バイエルン所属", "左足の正確なロングパス", "—", "世界一有名なクラブの一つにいる日本人"],
    [22, "冨安健洋", "DF", 27, "アヤックス(蘭)", "実力は世界級も2年ぶり代表(怪我明け)", "稼働できれば最強の切り札", "—", "ケガさえなければ日本最強DFと言われる人"],
    [25, "鈴木淳之介", "DF", 22, "コペンハーゲン(丁)", "22歳でCL経験クラブ主力", "大抜擢の新星", "新世代イレブン", "若手のびっくり枠。ブレイク候補"],
    [6, "遠藤航", "MF", 33, "リバプール(英)", "プレミア王者の一員", "❌6/12負傷離脱(無念)", "ノエル・ノア(精神的支柱)の不在", "本当のキャプテン。直前のケガで出られず…"],
    [7, "田中碧", "MF", 27, "リーズ(英)", "昇格の立役者→プレミアで主力", "中盤の頭脳", "御影玲王(万能の頭脳)", "ゲームを組み立てる頭のいい人"],
    [15, "鎌田大地", "MF", 29, "クリスタル・パレス(英)", "プレミアで攻撃の軸", "攻撃の設計者", "糸師冴(攻撃を設計する兄)", "パスで試合を支配するクールな人"],
    [24, "佐野海舟", "MF", 25, "マインツ(独)", "ブンデス屈指のボール回収力", "⭐遠藤の穴を埋める最重要人物", "BLに存在しない“回収屋”=現実では最強職", "地味だけど一番効いてる人。玄人はここを見る"],
    [8, "久保建英", "FW", 24, "レアル・ソシエダ(西)", "ラ・リーガのスター(怪我から半年ぶり復帰)", "⭐エース格。コンディションが鍵", "糸師凛(クールな完成形の天才)", "日本のスター。元バルセロナ育ちの天才"],
    [9, "後藤啓介", "FW", 20, "シント=トロイデン(白)", "20歳の大型ストライカー", "スーパーサブ→ブレイク候補", "國神錬介(パワー型・覚醒待ち)", "でかくて強い20歳。1点取ったら人生変わる"],
    [10, "堂安律", "FW", 27, "フランクフルト(独)", "ブンデスで主力アタッカー", "「俺が決める」の10番", "ミヒャエル・カイザー(ドイツ仕込みの皇帝エゴ)", "自信家。ゴール後のキメ顔に注目"],
    [11, "前田大然", "FW", 28, "セルティック(蘇)", "スコットランド王者の核", "爆速プレスで相手を壊す", "千切豹馬(チーター級スピード)", "足が速すぎる人。相手が一番嫌がるタイプ"],
    [13, "中村敬斗", "FW", 25, "スタッド・ランス(仏)", "リーグアンで得点量産歴", "左サイドの切り札", "凪誠士郎(ふわっと天才・神トラップ)", "ゆるい雰囲気で点を取る不思議な天才"],
    [14, "伊東純也", "FW", 33, "ヘンク(白)", "33歳でも衰えぬ快速", "右サイドの一刺し", "乙夜影汰(静かな仕事人)", "気づいたら裏に走ってる超速おじさん"],
    [17, "鈴木唯人", "FW", 24, "フライブルク(独)", "ブンデスで評価急上昇", "途中投入で流れを変える", "蜂楽廻(ドリブルの遊び心)", "ドリブルで遊ぶのが上手い人"],
    [18, "上田綺世", "FW", 27, "フェイエノールト(蘭)", "オランダ名門のエースFW", "⭐⭐1トップの大本命・得点源", "潔世一(ゴール前で進化し続ける主人公)", "日本のエース。点を取ったらこの人の名前が出る"],
    [19, "小川航基", "FW", 28, "NECナイメヘン(蘭)", "オランダで二桁得点級", "箱の中の仕留め役", "士道龍聖(ゴール前の野性)", "ゴール前だけ急に怖くなる人"],
    ["追", "町野修斗", "FW", "-", "（遠藤の代替で追加招集）", "6/12追加招集", "ワイルドカード", "二次選考から這い上がった男", "直前に呼ばれた補欠逆転ストーリーの人"],
]
set_rows(ws, 4, players, widths=[6, 13, 6, 6, 22, 30, 28, 30, 38])
for i in range(len(players)):
    for c in range(1, 10):
        if i % 2 == 1:
            ws.cell(row=4 + i, column=c).fill = PatternFill("solid", start_color=C_CREAM)

# ============ 04_追っかけ×マネー ============
ws = wb.create_sheet("04_追っかけ×マネー")
ws.sheet_properties.tabColor = "6FA8DC"
title(ws, "🟦 追っかけ選手×お金（オーナー視点・週1日曜21:00更新）", 8)
ws.cell(row=2, column=1, value="見方: W杯=世界最大の展示会。活躍→市場価値↑→夏の移籍金が動く。『仕入れ値と売値』で選手を見る").font = Font(name=FONT, size=10, bold=True)
head = ["選手", "所属", "💰お金の注目点", "W杯で何を見るか", "🏠不動産で例えると", "市場価値メモ(🟡週1更新)", "情報リンク", "状態"]
set_rows(ws, 3, [head])
style_header(ws, 3, 8)
follows = [
    ["上田綺世", "フェイエノールト(蘭)", "W杯得点→プレミア移籍金跳ね上げの筆頭", "ゴール数と決定率", "割安で仕入れ済→出口(売却)局面の優良物件", "", "transfermarkt.com/ayase-ueda", "🇯🇵"],
    ["鈴木彩艶", "パルマ(伊)", "23歳GK。ビッグクラブが常に探すポジション", "ビッグセーブ数", "駅近×築浅。値上がり確実エリアの新築", "", "transfermarkt.com/zion-suzuki", "🇯🇵"],
    ["佐野海舟", "マインツ(独)", "遠藤の後継=守備的MFは英で高騰職種", "ボール回収数・無失点", "地味な一棟アパート。利回りは実は最高", "", "", "🇯🇵"],
    ["久保建英", "R・ソシエダ(西)", "24歳=ビッグ移籍の最終ゴールデン世代。契約条項に注目", "怪我明けの稼働率と決定的仕事", "好立地の有名物件。指値(条項)勝負", "", "", "🇯🇵"],
    ["後藤啓介", "シント=トロイデン(白)", "20歳。1ゴールで価値が倍動く", "途中出場の結果", "再建築可の古家付き土地=化け枠", "", "", "🇯🇵"],
    ["鈴木淳之介", "コペンハーゲン(丁)", "北欧→五大リーグの王道昇格ルート上", "守備の安定感", "区画整理予定地の先行仕込み", "", "", "🇯🇵"],
    ["中村敬斗", "ランス(仏)", "得点量産歴あり。ステップアップ移籍候補", "スーパーゴール", "リフォーム済で内見映えする物件", "", "", "🇯🇵"],
    ["板倉滉", "アヤックス(蘭)", "主将ブランドで価値上昇中", "統率とビルドアップ", "管理良好な築浅レジデンス", "", "", "🇯🇵"],
    ["ジェケレシュ", "アーセナル(英)", "今季公式戦21得点。移籍金回収劇の進行形", "大舞台で通用するか最終査定", "高値仕入れでも家賃(得点)で回収中の収益物件", "", "", "🇸🇪"],
    ["イサク", "リバプール(英)", "英史上最高額級移籍→負傷続き(先発8試合)", "W杯で復活なら評価急回復", "高値掴み疑惑物件が底値反転するかの見極め", "", "", "🇸🇪"],
    ["ハンニバル・メイブリ", "バーンリー(英)", "元マンU。再ブレイクなら転売益大", "チュニジアの王様プレー", "訳あり物件の再生案件", "", "", "🇹🇳"],
    ["シャビ・シモンズ等 蘭の若手", "—", "オランダは育成→高値売却の国家モデル", "誰が「次の商品」か目利き", "デベロッパーの分譲地を見る目", "", "", "🇳🇱"],
]
set_rows(ws, 4, follows, widths=[16, 18, 36, 24, 34, 22, 28, 6])
for i in range(len(follows)):
    ws.cell(row=4 + i, column=6).fill = PatternFill("solid", start_color=C_INPUT)

# ============ 05_観戦日報 ============
ws = wb.create_sheet("05_観戦日報")
ws.sheet_properties.tabColor = "F6B26B"
title(ws, "🟧 観戦日報（毎日21:30枠で1行。W杯後も続ける趣味台帳）", 7)
head = ["日付", "試合・トピック", "結果", "気づき(俺メモ)", "💰お金の動きメモ", "温度感", "ソース"]
set_rows(ws, 3, [head])
style_header(ws, 3, 7)
example = [["2026-06-13", "(例)開幕戦ウォッチ", "—", "(例)48カ国制で3位でも突破あり=消化試合が減る設計", "(例)放映権はDAZN全試合。日本戦無料=客寄せ戦略", "🔥", "FIFA公式"]]
set_rows(ws, 4, example, widths=[12, 26, 10, 40, 40, 8, 14])
for i in range(4, 40):
    for c in range(1, 8):
        ws.cell(row=i + 1, column=c).border = BORDER
        ws.cell(row=i + 1, column=c).fill = PatternFill("solid", start_color=C_INPUT if c <= 7 else "FFFFFF")

# ============ 06_TIPS_師匠とリンク集 ============
ws = wb.create_sheet("06_TIPS_師匠とリンク集")
ws.sheet_properties.tabColor = "999999"
title(ws, "TIPS: 誰の視点を真似るか（ゴール逆算）＋リンク集＋不動産⇄サッカー対訳", 5)

ws.cell(row=3, column=1, value="◆ ゴール逆算マップ: 『お金×活躍の相関をオーナー視点で語れる人になる』").font = Font(name=FONT, bold=True, size=11, color=C_HEAD)
g = [
    ["レイヤー", "頻度", "やること", "真似る師匠", "ツール"],
    ["①速報", "毎日3分", "結果と移籍速報を拾う", "Fabrizio Romano(X)=移籍速報の世界標準「Here we go」", "FotMobアプリ＋X"],
    ["②値動き", "週1", "市場価値の上下を見る", "Transfermarkt=選手の路線価マップ", "transfermarkt.com"],
    ["③財務分析", "月1", "クラブ決算を読む", "Swiss Ramble(Substack)=クラブ決算図解の第一人者", "swissramble.substack.com"],
    ["④日本語の語り口", "随時", "移籍市場の文脈を日本語で", "小澤一郎(YouTube/Periodista)=育成×移籍市場", "YouTube"],
    ["⑤実践オーナー", "ロールモデル", "クラブ経営の一次情報", "岡田武史(FC今治オーナー)/本田圭佑(クラブ投資)", "書籍・インタビュー"],
]
set_rows(ws, 4, g, widths=[14, 12, 30, 48, 26])
style_header(ws, 4, 5)

r = 11
ws.cell(row=r, column=1, value="◆ リンク集（観戦と速報）").font = Font(name=FONT, bold=True, size=11, color=C_HEAD)
links = [
    ["用途", "名前", "URL/場所", "メモ", ""],
    ["速報アプリ", "FotMob", "アプリStoreで入手", "試合速報+選手採点+市場価値が1本で見られる", ""],
    ["ハイライト", "DAZN Japan公式YouTube", "youtube.com/@DAZNJapan", "W杯ハイライト。日本戦はDAZN本体で無料", ""],
    ["ハイライト", "FIFA公式YouTube", "youtube.com/@FIFA", "全試合ハイライト(英語)", ""],
    ["地上波", "NHK/日テレ/フジ", "TVer/NHKプラス", "日本のGS3試合はNHK2+日テレ1", ""],
    ["市場価値", "Transfermarkt", "transfermarkt.com", "選手の市場価値・契約年数・移籍履歴", ""],
    ["移籍速報", "Fabrizio Romano", "X: @FabrizioRomano", "「Here we go」=成約サイン", ""],
    ["財務", "Swiss Ramble", "swissramble.substack.com", "クラブ決算分析(英語・図解)", ""],
    ["財務", "Deloitte Football Money League", "deloitte.com (年次無料)", "クラブ収益ランキング=業界の路線価表", ""],
    ["戦術", "戸田和幸/leo the football", "YouTube", "試合の見方が深くなる", ""],
    ["記事", "footballista", "footballista.jp", "経営・戦術の読み物", ""],
]
set_rows(ws, r + 1, links, widths=[14, 30, 34, 44, 6])
style_header(ws, r + 1, 4)

r = r + len(links) + 3
ws.cell(row=r, column=1, value="◆ 不動産⇄サッカー移籍 対訳表（オーナー視点の最重要ツール）").font = Font(name=FONT, bold=True, size=11, color=C_HEAD)
trans = [
    ["サッカー用語", "不動産でいうと", "オーナー視点の見方", "", ""],
    ["市場価値(Transfermarkt)", "路線価・実勢価格", "毎週変わる。W杯は地価が一気に動く再開発イベント", "", ""],
    ["移籍金", "売買価格", "クラブの売却益=キャピタルゲイン", "", ""],
    ["契約残年数", "賃貸借の残期間", "残り1年=指値し放題(安く買える)。だから契約延長交渉が大事", "", ""],
    ["フリー移籍(契約満了)", "立退き完了後のゼロ円退去", "売主(クラブ)に1円も入らない最悪の出口", "", ""],
    ["レンタル移籍+買取OP", "賃貸+購入選択権付き契約", "使ってみて良ければ買う。リスク管理の王道", "", ""],
    ["育成→高値売却(蘭/ポルトガル型)", "ボロ戸建て再生→再販", "安く仕入れ、磨いて、高く売る。KHDと同じ商売", "", ""],
    ["連帯貢献金(移籍金の最大5%)", "仲介手数料の還流", "育てたクラブに自動で入る仕組み=ストック収益", "", ""],
    ["放映権/スポンサー収益", "家賃収入(インカム)", "移籍金(キャピタル)との両輪。古田土のMQ構造で読める", "", ""],
]
set_rows(ws, r + 1, trans, widths=[26, 26, 56, 6, 6])
style_header(ws, r + 1, 3)

out = "/Users/kikuchikenta/01_honbu_docs_automation/260613_W杯2026観戦ハブ_v1.xlsx"
wb.save(out)
print("saved:", out)

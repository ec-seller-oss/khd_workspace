# -*- coding: utf-8 -*-
# KHD 財務司令塔 再設計版 v1（2026-06-01）
# 菊池の思想＝4主格SSoT→判断を絞ったダッシュボード→予測連動 を実スプシ化。
# タブ: ①司令塔(判断1枚) ②資産負債SSoT(4主格×流動性×月次) ③資金繰り予測 ④借入SSoT ⑤使い方
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ---- スタイル ----
H  = Font(bold=True, size=12, color="FFFFFF")
HF = PatternFill("solid", fgColor="AA2E26")   # レンガ赤
SUB= Font(bold=True, size=11)
SUBF=PatternFill("solid", fgColor="F0E2DF")
HI = PatternFill("solid", fgColor="DDF3DD")    # ハイ=緑
MID= PatternFill("solid", fgColor="FFF4D6")    # 中=黄
LOW= PatternFill("solid", fgColor="ECECEC")    # 非流動=灰
YEN= '#,##0"円"'
thin=Side(style="thin", color="CCCCCC")
BD = Border(left=thin,right=thin,top=thin,bottom=thin)
def st(c, font=None, fill=None, num=None, al=None, bd=False):
    if font:c.font=font
    if fill:c.fill=fill
    if num:c.number_format=num
    if al:c.alignment=Alignment(horizontal=al,vertical="center")
    if bd:c.border=BD

# ============ ② 資産負債SSoT ============
ss = wb.active; ss.title="②資産負債SSoT"
ss["A1"]="② 資産負債 SSoT（4主格 × 流動性 × 月次）"; st(ss["A1"],H,HF);
for col in "BCDEFG": st(ss[col+"1"],fill=HF)
ss["A2"]="基準日：2026/06/01（毎月初に「6/1」列を更新→他タブは自動連動）"
hdr=["区分","主格","種類","口座/名称","流動性","6/1残高","データ元/備考"]
for i,h in enumerate(hdr):
    c=ss.cell(4,i+1,h); st(c,SUB,SUBF,al="center",bd=True)

# (区分,主格,種類,名称,流動性,値,備考)
assets=[
 ("資産","①法人","銀行","城北信金","ハイ",281012,"260601_城北信金_明細.csv"),
 ("資産","①法人","銀行","法人SBI(代表)","ハイ",139320,"6/1実読・5/19役員借入で減"),
 ("資産","①法人","銀行","朝日信金(普通)","ハイ",25000,""),
 ("資産","①法人","銀行","朝日信金(積立)","ハイ",360000,"付合い定額積立+3万/月"),
 ("資産","①法人","銀行","大東京信金","ハイ",769486,"※6/1未更新(5/1値)"),
 ("資産","①法人","銀行","法人TB(東京ベイ)","ハイ",1677563,"※6/1未更新(5/1値)"),
 ("資産","①法人","保険","経営セーフティ共済","非流動",1100000,"解約返戻・拘束"),
 ("資産","②③研太","銀行","個人事業SBI(SBI1/2613479)","ハイ",1224610,"260601_個人事業SBI_明細.csv"),
 ("資産","②③研太","銀行","みずほ","ハイ",71,""),
 ("資産","②③研太","銀行","ゆうちょ","ハイ",243582,"※6/1未更新(5/1値)"),
 ("資産","②③研太","銀行","楽天銀行RB1","ハイ",97587,"※6/1未更新(5/1値)"),
 ("資産","②③研太","銀行","楽天銀行RB2","ハイ",1674,""),
 ("資産","②③研太","銀行","東京ベイTB(研太)","ハイ",404939,"260601_東京ベイ信金TB_明細.csv"),
 ("資産","②③研太","現金","現金","ハイ",40000,""),
 ("資産","②③研太","証券","楽天証券(NISA+特定/持株/DC込)","中",11287071,"※6/1未更新・株式評価"),
 ("資産","②③研太","金","金(研太)","非流動",2000000,""),
 ("資産","②③研太","現物","現物","中",250000,""),
 ("資産","②③研太","不動産","自宅","非流動",42472875,"住宅ローン対応"),
 ("資産","②③研太","不動産","Kハウス北千住","中",13000000,""),
 ("資産","②③研太","その他","マリパパ譲渡","中",2000000,""),
 ("資産","②③研太","年金","年金(研太)","非流動",3300000,"将来資産"),
 ("資産","④麻梨奈","銀行","麻梨奈SBI2(普通)","ハイ",1117644,"※6/1未更新(5/1値)"),
 ("資産","④麻梨奈","証券","麻梨奈SBI2(証券)","中",2763600,""),
 ("資産","④麻梨奈","証券","麻梨奈 楽天2 NISA","非流動",2444704,""),
 ("資産","④麻梨奈","金","金(麻梨奈)","非流動",2000000,""),
 ("資産","④麻梨奈","年金","年金(麻梨奈)","非流動",3000000,"将来資産"),
]
liabs=[
 ("負債","①法人","借入","城北信金(東京保証)","-",3000000,"0借入一覧最新・返済予定表241115"),
 ("負債","①法人","借入","朝日信金(東京保証)","-",792000,"0借入一覧最新"),
 ("負債","①法人","借入","大東京(オリコ保証)","-",984000,"0借入一覧最新・250929確認"),
 ("負債","①法人","カード","JAL VIEWカード未払","-",1907342,""),
 ("負債","②③研太","借入","住宅ローン(自宅)","-",42355517,"2065年"),
 ("負債","②③研太","借入","公庫コロナ","-",1490000,"0借入一覧最新(残約149万)"),
 ("負債","②③研太","借入","浦安創業TB","-",4120000,"0借入一覧最新(残約412万)"),
 ("負債","②③研太","借入","奨学金","-",930700,"2040年"),
 ("負債","②③研太","カード","カード未払(MB/AGP)","-",218558,""),
 ("負債","②③研太","カード","カード未払(R1楽天)","-",180921,""),
 ("負債","②③研太","カード","PayPayカード","-",10000,""),
]
r=5; first_asset=r
for row in assets+liabs:
    for i,v in enumerate(row):
        c=ss.cell(r,i+1,v); st(c,bd=True)
        if i==5: st(c,num=YEN,al="right")
    liq=row[4]
    fill=HI if liq=="ハイ" else MID if liq=="中" else LOW if liq=="非流動" else None
    if fill: st(ss.cell(r,5),fill=fill,al="center")
    r+=1
last=r-1
asset_last=first_asset+len(assets)-1
liab_first=asset_last+1

# サマリー
r+=1
def line(lbl, formula, bold=True, fill=None):
    global r
    c1=ss.cell(r,4,lbl); st(c1,SUB if bold else None,fill)
    c2=ss.cell(r,6,formula); st(c2,SUB if bold else None,fill,num=YEN,al="right")
    r+=1
line("流動性ハイ（即現金化）", f'=SUMIFS(F{first_asset}:F{last},E{first_asset}:E{last},"ハイ")', fill=HI)
line("中（数ヶ月）", f'=SUMIFS(F{first_asset}:F{last},E{first_asset}:E{last},"中")', fill=MID)
line("非流動", f'=SUMIFS(F{first_asset}:F{last},E{first_asset}:E{last},"非流動")', fill=LOW)
line("総資産", f'=SUMIF(A{first_asset}:A{last},"資産",F{first_asset}:F{last})')
line("総負債", f'=SUMIF(A{first_asset}:A{last},"負債",F{first_asset}:F{last})')
networth_row=r
line("★純資産（自己資本）", f'=F{networth_row-2}-F{networth_row-1}', fill=SUBF)
ss.column_dimensions["D"].width=30; ss.column_dimensions["F"].width=14
ss.column_dimensions["G"].width=34
for col in "ABCE": ss.column_dimensions[col].width=10
SSREF=f"'②資産負債SSoT'"
HI_CELL=f"{SSREF}!F{networth_row-5}"   # 流動性ハイ
ASSET_CELL=f"{SSREF}!F{networth_row-2}"
LIAB_CELL=f"{SSREF}!F{networth_row-1}"
NET_CELL=f"{SSREF}!F{networth_row}"

# ============ ③ 資金繰り予測 ============
cf=wb.create_sheet("③資金繰り予測")
cf["A1"]="③ 資金繰り予測（6ヶ月・月初現金はSSoT連動）"; st(cf["A1"],H,HF)
for col in "BCDEFGHI": st(cf[col+"1"],fill=HF)
months=["2026/06","2026/07","2026/08","2026/09","2026/10","2026/11","2026/12","2027/01"]
cf.append(["項目"]+months)
for i in range(9): st(cf.cell(2,i+1),SUB,SUBF,al="center",bd=True)
cols=[get_column_letter(i+2) for i in range(8)]
def crow(label, vals):
    cf.append([label]+vals)
    rr=cf.max_row
    for i in range(9):
        c=cf.cell(rr,i+1); st(c,bd=True)
        if i>=1: st(c,num=YEN,al="right")
    return rr
# 月初: 6月=SSoT流動性ハイ、以降=前月末
r3=["月初現金（流動性ハイ・SSoT連動）", f"={HI_CELL}"]+[f"={get_column_letter(i+2)}99" for i in range(7)]
# 後で月末行番号確定して差し替えるためプレースホルダ→簡易に直接構築
cf.append(["■入金"])
栄町=crow("物件売却(栄町6月/新規)",[1900000,0,0,0,0,0,0,0])
医療=crow("医療コンサル(9月想定)",[0,0,0,660000,0,0,0,0])
ecr=crow("EC粗利(クーパン1+2)",[150000]*8)
iku=crow("麻梨奈 育休給付(2か月毎・月割)",[156659]*8)
in_first=栄町; in_last=iku
in_tot=crow("入金合計",[f"=SUM({c}{in_first}:{c}{in_last})" for c in cols])
cf.append(["■出金"])
fixed_rows=[
 ("法人:税理士(ZH橋本)",110000),("法人:外注(江藤)",99000),("法人:社保",22167),
 ("個人事業:税理士",22000),("住宅ローン",130668),("SMBC管理費",22130),("PayPay",10000),
 ("★世帯生活費(楽天18+麻梨奈12.1)",301000),
 ("★法人借入返済(大東京+朝日+城北)",76035),("★個人事業借入返済(TB2+公庫)",145001),
 ("投信積立(任意)",100000),("役員賞与(見送り確定=0)",0),
]
out_first=None
for lbl,val in fixed_rows:
    rr=crow(lbl,[val]*8)
    if out_first is None: out_first=rr
    out_last=rr
out_tot=crow("出金合計",[f"=SUM({c}{out_first}:{c}{out_last})" for c in cols])
net=crow("当月純増減",[f"={c}{in_tot}-{c}{out_tot}" for c in cols])
# 月初行を一番上に置きたいが簡易に：月初行をここで定義し直し（行3固定にできないので末尾にサマリ）
end=crow("月末現金",[f"={HI_CELL}+{cols[0]}{net}" if i==0 else f"={cols[i-1]}9999" for i in range(8)])
# 月末の連鎖を正しく：end行を作り直し
er=end
for i in range(8):
    if i==0:
        cf.cell(er,2,f"={HI_CELL}+B{net}")
    else:
        cf.cell(er,i+2,f"={get_column_letter(i+1)}{er}+{cols[i]}{net}")
    st(cf.cell(er,i+2),SUB,num=YEN,al="right")
cf.cell(er,1,"月末現金（体力計）"); st(cf.cell(er,1),SUB)
# ランウェイ・信号
cf.append([])
rwr=crow("純月次燃焼(資金分岐−経常入金)",[f"={c}{out_tot}-{c}{ecr}-{c}{iku}-100000" for c in cols])  # 積立除く近似
sig=cf.max_row+1
cf.cell(sig,1,"ランウェイ(月初現金÷燃焼)"); st(cf.cell(sig,1),SUB)
cf.cell(sig,2,f"={HI_CELL}/B{rwr}"); st(cf.cell(sig,2),SUB,num='0.0"ヶ月"')
cf.cell(sig,3,f'=IF(B{sig}<3,"🔴守り:投資凍結",IF(B{sig}<6,"🟡注意","🟢攻めOK"))'); st(cf.cell(sig,3),SUB)
cf.column_dimensions["A"].width=30
for c in cols: cf.column_dimensions[c].width=12

# ============ ① 司令塔（判断1枚）============
db=wb.create_sheet("①司令塔",0)
db["A1"]="① 財務司令塔 — 毎月初これだけ見る"; st(db["A1"],Font(bold=True,size=14,color="FFFFFF"),HF)
for col in "BCDE": st(db[col+"1"],fill=HF)
db["A2"]="基準日 2026/06/01"
def kpi(rr,label,formula,num=YEN,note=""):
    db.cell(rr,1,label); st(db.cell(rr,1),SUB,SUBF)
    c=db.cell(rr,2,formula); st(c,Font(bold=True,size=12),num=num,al="right")
    if note: db.cell(rr,3,note)
db["A4"]="■ 今の現在地（②SSoT連動）"; st(db["A4"],SUB)
kpi(5,"純資産（自己資本）",f"={NET_CELL}",note="世間中央値超だが流動性は薄い")
kpi(6,"総資産",f"={ASSET_CELL}")
kpi(7,"総負債",f"={LIAB_CELL}",note="法人は債務超過に注意")
kpi(8,"現預金（流動性ハイ）",f"={HI_CELL}",note="即現金化できる額")
db["A10"]="■ ランウェイ信号（③予測連動）"; st(db["A10"],SUB)
db.cell(11,1,"ランウェイ"); st(db.cell(11,1),SUB,SUBF)
db.cell(11,2,f"='③資金繰り予測'!B{sig}"); st(db.cell(11,2),Font(bold=True,size=12),num='0.0"ヶ月"')
db.cell(11,3,f"='③資金繰り予測'!C{sig}"); st(db.cell(11,3),Font(bold=True,size=12))
db.cell(12,1,"通期末(2027/1)現金"); st(db.cell(12,1),SUB,SUBF)
db.cell(12,2,f"='③資金繰り予測'!I{er}"); st(db.cell(12,2),num=YEN,al="right")
db["A14"]="■ 今月の判断ポイント（絞る）"; st(db["A14"],SUB)
db["A15"]="□ 法人現預金の内、運転資金(EC仕入で消える)と自由現金の切り分け＝真のランウェイ確定の鍵"
db["A16"]="□ 借入残高はタブ間で不一致だった→④借入SSoTに一本化(本シートで採用)"
db["A17"]="□ ランウェイ🟢6ヶ月以上で初めて『資産拡大の投資』に現金を回す"
db["A19"]="■ 投資判断ハードルレート"; st(db["A19"],SUB)
db["A20"]="機会コスト＝投信/NISA期待 年5〜7%。これを下回る事業に現金を張らない。"
db["A21"]="不動産＝CF率1.5%/CCR15%/IRR8%/粗利20%/土地値0.4。事業＝事業計画で同等以上+回収期間明示。"
db.column_dimensions["A"].width=40; db.column_dimensions["B"].width=16; db.column_dimensions["C"].width=34

# ============ ④ 借入SSoT ============
ln=wb.create_sheet("④借入SSoT")
ln["A1"]="④ 借入 SSoT（残高の単一真実源・返済予定表リンク）"; st(ln["A1"],H,HF)
for col in "BCDEFG": st(ln[col+"1"],fill=HF)
lh=["主格","借入先","残高","金利","月返済","返済予定表/明細リンク","備考"]
ln.append([""]);
for i,h in enumerate(lh): c=ln.cell(2,i+1,h); st(c,SUB,SUBF,al="center",bd=True)
loans=[
 ("①法人","城北信金(東京保証)",3000000,"1.8%固定",21253,"※Driveリンク要","R7.8.25より元金均等"),
 ("①法人","朝日信金(東京保証)",792000,"2.575%変動",17732,"※Driveリンク要","年1回11月改定"),
 ("①法人","大東京(オリコ保証)",984000,"2.8%固定",18340,"※Driveリンク要","250929確認"),
 ("②③研太","公庫コロナ",1490000,"1.21%固定",22531,"※Driveリンク要","利子補給終了"),
 ("②③研太","浦安創業TB",4120000,"2.1%変動",91348,"※Driveリンク要","R8.2.2金利変更"),
 ("②③研太","住宅ローン(自宅)",42355517,"要確認",130668,"※Driveリンク要","2065年"),
 ("②③研太","奨学金",930700,"-","-","","2040年"),
]
for row in loans:
    ln.append(list(row))
    rr=ln.max_row
    for i in range(7):
        c=ln.cell(rr,i+1); st(c,bd=True)
        if i==2: st(c,num=YEN,al="right")
ln.append([])
tr=ln.max_row+1
ln.cell(tr,2,"借入合計"); st(ln.cell(tr,2),SUB)
ln.cell(tr,3,f"=SUM(C3:C{tr-2})"); st(ln.cell(tr,3),SUB,num=YEN,al="right")
ln.column_dimensions["B"].width=22; ln.column_dimensions["C"].width=14
ln.column_dimensions["F"].width=20; ln.column_dimensions["G"].width=20

# ============ ⑤ 使い方 ============
hp=wb.create_sheet("⑤使い方・データ元")
hp["A1"]="⑤ 使い方・データ元リンク"; st(hp["A1"],H,HF)
notes=[
 "【思想】②資産負債SSoTが単一入力源。毎月初に『6/1』列(各口座残高)を更新→①司令塔・③予測が自動連動。",
 "【月初ルーティン】各行ログインDL→~/Downloads→Claudeが khd_sort_meisai.py で各フォルダ振分→残高を②に入力。",
 "【収集制約】SBIはClaude代行可。トークン式(ゆうちょ/楽天/信金)は菊池がDL→Downloads→Claude仕分け。",
 "【※6/1未更新】の口座は5/1値のまま。ログインDLでき次第、実値に更新する。",
 "【ポイント/マイル/チャージは資産計上しない方針】(換金性低・旧収支タブの水増しを排除)。",
 "【借入】④借入SSoTを唯一の真実源に(旧:収支/借入残高/0借入一覧で不一致だった)。返済予定表Driveリンクを各行に貼る。",
 "【元データ】明細CSVは主格別フォルダ(02_KHD/法人SBI・城北 等／01_個人/SBI・TB／00_プライベート/ゆうちょ・RB1・RS1)。",
]
for i,n in enumerate(notes):
    hp.cell(i+3,1,n)
hp.column_dimensions["A"].width=110

out="/Users/kikuchikenta/Library/CloudStorage/GoogleDrive-ec-seller@kikuchi-hd.net/マイドライブ/KHD_財務司令塔_v1_提案.xlsx"
wb.save(out)
print("保存完了:",out)
print("タブ:", wb.sheetnames)
